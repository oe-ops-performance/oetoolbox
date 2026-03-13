import calendar
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import pandas as pd
from pathlib import Path
from sqlalchemy import create_engine
from types import SimpleNamespace

from ..utils import oepaths
from ..utils.datetime import remove_tzinfo_and_standardize_index
from ..utils.helpers import with_retries, quiet_print_function

## function to load most recently-created file in fpath list
get_file = lambda fp_list: max((fp.stat().st_ctime, fp) for fp in fp_list)[1] if fp_list else None


REPORT_TEMPLATE_PATH = Path(
    oepaths.solar, "Data", "Comanche", "Templates", "Curtailment_Report_TEMPLATE.xlsx"
)
FMT = SimpleNamespace(
    datetime="yyyy-mm-dd hh:mm;@",
    date="yyyy-mm-dd;@",
    z4="0.0000",
    z2="0.00",
    rev="#,##0.00",
    int="0",
)


class ComancheCurtailment:
    def __init__(self, year: int, month: int, scaling_factor: float):
        self.year = year
        self.month = month
        self.flashreport_path = oepaths.frpath(year, month, ext="solar", site="Comanche")
        self.scaling_factor = scaling_factor
        self.data = {  # everything is minute-level
            "poa": None,  # dataframe with poa from pvlib file (sourced from PI or DTN)
            "pi": None,  # dataframe with pvlib, pi meter, and ppc setpoint
            "sql": None,  # dataframe with expected, sql meter, and curtailment
        }
        self.report_data = {
            "pi": {"interval": None, "summary": None},
            "sql": {"interval": None, "summary": None},
        }
        self.source_files = []
        self._load_data_from_files()
        self._query_sql_data()

    def _load_data_from_files(self):
        """Loads data from PI query and PVLib files in FlashReport folder.
        -> self.data["poa"] - columns: ["POA" or "POA_DTN"]
        -> self.data["pi"] - columns: ["possible", "meter", "xcel_MW_setpoint", "curtailment"]
        -> updates self.source_files
        """

        def getloadfile(pattern) -> pd.DataFrame:
            fpath = oepaths.latest_file(list(self.flashreport_path.glob(pattern)))
            if fpath is None:
                raise ValueError(f"Missing required file: {pattern}")
            self.source_files.append(fpath.name)
            df = pd.read_csv(fpath, index_col=0, parse_dates=True)
            if df.index.tzinfo is not None:
                df = remove_tzinfo_and_standardize_index(df)
            return df

        # load pvlib file and get poa & site-level possible generation
        df_pvl = getloadfile("PVLib_InvAC*.csv")
        poacol = "POA" if "POA" in df_pvl.columns else "POA_DTN"
        if poacol not in df_pvl.columns:
            raise ValueError("Could not find POA column in pvlib dataset.")
        if df_pvl.index.diff().max() > pd.Timedelta(minutes=1):
            df_pvl = df_pvl.resample("1min").ffill()
            df_pvl = df_pvl.reindex(df_mtr.index)
            print("Warning: possible power native freq is 1-hour (uses DTN POA).")

        # get poa data
        self.data["poa"] = df_pvl[[poacol]].copy()

        # calculate pvlib inverter sum
        pvlib_inv_sum = df_pvl.filter(like="Possible_Power").sum(axis=1).div(1e3)  # kW to MW
        pvlib_inv_sum = pvlib_inv_sum.clip(upper=120.0)  # limit to site capacity
        if self.scaling_factor != 1:
            pvlib_inv_sum = pvlib_inv_sum.mul(self.scaling_factor)

        # create df for data sourced from pi
        df = pd.Series(pvlib_inv_sum, name="possible").to_frame()

        # get pi meter data
        df_mtr = getloadfile("PIQuery_Meter_*.csv")
        df["meter"] = df_mtr.iloc[:, 0].copy()

        # get ppc data
        sp_col = "xcel_MW_setpoint"
        df_ppc = getloadfile("PIQuery_PPC_*.csv")
        if sp_col not in df_ppc.columns:
            raise ValueError(f"Column '{sp_col}' not found in PPC dataset.")
        df[sp_col] = df_ppc[sp_col].copy()

        # create column for curtailed kW
        is_curtailed = (
            df[sp_col].lt(120.0) & df[sp_col].lt(df["possible"]) & df["possible"].gt(df["meter"])
        )
        # NOTE: this is in kW; converted to kWh in subsequent steps
        df["curtailment"] = np.where(  # poss/meter are in MW
            is_curtailed, df["possible"].sub(df["meter"]).mul(1e3), 0.0
        )

        # assign pi data - columns: ["possible", "meter", "xcel_MW_setpoint", "curtailment"]
        self.data["pi"] = df.copy()

    def _query_sql_data(self):
        """Queries interval data from SQL.
        -> self.data["sql"] - columns: ["possible", "meter", "xcel_MW_setpoint", "curtailment"]
        """
        df = comanche_reporting_timeseries(self.year, self.month)
        df = df.rename(
            columns={
                "Expected_KW": "possible",
                "Meter_KW": "meter",
                "Power_SP_KW": "xcel_MW_setpoint",
                "Curtailed_KW": "curtailment",
            }
        )
        mw_cols = ["possible", "meter", "xcel_MW_setpoint"]
        df[mw_cols] = df[mw_cols].div(1e3)  # convert kW to MW
        self.data["sql"] = df.copy()

    def _generate_report_interval_data(self, source):
        # combine: [poa, possible, meter, xcel_MW_setpoint, curtailment]
        df = self.data[source].copy()
        df.insert(0, "poa", self.data["poa"])
        df["curt_rate"] = 0.0625  # units: $/kWh

        # convert curtailment data from power (kW) to energy (kWh)
        df["curtailment"] = df["curtailment"].div(60)  # minute-level data

        df["lost_revenue"] = df["curtailment"].mul(df["curt_rate"])
        df["n_curtailments"] = np.where(df["curtailment"].gt(0), 1, 0)  # for summary table
        self.report_data[source]["interval"] = df.copy()

    def _generate_daily_summary_table(self, source):
        # get report interval data
        df = self.report_data[source]["interval"].copy()
        summary_cols = ["curtailment", "n_curtailments", "lost_revenue"]
        df = df[summary_cols].rename(columns={"n_curtailments": "curt_hours"})
        df = df.groupby(df.index.day).sum()
        df["curt_hours"] = df["curt_hours"].div(60)
        df.index = pd.to_datetime(
            f"{self.year}-{self.month:02d}-" + df.index.astype(str).str.zfill(2),
            format="%Y-%m-%d",
        )
        self.report_data[source]["summary"] = df.copy()

    def _generate_data_for_report(self):
        for source in ("pi", "sql"):
            self._generate_report_interval_data(source=source)
            self._generate_daily_summary_table(source=source)

    @classmethod
    def load(cls, year: int, month: int, scaling_factor: float = 1.0):
        curt = cls(
            year=year,
            month=month,
            scaling_factor=scaling_factor,
        )
        curt._generate_data_for_report()
        return curt

    def compare_totals(self):
        print(f"Total curtailment losses:")
        for source, dict_ in self.report_data.items():
            lost_mw = dict_["interval"]["curtailment"].sum() / 1e3  # curtailment is in kW
            print(f"   {source.rjust(3)} -> {lost_mw:.2f}")
        return

    def generate_report(self, source, q=True, local=False):
        qprint = quiet_print_function(q=q)
        df = self.report_data[source]["interval"].copy()
        dfs = self.report_data[source]["summary"].copy()

        # resample to 5-minute interval data, reset index for dataframe_to_rows function
        df = df.resample("5min").mean().copy()
        df = df.reset_index(drop=False)
        dfs = dfs.reset_index(drop=False)

        # open Excel template workbook
        qprint("Generating report... ", end="")
        wb = openpyxl.load_workbook(REPORT_TEMPLATE_PATH)
        ws = wb["Curtailment_ONW"]

        # write main data table using interval dataframe
        # cols. A-I:  [Timestamp, poa,pvlib,meter, xcel_sp, curtailment, rate, lost_rev, n_curt]
        rng1_formats = [FMT.datetime] + [FMT.z4] * 3 + [FMT.z2, FMT.z4, FMT.z4, FMT.rev, FMT.int]
        for r, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
            for c, val in enumerate(row):
                ws.cell(r + 2, c + 1).value = val
                ws.cell(r + 2, c + 1).number_format = rng1_formats[c]

        # SUMMARY TABLE (from dfs)
        # cols. L-O: [Timestamp, curtailment, curt_hrs, lost_rev]
        rng2_formats = [FMT.date, FMT.z2, FMT.z2, FMT.rev]
        rng2_offset = 12  # starts at col idx 12
        for r, row in enumerate(dataframe_to_rows(dfs, index=False, header=False)):
            for c, val in enumerate(row):
                cell = ws.cell(r + 2, c + rng2_offset)
                cell.value = val
                cell.number_format = rng2_formats[c]

        totals_row = 33  # hard-coded (to accommodate months with diff # of days)
        totals_cols = ["curtailment", "curt_hours", "lost_revenue"]
        totals_formats = [FMT.rev, FMT.z2, FMT.rev]
        for c, col in enumerate(totals_cols):
            col_idx = c + rng2_offset + 1
            cell = ws.cell(totals_row, col_idx)
            cell.value = dfs[col].sum()
            cell.number_format = totals_formats[c]
            if col_idx == 13:
                ws.cell(totals_row + 1, col_idx).value = dfs[col].sum() / 1000
                ws.cell(totals_row + 1, col_idx).number_format = totals_formats[c]

        # remove columns previously used for including sql data - TODO: update template
        ws.delete_cols(idx=17, amount=7)  # cols Q through W

        if local is True:
            savedir = Path.home().joinpath("Downloads")
        else:
            savedir = self.flashreport_path
        fname_ext = f"{calendar.month_abbr[self.month]}-{self.year}"
        filename = f"Comanche_Curtailment_Report_{fname_ext}.xlsx"
        savepath = oepaths.validated_savepath(Path(savedir, filename))

        wb.save(savepath)
        wb.close()
        qprint(f'done!\n    >> saved file: "{filename}"')
        return


@with_retries(n_max=3)
def comanche_reporting_timeseries(year, month, q=True) -> pd.DataFrame:
    """Queries interval data from Comanche SQL database.

    Returns
    -------
    pd.DataFrame
        A dataframe with columns: Expected_KWh, Meter_KWh, Power_SP, Curtailed_KWh
    """
    qprint = quiet_print_function(q=q)

    start = pd.Timestamp(year, month, 1)
    end = start + pd.DateOffset(months=1)
    start_date, end_date = map(lambda ts: ts.strftime("%Y-%m-%d"), [start, end])

    # create engine for connection to SQL server, then use pandas to get query results
    engine = create_engine(
        "mssql+pyodbc://CORP-OPSSQL/Comanche?driver=ODBC+Driver+17+for+SQL+Server"
    )
    with engine.connect() as conn, conn.begin():
        df = pd.read_sql_query(
            """
            DECLARE @startDate date;
            DECLARE @endDate date;

            SET @startDate = ?;
            SET @endDate = ?;

            SELECT
                Timestamp_UTC
                ,Park_Potential_KW AS Expected_KW
                ,Meter_KW AS Meter_KW
                ,Power_Limit_SP AS Power_SP_KW
                ,CASE
                    WHEN (Park_Potential_KW > Meter_KW)
                        AND (Power_Limit_SP < 120000)
                        AND (Park_Potential_KW > Power_Limit_SP)
                    THEN Park_Potential_KW - Meter_KW
                    ELSE 0
                END AS Curtailed_KW
            FROM [Comanche].dbo.POIData
            WHERE DATEADD(HOUR, -6, Timestamp_UTC) BETWEEN @startDate AND @endDate
            ORDER BY Timestamp_UTC;
            """,
            con=conn,
            params=(start_date, end_date),
        )

    df["Timestamp"] = df["Timestamp_UTC"] - pd.Timedelta(hours=6)
    df = df.set_index("Timestamp").drop(columns=["Timestamp_UTC"])

    # format/validate index
    idx_kwargs = dict(start=start_date, end=end_date, freq="1min")
    expected_sql_index = pd.date_range(**idx_kwargs, inclusive="both")
    n_diff = len(df.index) - len(expected_sql_index)
    if n_diff != 0:
        if month == 3 and n_diff < 0:
            # missing hour (60 missing timestamps)
            pass  # reindexed below
        elif month == 11 and n_diff > 0:
            # extra hour (60 extra/duplicate timestamps)
            df = df[~df.index.duplicated(keep="first")]
        else:
            raise ValueError("Unexpected timestamp mismatch in SQL output.")
        df = df.reindex(expected_sql_index)
        qprint("Reindexed DST condition")

    expected_index = pd.date_range(**idx_kwargs, inclusive="left")
    df = df.reindex(expected_index)
    return df


def get_comanche_curtailment(year, month, return_fpath: bool = False):
    """Loads total curtailment from Comanche curtailment report file (if exists)"""
    site_fpath = oepaths.frpath(year, month, ext="solar", site="Comanche")
    valid_files = list(site_fpath.glob("*Curtailment*Report*"))
    if not valid_files:
        raise ValueError("No curtailment report file found.")
    fpath = oepaths.latest_file(valid_files)
    df = pd.read_excel(fpath, engine="calamine")
    ref_col = "DATE" if "DATE" in df.columns else "Sum"
    if ref_col not in df.columns:
        raise KeyError("Could not find curtailment column in file. Check format.")
    target_col_index = df.columns.get_loc(ref_col) + 1
    target_col = df.columns[target_col_index]
    target_row = df[target_col].last_valid_index()
    curt_total = df.at[target_row, target_col]
    if return_fpath:
        return {"total_mwh": curt_total, "source": fpath}
    return curt_total


def commercial_folder():
    for fpath in oepaths.onward_fp.glob("*Commercial*Documents*"):
        if Path(fpath, "Commercial").exists():
            return Path(fpath, "Commercial")
    return


def get_commercial_folder(site: str, fleet: str = "solar"):
    commercial_dir = commercial_folder()
    if commercial_dir is None:
        raise ValueError("Commercial directory not found in expected location.")
    if fleet.lower() not in ("solar", "thermal", "wind"):
        raise KeyError("Invalid fleet specified.")
    folder_name = site  # temp; might need mapping for certain sites
    site_folder = Path(commercial_dir, fleet.capitalize(), folder_name)
    if not site_folder.exists():
        raise KeyError("Unsupported site; temp.")
    return site_folder


def _curtailment_from_settlement_invoice(filepath):
    sheet_name = "Curtailment Input" if "Maplewood 1" in str(filepath) else "Curtailment INPUT"
    df = pd.read_excel(filepath, engine="calamine", sheet_name=sheet_name)

    checkvals_hdr = df.iloc[:10, 0].to_list()
    if "Element" not in checkvals_hdr:
        raise KeyError("Unexpected format in file.")
    header_row = checkvals_hdr.index("Element")
    table_end_col_name = df.iloc[header_row].last_valid_index()
    table_end_col = df.columns.get_loc(table_end_col_name)

    if "Daily View" in df.columns:
        table_start_col = df.columns.get_loc("Daily View")
    else:
        checkvals_tbl = df.loc[0].to_list()
        if "Daily View" not in checkvals_tbl:
            raise KeyError("Unexpected format in file.")
        table_start_col = checkvals_tbl.index("Daily View")
    end_row = df.iloc[:, table_start_col].last_valid_index()

    table_end_col_name = df.iloc[header_row].last_valid_index()
    table_end_col = df.columns.get_loc(table_end_col_name)

    df_tbl = df.iloc[header_row : end_row + 1, table_start_col : table_end_col + 1].copy()
    df_tbl.columns = df_tbl.iloc[0].values
    df_tbl = df_tbl.iloc[1:, :].reset_index(drop=True)

    date_col, curt_col = df_tbl.columns
    df_tbl[date_col] = pd.to_datetime(df_tbl[date_col])
    df_tbl[curt_col] = pd.to_numeric(df_tbl[curt_col])
    return df_tbl


def get_maplewood_curtailment(site: str, year: int, month: int) -> dict:
    """NOTE: this will only work if the Commercial SharePoint folder is synced to the user's "Onward Energy" folder"""
    if site not in ("Maplewood 1", "Maplewood 2"):
        raise ValueError("Invalid site specified.")
    folder = get_commercial_folder(site=site)
    settlements_folder = Path(folder, "Settlements")
    if not settlements_folder.exists():
        raise ValueError("Could not find Settlements folder.")
    fpattern = f"*{site}*Energy*Invoice*{year}*{month:02d}*.xlsx"
    invoice_fpaths = list(Path(folder, "Settlements").glob(fpattern))
    if not invoice_fpaths:
        raise ValueError("No settlement files found for specified period.")

    # use final if exist (MW1), otherwise preliminary
    final_fpaths = [fp for fp in invoice_fpaths if "final" in fp.name.lower()]
    if final_fpaths:
        invoice_fpaths = final_fpaths
    invoice_fp = oepaths.latest_file(invoice_fpaths)
    df = _curtailment_from_settlement_invoice(invoice_fp)
    total_mwh = df.iloc[:, -1].sum()
    return {"data": df, "total_mwh": total_mwh, "source": invoice_fp}


def load_external_curtailment_totals(site: str, year: int, month: int) -> dict:
    # for flashreport function
    if site == "Comanche":
        try:
            output = get_comanche_curtailment(year, month, return_fpath=True)
        except:
            output = {}
    elif site in ("Maplewood 1", "Maplewood 2"):
        try:
            output = get_maplewood_curtailment(site, year, month)
        except:
            output = {}
    else:
        return {}

    if not all(k in output.keys() for k in ["total_mwh", "source"]):
        raise ValueError("Unexpected output format.")

    if "data" in output.keys():
        del output["data"]

    return output
