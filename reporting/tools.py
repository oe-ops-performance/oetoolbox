import time
import calendar
import openpyxl
from openpyxl.styles import PatternFill
import numpy as np
import pandas as pd
import itertools
import tempfile, shutil, pythoncom
import xlwings as xw
from pathlib import Path
from ..utils import oepaths, oemeta
from ..dataquery import pireference as ref
from ..dataquery.external import query_DTN


ALL_SOLAR_SITES = list(oemeta.data["AF_Solar_V3"].keys())
ALL_WIND_SITES = list(oemeta.data["AF_Wind_V3"].keys())
ALL_GAS_SITES = list(oemeta.data["AF_Gas_V3"].keys())

is_solarfp = lambda fp: (fp.name in ALL_SOLAR_SITES)
frfolders = lambda y, m: list(oepaths.frpath(y, m, ext="Solar").glob("*"))

CAISO_SITES = [
    "Adams East",
    "Alamo",
    "Camelot",
    "Catalina II",
    "CID",
    "Columbia II",
    "CW-Corcoran",
    "CW-Goose Lake",
    "Kansas",
    "Kent South",
    "Maricopa West",
    "Old River One",
    "West Antelope",
]


def solar_frpaths(year, month, all_fps=False):
    fpaths = [fp for fp in frfolders(year, month) if is_solarfp(fp)]
    site_fp_dict = {fp.name: fp for fp in fpaths}

    sorted_fpaths = []
    for site in ALL_SOLAR_SITES:
        sfp = site_fp_dict.get(site)
        if sfp:
            sorted_fpaths.append(sfp)

    if not all_fps:
        isvalid_ = lambda fp: all(
            k in oemeta.data["AF_Solar_V3"].get(fp.name) for k in ["Inverters", "Met Stations"]
        )
        sorted_fpaths = [
            fp for fp in sorted_fpaths if isvalid_(fp) or fp.name in ["CW-Marin", "FL1"]
        ]
    return sorted_fpaths


site_frpath = lambda site, y, m: Path(oepaths.frpath(y, m, ext="Solar"), site)

latest_file = lambda fplist: (
    None if (not fplist) else max((fp.stat().st_ctime, fp) for fp in fplist)[1]
)
earliest_file = lambda fplist: (
    None if (not fplist) else min((fp.stat().st_ctime, fp) for fp in fplist)[1]
)


def site_fr_fpaths(site, year, month, types_=None, return_dict=False):
    supported_types = ["Meter", "Inverters", "MetStations", "PVLib"]
    selected_ = (
        supported_types
        if (types_ is None)
        else [t.replace(" ", "") for t in types_ if t.replace(" ", "") in supported_types]
    )
    if not selected_:
        print('error: check "types_" arg.\nexiting..')
        return []
    frpath_ = site_frpath(site, year, month)
    getfpaths = lambda glob_str: list(frpath_.glob(glob_str))
    sorted_ = lambda fplist: (
        [] if not fplist else list(sorted(fplist, key=lambda fp: fp.stat().st_ctime, reverse=True))
    )
    fp_dict = {
        "MetStations": sorted_(getfpaths(f"PIQuery_MetStations*.csv")),
        "Inverters": sorted_(getfpaths(f"PIQuery_Inverters*.csv")),
        "Meter": sorted_(getfpaths(f"PIQuery_Meter*.csv")),
        "PVLib": sorted_(getfpaths(f"PVLib_InvAC*.csv")),
    }
    if return_dict:
        return fp_dict

    fr_fpaths = []
    for type_, fplist_ in fp_dict.items():
        if type_ not in selected_:
            continue
        fr_fpaths.extend(fplist_)
    return list(sorted(fr_fpaths))


def solar_fr_fpath_dict(year, month, allfiles=False):
    fplist_ = lambda site_fp: [fp for fp in list(site_fp.glob("*")) if fp.is_file()]
    filekeys_ = ["PIQuery_", "PVLib_InvAC", "FlashReport"]
    is_relevant = lambda fp: (fp.suffix in [".csv", ".xlsx"]) and any(
        k in fp.name for k in filekeys_
    )
    filtered_fplist = lambda sfp_: [fp for fp in fplist_(sfp_) if is_relevant(fp)]
    return {
        sfp.name: fplist_(sfp) if allfiles else filtered_fplist(sfp)
        for sfp in solar_frpaths(year, month, all_fps=True)
    }


def printsites(sitelist, show_total=True, n_=8):
    if show_total:
        print(f"Total number of sites = {len(sitelist)}")
    print("[")
    w, s0, s1 = [len(sitelist), 0, n_]
    while w > 0:
        print((" " * 4) + f"{str(sitelist[s0:s1])[1:-1]},")
        w -= n_
        s0 = s1
        s1 += n_
    print("]")
    return


def get_solar_fr_filepaths(site, year, month):
    site_fp = Path(oepaths.frpath(year, month, ext="Solar"), site)
    fplist_ = lambda glob_str: list(site_fp.glob(glob_str))
    get_fpath = lambda g: latest_file(fplist_(g)) if fplist_(g) else None

    isprocessed = lambda fp: ("PROCESSED" in fp.name)
    isclean = lambda fp: ("CLEANED" in fp.name) and (not isprocessed(fp))
    israw = lambda fp: (not isclean(fp)) and (not isprocessed(fp))

    inv_fps = fplist_("PIQuery_Inverters*.csv")
    ifp0 = latest_file([fp for fp in inv_fps if israw(fp)])  # 0: RAW
    ifp1 = latest_file([fp for fp in inv_fps if isclean(fp)])  # 1: CLEANED

    met_fps = fplist_("PIQuery_MetStations*.csv")
    mfp0 = latest_file([fp for fp in met_fps if israw(fp)])  # 0: RAW
    mfp1 = latest_file([fp for fp in met_fps if isclean(fp)])  # 1: CLEANED
    mfp2 = latest_file([fp for fp in met_fps if isprocessed(fp)])  # 2: PROCESSED

    pfp = latest_file(fplist_(f"PVLib_InvAC_{site}*.csv"))
    rfp = latest_file(fplist_("*FlashReport*.xlsx"))

    placeholder = lambda type_: f"({type_} PI query file)"
    if ifp0 is None:
        ifp0 = placeholder("raw")
    if ifp1 is None:
        ifp1 = placeholder("clean")
    if mfp0 is None:
        mfp0 = placeholder("raw")
    if mfp1 is None:
        mfp1 = placeholder("clean")
    if mfp2 is None:
        mfp2 = placeholder("processed")
    if pfp is None:
        pfp = "(pvlib data file)"
    if rfp is None:
        rfp = "(flashreport file)"

    fp_dict = {
        "inv0": ifp0,
        "inv1": ifp1,
        "met0": mfp0,
        "met1": mfp1,
        "met2": mfp2,
        "pvl": pfp,
        "rpt": rfp,
    }

    return fp_dict


def get_solar_FR_status(year, month, sitelist=None, showfiles=False, output=False, all_fps=False):
    """note: prints summary & optionally returns dictionary of sites organized by status"""
    oprint = lambda msg, end="\n": print(msg, end=end) if (not output) else None
    # get solar flashreport folders/paths
    site_fpaths = solar_frpaths(year, month, all_fps=all_fps)

    completed = []
    completed_with_meter = []
    completed_no_meter = []
    ready_for_FR = []
    ready_for_PVLib = []
    ready_for_metbackfill = []
    needs_metQC = []
    needs_invQC = []

    for site_fp in site_fpaths:
        site = site_fp.name
        get_files = lambda glob_str: list(site_fp.glob(glob_str))

        # inverter file (cleaned)
        inv_fps = get_files("PIQuery*Inverters*CLEANED*.csv")
        invfp = latest_file(inv_fps)

        # met station files (cleaned & processed)
        isprocessed = lambda fp: ("PROCESSED" in fp.name)
        met_fps = get_files("PIQuery*MetStations*CLEANED*.csv")
        metfp1 = latest_file([fp for fp in met_fps if not isprocessed(fp)])  # 1: CLEANED
        metfp2 = latest_file([fp for fp in met_fps if isprocessed(fp)])  # 2: PROCESSED

        # pvlib file (minute-level)
        pvl_fps = get_files("PVLib*.csv")
        pvlfp = latest_file([fp for fp in pvl_fps if ("Hourly" not in fp.name)])

        # flashreport file
        fr_fps = get_files("*FlashReport*.xlsx")
        frfp = latest_file(fr_fps)

        exists_ = lambda fp: (fp is not None)

        if exists_(frfp):
            completed.append(site)
            msg_ = "(COMPLETED)"
            if "Rev0" in frfp.name:
                msg_ = "**note: waiting on utility meter - substituted with PI data **"
                if site not in ["FL1", "FL4", "Indy I", "Indy II", "Indy III"]:
                    completed_no_meter.append(site)  # for output dict
            else:
                completed_with_meter.append(site)  # for output dict
            next_step = f"NOTHING {msg_}"
        elif exists_(pvlfp) and exists_(invfp):
            ready_for_FR.append(site)
            next_step = "GENERATE FLASHREPORT"
        elif exists_(metfp2) and not exists_(pvlfp):
            ready_for_PVLib.append(site)
            next_step = "RUN PVLIB MODEL"
        elif exists_(metfp1) and not exists_(metfp2):
            ready_for_metbackfill.append(site)
            next_step = "RUN METEO BACKFILL SCRIPT"
        else:
            next_step = "MANUAL QC OF PI QUERY FILES"

        if not exists_(metfp1):
            needs_metQC.append(site)
        if (not exists_(invfp)) and (site != "CW-Marin"):
            needs_invQC.append(site)

        # full printouts
        if showfiles:
            oprint(f"\n{site.upper().ljust(23)}   ---   <||| NEXT STEP: {next_step} |||>")
            oprint(f"     cleaned inv file:  " + (invfp.name if exists_(invfp) else ""))
            oprint(f"     cleaned met file:  " + (metfp1.name if exists_(metfp1) else ""))
            oprint(f"   processed met file:  " + (metfp2.name if exists_(metfp2) else ""))
            oprint(f"           pvlib file:  " + (pvlfp.name if exists_(pvlfp) else ""))
            oprint(f"     flashreport file:  " + (frfp.name if exists_(frfp) else ""))

    needs_manual_qc = list(set(sorted(needs_metQC + needs_invQC)))

    def getQCmsg(site):
        msg = None
        if (site in needs_metQC) and (site in needs_invQC):
            msg = "met+inv"
        elif site in needs_metQC:
            msg = "met"
        elif site in needs_invQC:
            msg = "inv"
        return msg

    oprint("\n---------------- STATUS OVERVIEW ----------------")
    n_sites = len(site_fpaths)
    site_statuses = {
        "already have FlashReport files -- **waiting on utility meter (substituted with PI data)": [
            s if (s not in completed_no_meter) else f"**{s}" for s in completed
        ],
        "ready to run FlashReport script": ready_for_FR,
        "ready to run PVLib script": ready_for_PVLib,
        "ready to run Meteo QC/Backfill script": ready_for_metbackfill,
        "need manual QC of meteo and/or inverter data": [
            f"{s} ({getQCmsg(s)})" for s in needs_manual_qc
        ],
    }
    for message, sitelist in site_statuses.items():
        oprint(f"\n({len(sitelist)} of {n_sites}) sites {message}")
        if sitelist and (not output):
            printsites(sitelist, show_total=False, n_=4)
    allsites_ = [fp.name for fp in site_fpaths]
    oprint("")
    output_dict = {
        "has report with meter": completed_with_meter,
        "has report no meter": completed_no_meter,
        "needs report": ready_for_FR,
        "needs pvlib": ready_for_PVLib,
        "needs meteo backfill": ready_for_metbackfill,
        "needs manual qc": needs_manual_qc,
        "qc complete": [s for s in allsites_ if s not in (needs_metQC + needs_invQC)],
        "needs met qc": needs_metQC,
        "needs inv qc": needs_invQC,
    }
    if output:
        return output_dict
    return


def date_range(year, month, freq):
    t0 = pd.Timestamp(year=year, month=month, day=1)
    t1 = t0 + pd.DateOffset(months=1)
    date_range = pd.date_range(t0, t1, freq=freq)[:-1]
    return date_range


def load_monthly_query_files(
    site,
    year,
    month,
    types_=None,
    separate_dfs=False,
    q=True,
    return_fpaths=False,
):
    qprint = lambda msg, end="\n": None if q else print(msg, end=end)
    frpath = Path(oepaths.frpath(year, month, ext="Solar"), site)
    latestfile = lambda fplist: max((fp.stat().st_ctime, fp) for fp in fplist)[1]
    getfiles_ = lambda str_: list(frpath.glob(str_))
    getfile = lambda str_: None if (not getfiles_(str_)) else latestfile(getfiles_(str_))

    fp_dict = {
        "Meter": getfile(f"PIQuery_Meter*.csv"),
        "Inverters": getfile(f"PIQuery_Inverters*.csv"),
        "MetStations": getfile(f"PIQuery_MetStations*.csv"),
        "PVLib": getfile(f"PVLib_InvAC_{site}*.csv"),
    }
    if isinstance(types_, list):
        fp_dict = {k: fp for k, fp in fp_dict.items() if k in types_}

    fpath_list = []
    df_list = []
    df_dict = {}
    for key_, fpath in fp_dict.items():
        if fpath is None:
            qprint(f"No {key_} file found!")
            if not separate_dfs:
                continue
            df_dict[key_] = pd.DataFrame()

        df = pd.read_csv(fpath, index_col=0, parse_dates=True)
        fpath_list.append(str(fpath))

        if key_ == "Meter":
            mcol_ = [c for c in df.columns if "processed" not in c.casefold()]
            df = df[mcol_].copy()
            df.columns = ["PI_Meter_MW"]

        elif key_ == "Inverters":
            invcols = list(filter(lambda c: "OE.ActivePower" in c, df.columns))
            df = df[invcols]
            df["Actual_MW"] = df[invcols].sum(axis=1).div(1000).copy()

        elif key_ == "MetStations":
            cols = list(df.columns)
            end_col_idx = cols.index("Hour")
            df = df.iloc[:, :end_col_idx].copy()
            df = df[[c for c in df.columns if c != "PROCESSED"]]

        elif key_ == "PVLib":
            pvl_inv_cols = [c for c in df.columns if "Possible_Power" in c]
            keepcols = ["module_temperature", "POA"] + pvl_inv_cols
            df = df[keepcols]
            df["Possible_MW"] = df[pvl_inv_cols].sum(axis=1).div(1000).copy()

        qprint(f'Loaded "{fpath.name}"')
        df_list.append(df)
        df_dict[key_] = df.copy()

    if not df_list:
        if return_fpaths:
            return pd.DataFrame(), []
        return pd.DataFrame()

    if separate_dfs:
        if return_fpaths:
            return df_dict, fpath_list
        return df_dict

    df_all = pd.concat(df_list, axis=1)
    if return_fpaths:
        return df_all, fpath_list
    return df_all


# function to load utility meter data from historian
def load_meter_historian(year=None, month=None):
    with Path(oepaths.meter_generation_historian) as fpath:
        df = pd.read_excel(fpath, engine="openpyxl")
    df = df.iloc[: df.Day.last_valid_index() + 1, :].copy()
    df.index = pd.to_datetime(
        df["Day"].astype(str)
        + " "
        + df["Hour Ending"].astype(int).sub(1).astype(str).str.zfill(2)
        + ":00:00",
        format="%Y-%m-%d %H:%M:%S",
    )
    dropcols = ["Year", "Month", "Day", "Hour Ending"] + [c for c in df.columns if "Unnamed" in c]
    df = df.drop(columns=dropcols)
    if year is not None:
        df = df[(df.index.year == year)].copy()
    if month is not None:
        df = df[(df.index.month == month)].copy()

    df = df.loc[~df.index.duplicated()]

    return df


# function to load kpi tracker (updated Aug.2024 for rev1)
def load_kpi_tracker(sheet="ALL PROJECTS"):
    kwargs = dict(sheet_name=sheet, engine="openpyxl")
    header_row_dict = {"Ref Tables": 1, "Site List": 1, "ALL PROJECTS": 0}
    hdr_row = header_row_dict.get(sheet)
    keepcols = lambda df: [c for c in df.columns if "Unnamed" not in c]

    kpi_fpath = Path(oepaths.kpi_tracker_rev1)

    if hdr_row is not None:
        with kpi_fpath as fpath:
            df_ = pd.read_excel(fpath, **kwargs, header=hdr_row)
        if sheet != "Ref Tables":
            df = df_[keepcols(df_)].copy()
        else:
            df_.columns = df_.iloc[0].astype(str).str.replace(".0", "")
            end_row = df_[df_["Project Name"].isna()].index[0]
            end_col = list(df_.columns).index("Notes")
            df = df_.iloc[1:end_row, :end_col].dropna(axis=1, how="all").copy()
            df = df.set_index("Project Name")
    else:
        with kpi_fpath as fpath:
            df = pd.read_excel(fpath, **kwargs)

    if sheet == "ALL PROJECTS":
        end_row = df.Project.last_valid_index() + 1
        df = df.iloc[:end_row, :].copy()
        for col in ["Year", "Month"]:
            df[col] = df[col].astype(int)

    return df


# function to get ppa rate from kpi tracker  (used in flashreport script)
def get_ppa_rate(site, year):
    dfk = load_kpi_tracker(sheet="Ref Tables")
    try:
        ppa = dfk.at[site, str(year)] if (site in dfk.index) else None
    except:
        ppa = None
    return ppa


## function to load kpi tracker data
def get_kpis_from_tracker(sitelist, yearmonth_list):
    dfk = load_kpi_tracker()
    df_list = []
    ym_list = list(sorted(yearmonth_list, reverse=True))
    for site, ym_tup in itertools.product(sitelist, ym_list):
        year, month = ym_tup
        conditions = dfk.Project.eq(site) & dfk.Year.eq(year) & dfk.Month.eq(month)
        df_ = dfk[conditions].copy()
        df_list.append(df_)

    dfKPI = pd.concat(df_list, axis=0, ignore_index=True)
    # dfKPI = dfKPI.drop(columns=['Notes'])
    return dfKPI


def get_solar_budget_values(year, month):
    dfk = load_kpi_tracker()
    budget_cols = [c for c in dfk.columns if "budget" in c.casefold()]
    conditions_ = dfk.Year.eq(year) & dfk.Month.eq(month)
    target_cols = ["Project"] + budget_cols
    df_budget = dfk.loc[conditions_, target_cols].reset_index(drop=True).copy()
    df_budget["sortcol"] = df_budget.Project.str.casefold().copy()
    df_budget = df_budget.sort_values(by=["sortcol"])
    df_budget = df_budget.drop(columns=["sortcol"])
    return df_budget


##functions to read from Excel monthly FlashReport files (or any Excel file)
def calculate_and_read_excel_file(filepath, sheet, q=True, use_tempdir=True):
    fpath = Path(filepath)
    parent_dir = tempfile.TemporaryDirectory() if use_tempdir else fpath.parent
    qprint = lambda msg, end="\n": None if q else print(msg, end=end)
    showstatus = lambda msg: qprint(f">> {msg}".ljust(50), end="\r")

    if use_tempdir:
        msg_ = "(note: used temp directory - original file remains uncalculated)"
    else:
        msg_ = "(note: re-saved original file with calculated values)"

    with parent_dir as dir_:
        excelfp = Path(dir_, fpath.name)
        if use_tempdir:
            shutil.copy2(fpath, excelfp)
        pythoncom.CoInitialize()
        showstatus("co-initialized")
        xw_app = xw.App(visible=False)
        showstatus("opened Excel instance")
        wb = xw_app.books.open(excelfp)
        showstatus("opened target file")
        xw_app.calculate()
        showstatus("completed calculations")
        wb.save()
        showstatus("saved target file")
        wb.close()
        showstatus("closed target file")
        xw_app.quit()
        showstatus("quit Excel instance")
        pythoncom.CoUninitialize()
        showstatus("co-uninitialized")
        time.sleep(2)

        df = pd.read_excel(excelfp, sheet_name=sheet, engine="openpyxl")
        qprint(f"done! {msg_}".ljust(50))

    return df


# function to load flashreport file
def load_solar_flashreport(site, year, month, q=True, use_tempdir=True, return_df_and_fpath=False):
    qprint = lambda msg, end="\n": print(msg, end=end) if (not q) else None

    # get solar flashreport folders/paths
    site_fp = Path(oepaths.frpath(year, month, ext="Solar"), site)
    frfpaths = list(site_fp.glob("*FlashReport*.xlsx"))
    frfp = latest_file(frfpaths)

    if frfp is None:
        qprint("no FlashReport file found!")
        return pd.DataFrame()

    qprint(f'{(site+"..").ljust(20)}', end="")
    with open(frfp, "r") as fr_file:
        df = pd.read_excel(frfp, sheet_name="FlashReport", engine="openpyxl")
    qprint(f'loaded file: "{frfp.name}"')

    # check if equations have been calculated
    summary_is_blank = lambda df_FR: df_FR.iloc[6:9, 1:3].isna().all().all()

    if summary_is_blank(df):
        qprint("!! equations not calculated !!\nCalculating..")
        df = calculate_and_read_excel_file(frfp, sheet="FlashReport", q=q, use_tempdir=use_tempdir)
        if df.iloc[1:9, 1:3].isna().all().all():
            qprint("ERROR".ljust(50))
            df = pd.DataFrame()

    if return_df_and_fpath:
        return df, str(frfp)
    return df


def get_flashreport_summary_table(
    site, year, month, inverter_level=False, q=True, return_df_and_fpath=False
):
    # load flashreport
    output = load_solar_flashreport(site, year, month, q=q, return_df_and_fpath=return_df_and_fpath)
    df_ = output if not return_df_and_fpath else output[0]
    if df_.empty:
        if not q:
            print("error loading flashreport file.\nexiting..")
        return df_, output[1]

    if site in df_.loc[0].values:
        df_.columns = df_.loc[0].values
        df_ = df_.iloc[1:, :].copy()

    if not inverter_level:
        ## get site-level summary table
        end_row_idx = df_.loc[df_[site].astype(str).str.contains("Tcoeff")].index[0]
        df = df_.iloc[1:end_row_idx, :3].reset_index(drop=True).copy()
        df.columns = [site, "%", "Value"]

        # copy values from % to Value column for site capacity & insolation
        copy_idx_start = df.loc[df[site].str.contains("Meter Gen")].index[0] + 1
        for r in range(copy_idx_start, df.shape[0]):
            df.at[r, "Value"] = df.at[r, "%"]
            df.at[r, "%"] = np.nan

        df = df.set_index(site)
        df.index = df.index.str.strip()
        df = df[["Value", "%"]]
    else:
        ##get "Inverter Totals" summary table
        col_has_value = lambda idx_, n_rows, str_: (str_ in df_.iloc[:n_rows, idx_].values)
        start_col = 0  # init
        while (start_col < df_.shape[1]) and (not col_has_value(start_col, 5, "Inverter Totals")):
            start_col += 1

        end_col = 0  # init
        while (end_col < df_.shape[1]) and (not col_has_value(end_col, 20, "calcCol")):
            end_col += 1

        ## get start/end rows
        col0_vals = [v.strip() for v in list(df_.iloc[:15, start_col].values)]
        start_row = col0_vals.index("Inverter Totals")
        end_row = col0_vals.index("Possible Generation") + 1

        ## create dataframe from flashreport df_
        df = df_.iloc[start_row:end_row, start_col:end_col].copy()
        df.columns = df.loc[0].values
        df = df.iloc[1:, :].set_index("Inverter Totals").rename_axis(None)
        df = df.apply(pd.to_numeric, errors="coerce")
        df.index = [i.strip() for i in df.index.values]
        for col in df.columns:
            df.loc[df[col].lt(0), col] = 0.0

    if return_df_and_fpath:
        return df, output[1]
    return df


"""
BELOW FUNCTIONS ARE FOR COLLECTING KPIS FROM SOLAR FLASHREPORTS IN PREPARATION FOR TRANSFER TO TRACKER
"""
kpi_matching_cols = {
    "Site": "Project",
    "DTN Insolation [kWh/m^2]": "DTN POA Insolation (kWh/m2)",
    "Insolation [kWh/m^2]": "POA Insolation (kWh/m2)",
    "Possible Generation": "Possible Generation (MWh)",
    "Actual Generation": "Inverter Generation (MWh)",
    "Meter Generation": "Meter Generation (MWh)",
    "DC / System Health": "DC/System Health Loss (MWh)",
    "Downtime": "Downtime Loss (MWh)",
    "Curtailment": "Curtailment - Non_Compensable (MWh)",
    "Inverter Availability": "Inverter Uptime Availability (%)",
}
kpi_data_cols = [
    "DTN POA Insolation (kWh/m2)",
    "POA Insolation (kWh/m2)",
    "GHI Insolation (kWh/m2)",
    "Possible Generation (MWh)",
    "Inverter Generation (MWh)",
    "Meter Generation (MWh)",
    "Meter Generation - ADJUSTED (MWh)",
    "DC/System Health Loss (MWh)",
    "Snow Derate Loss (MWh)",
    "Downtime Loss (MWh)",
    "Curtailment - Compensable (MWh)",
    "Curtailment - Non_Compensable (MWh)",
    "Curtailment - Total (MWh)",
    "Insurance BI Adjustment (MWh)",
    "Inverter Uptime Availability (%)",
]


def get_flashreport_kpis(site, year, month, q=True, return_df_and_fpath=False):
    qprint = lambda msg, end="\n": print(msg, end=end) if (not q) else None

    # load summary table
    output = get_flashreport_summary_table(
        site, year, month, q=q, return_df_and_fpath=return_df_and_fpath
    )
    df__ = output if not return_df_and_fpath else output[0]
    if df__ is None:
        if return_df_and_fpath:
            return df__, output[1]
        return df__

    ## filter columns, move avail % to Value col, create entry for sitename
    df_ = df__.loc[df__.index.isin(kpi_matching_cols)].copy()
    avail_col = "Inverter Availability"
    df_.at[avail_col, "Value"] = df_.at[avail_col, "%"]
    df_.at["Site", "Value"] = df_.index.name

    ## transpose, reorder/rename columns
    df = df_[["Value"]].rename_axis(None).T.reset_index(drop=True).copy()
    reordered_cols = [c for c in kpi_matching_cols if c in df.columns]
    df = df[reordered_cols].rename(columns=kpi_matching_cols)

    ## add columns to match tracker
    for i, col in enumerate(kpi_data_cols):
        if col not in df.columns:
            df.insert(i + 1, col, np.nan)

    ## get GHI data from processed met file
    fp_dict = get_solar_fr_filepaths(site, year, month)
    mfp_ = fp_dict.get("met2")
    qprint(" " * 20, end="")
    if not isinstance(mfp_, Path):
        qprint(">> no processed met file! querying DTN GHI data..", end=" ")
        df_dtn = get_monthly_DTN_ghi_total(site, year, month)
        qprint("done!")
        ghi_total = df_dtn["DTN_GHI"].sum() / 1e3
    else:
        dfm = pd.read_csv(mfp_, index_col=0, parse_dates=True)
        qprint(f"loaded GHI from file: {mfp_.name}")
        ghi_total = dfm["Processed_GHI"].sum() / 1e3 / 60  # minute-level data

    df["GHI Insolation (kWh/m2)"] = (
        ghi_total  # column was previously inserted at the correct location
    )

    if return_df_and_fpath:
        return df, output[1]
    return df


def get_monthly_DTN_ghi_total(site, year, month, q=True):
    tz = oemeta.data["TZ"].get(site)
    lat, lon = oemeta.data["LatLong"].get(site)
    start = pd.Timestamp(year, month, day=1)
    end = start + pd.DateOffset(months=1)
    interval = "hour"
    fields = ["shortWaveRadiation"]
    df_dtn = query_DTN(lat, lon, start, end, interval, fields, tz=tz, q=q)
    df_dtn.columns = ["DTN_GHI"]
    dtn_ghi_adj = oemeta.data["DTN_GHI_adj"].get(site)
    df_dtn["DTN_GHI"] = df_dtn["DTN_GHI"].mul(dtn_ghi_adj)
    return df_dtn


def collect_solar_flashreport_kpis(year, month, sitelist=None, q=True, return_df_and_fpath=False):
    qprint = lambda msg: None if q else print(msg)
    if sitelist is not None:
        frsites_ = [s for s in sitelist if s in ALL_SOLAR_SITES]
    else:
        frsites_ = ALL_SOLAR_SITES
    skipped_sites = []  # init
    df_list = []  # init
    fpath_list = []
    for site in frsites_:
        if not list(site_frpath(site, year, month).glob("*FlashReport*.xlsx")):
            skipped_sites.append(site)
            continue
        output = get_flashreport_kpis(
            site, year, month, q=q, return_df_and_fpath=return_df_and_fpath
        )
        df_ = output if not return_df_and_fpath else output[0]
        df_list.append(df_.copy())
        if return_df_and_fpath:
            fpath_list.append(output[1])
    if len(skipped_sites) > 0:
        qprint(f"\nNote: KPIs were not collected for the following sites:\n{skipped_sites}\n")
    if not df_list:
        qprint("!! No kpis found !!\nexiting..")
        if return_df_and_fpath:
            return None, []
        return

    df = pd.concat(df_list, axis=0, ignore_index=True)

    ## overwrite negative values
    pos_cols = ["DC/System Health Loss (MWh)", "Downtime Loss (MWh)"]
    for col in pos_cols:
        df.loc[df[col].lt(0), col] = 0

    ## remove gross generation values that are less than net gen.
    gross_ = "Inverter Generation (MWh)"
    net_ = "Meter Generation (MWh)"
    df.loc[df[gross_].lt(df[net_]), gross_] = np.nan

    ## limit availability to 100%
    inv_avail_ = "Inverter Uptime Availability (%)"
    df.loc[df[inv_avail_].gt(1), inv_avail_] = 1.00

    if return_df_and_fpath:
        return df, fpath_list
    return df


def load_kpis_and_notes(year, month):
    notes_fp = Path(oepaths.kpi_notes_file)
    sheets_ = list(pd.read_excel(notes_fp, nrows=0, sheet_name=None).keys())
    target_sheet = f"{calendar.month_name[month]}_{year}"
    if target_sheet not in sheets_:
        print(f"No sheet exists in notes file for {year}-{month:02d}.\nExiting.")
        return

    dfn = pd.read_excel(notes_fp, sheet_name=target_sheet)
    hdr_idx = dfn[dfn.isin([target_sheet]).any(axis=1)].index[0]
    dfn.columns = dfn.iloc[hdr_idx].values
    row_end = dfn.iloc[:, 0].last_valid_index() + 1
    col_end = list(dfn.columns[1:]).index(target_sheet)  # second occurence of sheet name
    dfn = dfn.iloc[(hdr_idx + 1) : row_end, :col_end].reset_index(drop=True)
    return dfn


## add fill format for cell if using pi meter data (cell.fill = 'ddebf7'


def transfer_flashreport_kpis_to_notes_file(
    year, month, sitelist=None, df=None, overwrite=False, q=True
):
    qprint = lambda msg, end="\n": print(msg, end=end) if (not q) else None

    notes_fp = Path(oepaths.kpi_notes_file)
    sheets_ = list(pd.read_excel(notes_fp, nrows=0, sheet_name=None).keys())
    target_sheet = f"{calendar.month_name[month]}_{year}"
    if target_sheet not in sheets_:
        print(f"No sheet exists in notes file for {year}-{month:02d}.\nExiting.")
        return

    if not isinstance(df, pd.DataFrame):
        df = collect_solar_flashreport_kpis(year, month, sitelist=sitelist, q=q)

    # open workbook (note: had problems using as context-manager, so switched to explicit close() method)
    wb = openpyxl.load_workbook(notes_fp, rich_text=True)
    ws = wb[target_sheet]

    cell_vals = lambda id_="A": [c.value for c in ws[id_]]  # can be row or column
    site_row_number = lambda s: (cell_vals().index(s) + 1) if (s in cell_vals()) else None

    # get kpi col names from header row
    header_row = [
        (i + 1)
        for i, c in enumerate(cell_vals())
        if isinstance(c, openpyxl.worksheet.formula.ArrayFormula)
    ][0]
    header_vals = cell_vals(str(header_row))

    qprint(f"\nWriting new KPIs to notes file: ({overwrite=})")
    something_was_added = False  # init
    fmt_col = lambda col: col.split(" (")[0]

    # write to notes file
    for i, site in enumerate(cell_vals()):
        if site not in df.Project.values:
            continue
        qprint(f"\n{site}")

        added_ = []
        skipped_nan = []
        skipped_existing = []
        exclude_cols = [
            "Meter Generation - ADJUSTED (MWh)",
            "Curtailment - Compensable (MWh)",
            "Insurance BI Adjustment (MWh)",
        ]

        row = i + 1
        for j, kpi_col in enumerate(df.columns):
            if kpi_col in exclude_cols:
                continue
            if kpi_col in header_vals:
                col = header_vals.index(kpi_col) + 1
                kpi_value = df.loc[df.Project.eq(site), kpi_col].values[0]
                if pd.isnull(kpi_value):
                    skipped_nan.append(fmt_col(kpi_col))
                    continue
                if (ws.cell(row, col).value is not None) and (not overwrite):
                    skipped_existing.append(fmt_col(kpi_col))
                    continue
                ws.cell(row, col).value = kpi_value

                ## format cell fill color if PI meter data
                if kpi_col == "Meter Generation (MWh)":
                    frfp = latest_file(
                        list(site_frpath(site, year, month).glob("*FlashReport*.xlsx"))
                    )
                    if "rev0" in frfp.stem.casefold():
                        ws.cell(row, col).fill = PatternFill("solid", fgColor="ddebf7")

                something_was_added = True
                added_.append(kpi_col)
                qprint(f"  ++ added {kpi_col}")

        if not added_:
            qprint(f"  -- no new KPI values added to file ({overwrite=})")
            continue

        if skipped_nan:
            qprint(f"  -- skipped KPIs with no values: {skipped_nan}")
        if skipped_existing:
            qprint(f"  -- skipped existing KPIs ({overwrite=}): {skipped_existing}")

    # save/close
    if something_was_added:
        wb.save(notes_fp)
        qprint(f'\n\nKPI values successfully added to file: "{notes_fp.name}"')

    wb.close()

    return


# function to build dictionary of attribute paths from monthly query atts in pireference pqmeta
solarAFpath = "\\\\CORP-PISQLAF\\Onward Energy\\Renewable Fleet\\Solar Assets"


def get_query_attribute_paths(site):
    site_dict = ref.pqmeta.get(site)
    siteAFpath = f"{solarAFpath}\\{site}"

    # inverters
    inv_att_paths = []
    inverters_ = oemeta.data["AF_Solar_V3"][site].get("Inverters")
    if inverters_:
        inv_names = list(inverters_["Inverters_Assets"].keys())
        invPath = f"{siteAFpath}\\Inverters"
        inv_att_paths = [f"{invPath}\\{inv}|OE.ActivePower" for inv in inv_names]

    # combiners
    cmb_att_paths = []
    cmb_atts = site_dict.get("inv atts")
    cmb_sub_atts = site_dict.get("inv sub atts")
    if cmb_atts:
        for inv in inv_names:
            cmb_att_paths.extend([f"{invPath}\\{inv}|{att}" for att in cmb_atts])
    elif cmb_sub_atts:
        inv1 = inv_names[0]  # get cmb names from first inv (same for all)
        inv1_dict = inverters_["Inverters_Assets"].get(inv1)
        cmb_names = list(inv1_dict[f"{inv1}_Subassets"].keys())
        if site == "Imperial Valley":  # temp
            cmb_names = [c for c in cmb_names if c != "Combiner Boxes"]  # temp
        for inv in inv_names:
            for cmb in cmb_names:
                cmb_att_paths.extend([f"{invPath}\\{inv}\\{cmb}|{att}" for att in cmb_sub_atts])

    # met stations
    met_att_paths = []
    met_atts = site_dict.get("metsta atts")
    metPath = f"{siteAFpath}\\Met Stations"
    if met_atts:
        if isinstance(met_atts, dict):
            for met, att_list in met_atts.items():
                met_att_paths.extend([f"{metPath}\\{met}|{att}" for att in att_list])
        else:
            if "met ids" in site_dict:
                met_names = site_dict["met ids"]
            else:
                metstations_ = oemeta.data["AF_Solar_V3"][site].get("Met Stations")
                met_names = list(metstations_["Met Stations_Assets"].keys())

            for met in met_names:
                met_att_paths.extend([f"{metPath}\\{met}|{att}" for att in met_atts])

    # ppc
    ppc_atts = site_dict.get("ppc atts")
    ppc_att_paths = [f"{siteAFpath}\\PPC|{att}" for att in ppc_atts] if ppc_atts else []

    # trackers
    trk_att_paths = []
    trk_atts = site_dict.get("trk atts")
    trk_sub_atts = site_dict.get("trk sub atts")
    trkPath = f"{siteAFpath}\\Trackers"
    trackers_ = oemeta.data["AF_Solar_V3"][site].get("Trackers")
    trk_names = list(trackers_["Trackers_Assets"].keys()) if trackers_ else []
    if trk_atts:
        for trk in trk_names:
            t_atts = trk_atts.get(trk) if isinstance(trk_atts, dict) else trk_atts
            trk_att_paths.extend([f"{trkPath}\\{trk}|{att}" for att in t_atts])
    elif trk_sub_atts:
        trk1 = trk_names[0]
        trk1_dict = trackers_["Trackers_Assets"].get(trk1)
        mtr_names = list(trk1_dict[f"{trk1}_Subassets"].keys())
        for trk in trk_names:
            for mtr in mtr_names:
                trk_att_paths.extend([f"{trkPath}\\{trk}\\{mtr}|{att}" for att in trk_sub_atts])

    # meter
    meter_att_paths = []
    meter_ = oemeta.data["AF_Solar_V3"][site].get("Meter")
    all_meter_atts = meter_["Meter_Attributes"] if meter_ else []
    if "OE_MeterMW" in all_meter_atts:
        meter_att_paths.append(f"{siteAFpath}\\Meter|OE_MeterMW")

    ##temp
    if site == "GA4":
        meter_att_paths = [f"{siteAFpath}\\Meter|SEL3530.MTR_FSREV1.Sts.P_MW"]

    groups_ = ["Inverters", "Combiners", "Met Stations", "PPC", "Trackers", "Meter"]
    attpaths_ = [
        inv_att_paths,
        cmb_att_paths,
        met_att_paths,
        ppc_att_paths,
        trk_att_paths,
        meter_att_paths,
    ]
    output_dict = {grp: atts for grp, atts in zip(groups_, attpaths_) if atts}

    return output_dict


## sites with defined monthly query attributes in file "pireference.py"
pireference_sites = list(ref.pqmeta.keys())


def get_pqmeta_att_paths(sitelist=None):
    if sitelist is None:
        sitelist = pireference_sites
    else:
        sitelist = [s for s in sitelist if s in pireference_sites]

    attPath_dict = {site: get_query_attribute_paths(site) for site in sitelist}
    return attPath_dict
