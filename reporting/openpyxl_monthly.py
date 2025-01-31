import openpyxl
import pythoncom
import shutil
import tempfile
import pandas as pd
import xlwings as xw
from pathlib import Path
from tqdm.notebook import tqdm
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList
from openpyxl.formula.translate import Translator
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.chart import ScatterChart, BarChart, LineChart, Reference, Series
from oetoolbox.reporting.xlformatting import colors, formatting_props
from types import SimpleNamespace

# using "simple namespace" for dot notation reference to dict items
cc = SimpleNamespace(**formatting_props["colors"])
f = SimpleNamespace(**formatting_props["fill"])
b = SimpleNamespace(**formatting_props["border"])
a = SimpleNamespace(**formatting_props["align"])
fnt = SimpleNamespace(**formatting_props["font"])


def create_monthly_report(
    sitename,
    capacity,
    capacity_DC,
    mod_Tcoeff,
    dtn_insolation,  # NEW
    missinginvdata,
    peakkWh_dict,
    df4xl,
    dfm,
    dfc,
    fig,
    savepath,
    q=True,
    comanchePPCval=None,
    missingfiles=None,
    df_caiso=None,
    df_caiso2=None,
):
    """
    this function generates a monthly report Excel file using openpyxl
        inputs: (from flashreportfactory.py)
            sitename = name of site/project
            capacity = site capacity, MWac
            capacity_DC = site capacity, MWdc
            mod_Tcoeff = site mod. temp. coefficient
            missinginvdata = list of missing data rows per inverter
            peakkWh_dict = dict of peak kWh values (minute-level) for inv/pvlib
            df4xl = combined dataframe
            dfm = meter dataframe
            dfc = curtailment dataframe
            fig = missing data figure
            savepath = full report save path (incl. filename)
            q = quiet, bool (default: True) to disable printouts
            comanchePPCval = reported curtailment value for Comanche
            missingfiles = dict of data file existence bools
                         = {'Inverters': True/False,
                            'PPC': True/False,
                            'Meter': True/False}
            df_caiso = dataframe from site sheet in caiso file (if exists)
            df_caiso2 = second caiso df, 5min data - reindexed to fill missing timestamps
    """
    # define function to print events if q=False (adds padding of spaces for carriage return)
    print_event = lambda msg: print(msg.ljust(44), end="\r") if not q else None

    IS_MAPLEWOOD = "Maplewood" in sitename

    # create new blank workbook object & assign sheet to variable
    if not q:
        print("Generating report...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FlashReport"

    # get start/end time from df4xl
    t_start = pd.Timestamp(df4xl.Timestamp.min())
    t_end = t_start + pd.DateOffset(months=1)

    # variable for hourly timerange (used for df_caiso section)
    tstamps = pd.date_range(t_start, t_end, freq="h")[:-1]

    # variable for number of inverters
    numInv = len(missinginvdata)

    # variable for max. inv generation
    invcols = [c for c in df4xl.columns if "ActivePower" in c and "timeAvail" not in c]
    inv_maxMWh = 1  # default
    if invcols:
        dfi = df4xl[invcols].copy()
        inv_maxMWh = max(dfi.sum() / 1000)

    # calculation of total for tqdm progress bar
    """
        p1 = 10*(df4xl.shape[0]+1)               -df4xl in dataframe_to_rows
         >> len(row_range) = len(range(int(rStart)+1, int(rEnd)+1))
         >>                = (df4xl.shape[0]-1)
        p2 = 10*len(row_range)            -main data table formulas
        p3 = 10*len(row_range)            -meter data column
        p4 = len(row_range)*numInv     -pvlib adj. columns
        p5 = len(row_range)*numInv     -lost MWh columns
        if isinstance(df_caiso, pd.DataFrame):
            pc1 = df_caiso.shape[0]+1      -df_caiso in dataframe_to_rows
            pc2 = df_caiso.shape[0]-1      -formulas for minute timestamps
            pc3 = 10*(len(tstamps)-1)      -formulas for hourly SUPP/DOT
            pc4 = 10*len(row_range)        -adjusted caiso formula for curtSP
        p6 = 2400       -formatting; 100+100+100+300+300+300+300+300+400+200
        p7 = 800+       -charts/image; 200+200+100+200+100+110
        
        pTotal =           p1          +           p2          +          p3           ...
               = 10*(df4xl.shape[0]+1) + 10*(df4xl.shape[0]-1) + 10*(df4xl.shape[0]-1) ...
                ... +            p4             +            p5             +  p6  + p7
                ... + numInv*(df4xl.shape[0]-1) + numInv*(df4xl.shape[0]-1) + 2400 + 800
        
        pTotal = 30*df4xl.shape[0] + 10 - 10 - 10  + 2*numInv*df4xl.shape[0] - 2*numInv + 2400 + 910
        >>[[{ pTotal = (30+2*numInv)*df4xl.shape[0] - 2*numInv + 3300 }]]<<
        
        (if df_caiso)
        pcTotal =          pc1          +          pc2          +         pc3         +          pc4         
        pcTotal = (df_caiso.shape[0]+1) + (df_caiso.shape[0]-1) + 10*(len(tstamps)-1) + 10*(df4xl.shape[0]-1)
        >>[[{ pcTotal = 2*df_caiso.shape[0] + 10*len(tstamps) + 10*df4xl.shape[0] - 20 }]]<<
    """
    tqdm_total = (30 + 2 * numInv) * df4xl.shape[0] - 2 * numInv + 3300
    if isinstance(df_caiso, pd.DataFrame):
        pcTotal = 2 * df_caiso.shape[0] + 10 * len(tstamps) + 10 * df4xl.shape[0] - 20
        tqdm_total += pcTotal

    # create tqdm progress bar for manual updates
    pbar = tqdm(
        desc="REPORT STATUS",
        total=tqdm_total,
        disable=q,
        bar_format="{bar}|{desc}: {percentage:0.0f}% | {n_fmt}/{total_fmt} |",
    )

    # write all data to worksheet - [Timestamp, POA, Inv. Power, PVLib Power, Hourly Avail., Lost MWh]
    print_event("writing raw data to worksheet")
    for r in dataframe_to_rows(df4xl, index=False, header=True):
        ws.append(r)
        pbar.update(
            10
        )  # total: df4xl.shape[0] + 1    [[[700]]]  -744 if 31 days, 720 if 30 days, 672 if 28 days, or 696 if 29 days

    print_event("initializing rows and columns")
    """
    order of columns in df4xl:
        Timestamp | POA | POA_all_bad | module_temp | InverterColumns | pvlibColumns | AvailabilityColumns
    """
    # add columns/rows to separate data
    col_invData = 5  # in df4xl inv data is after tstamp, POA, badPOA, and modTemp
    col_pvlData = col_invData + numInv  # pvlib inv data start column
    col_hrAvail = col_pvlData + numInv  # hourly availability data start column
    ws.insert_cols(col_hrAvail, 1)  # +1 col before hourly avail data
    ws.insert_cols(col_pvlData, 1)  # +1 col before pvlib data (for calcCol)

    newcols = [
        "Actual",
        "Possible",
        "Tcell",
        "ENDCi",
        "Perf_Ratio",
        "Curt_Loss",
        "Curt_SP",
        "Flag",
        "Meter",
    ]

    if IS_MAPLEWOOD:
        newcols.insert(2, "Possible_Adj")

    n_newcols = len(newcols)  # add column for "Possible_Adj" (due to transformer derate)
    ws.insert_cols(
        col_invData, n_newcols
    )  # +9 cols before inv data (for actual, poss, (possadj), tcell, endci, pratio, cLoss, cSP, Flag, meter)

    header_row = 16  # set header row number
    ws.insert_rows(1, header_row - 1)  # insert rows at top of sheet for KPI/summary table

    """
    order of columns for report:
        Timestamp | POA | badPOA | modTemp | ActualGen | PossGen | Tcell | ENDCi | Perf. Ratio | CurtLoss | CurtSP | Flag | ...
        ... MeterGen | InverterColumns | calcCol | pvlibColumns | (emptycol) | AvailabilityColumns | ...
        ... (emptycol) | adjPVLibColumns | (emptycol) | LostMWhColumns
    """

    class Col:
        def __init__(self, ltr, width):
            self.ltr = ltr
            self.width = width
            self.num = "ABCDEFGHIJKLMNOPQRSTUVWXYZ".index(ltr) + 1

    sheetcols_ = ["Timestamp", "POA", "badPOA", "modTemp"] + newcols

    # create column objects/variables
    makecol = lambda ltr, wd: Col(ltr, wd)
    ltrs = "ABCDEFGHIJKLMNOPQ"[: len(sheetcols_)]
    wdths = [25, 12, 12, 12, 16, 16, 16, 16, 16, 12, 12, 6, 21]
    if IS_MAPLEWOOD:
        wdths.insert(2, 16)  # for possible_adj

    # create list of Col instances for cols A through I
    Col_list = list(map(makecol, ltrs, wdths))

    # assign variables to Cols using list of Col instances
    # 'A',  'B',   'C',    'D',  'E',   'F',    'G',    'H',     'I',     'J',     'K',   'L',  'M'
    # cTime, cPOA, cPOAb, cMTemp, cAct, cPoss, cTcell, cENDCi, cPRatio, cCrLoss, cCurtSP, cFlag, cMtr = Col_list

    # 'A',  'B',   'C',    'D',  'E',   'F'
    cTime, cPOA, cPOAb, cMTemp, cAct, cPoss = Col_list[:6]

    if IS_MAPLEWOOD:
        cPossAdj = Col_list[6]

    #  'G',    'H',     'I',     'J',     'K',   'L',  'M'
    cTcell, cENDCi, cPRatio, cCrLoss, cCurtSP, cFlag, cMtr = Col_list[-7:]
    #  'H',    'I',     'J',     'K',     'L',   'M',  'N'  <<for Maplewoods

    # get max row/column, then create variables to dynamically set table range
    rEnd = ws.max_row
    cEnd = ws.max_column

    cAvail_end = get_column_letter(cEnd)
    cAvail_num = cEnd - numInv + 1
    cAvail = get_column_letter(cAvail_num)

    cPVLib_end_num = cAvail_num - 2
    cPVLib_end = get_column_letter(cPVLib_end_num)
    cPVLib_num = cPVLib_end_num - numInv + 1
    cPVLib = get_column_letter(cPVLib_num)

    cInv_end_num = cPVLib_num - 2
    cInv_end = get_column_letter(cInv_end_num)
    cInv_num = cInv_end_num - numInv + 1
    cInv = get_column_letter(cInv_num)

    cInv_calc = get_column_letter(cInv_end_num + 1)
    cAdjFactor = get_column_letter(cEnd + 1)

    cPVLadj_num = cEnd + 2
    cPVLadj = get_column_letter(cPVLadj_num)
    cPVLadj_end_num = cPVLadj_num + numInv - 1
    cPVLadj_end = get_column_letter(cPVLadj_end_num)

    cMWhLost_num = cPVLadj_end_num + 2
    cMWhLost = get_column_letter(cMWhLost_num)
    cMWhLost_end_num = cMWhLost_num + numInv - 1
    cMWhLost_end = get_column_letter(cMWhLost_end_num)

    # map rows to variables
    rSet = header_row  # hard-coded; drives all other row number references
    rHeader = str(rSet)
    rStart = str(rSet + 1)
    r1b4hdr = str(rSet - 1)
    r2b4hdr = str(rSet - 2)

    poasource = "↙ from DTN data" if "POA_DTN" in df4xl.columns else "↙ from POA sensors"

    # create dictionary of cell values to be written
    dict_cellVals = {
        "A1": f"{sitename}",
        "B2": "%",
        "B1": "Site Totals",
        "C2": "MWh",
        f"{cMtr.ltr}2": "Inverter Totals",
        "A3": "Total Energetic Loss ",
        f"{cFlag.ltr}3": "MWh",
        f"{cMtr.ltr}3": "Total Energetic Loss ",
        "A4": "DC / System Health ",
        f"{cFlag.ltr}4": "MWh",
        f"{cMtr.ltr}4": "DC/System Health ",
        "A5": "Inverter Availability ",
        f"{cFlag.ltr}5": "%",
        f"{cMtr.ltr}5": "Inverter Availability ",
        "A6": "Downtime ",
        f"{cFlag.ltr}6": "MWh",
        f"{cMtr.ltr}6": "Downtime ",
        "A7": "Curtailment ",
        f"{cFlag.ltr}7": "MWh",
        f"{cMtr.ltr}7": "Curtailment ",
        "A8": "Actual Generation ",
        f"{cFlag.ltr}8": "MWh",
        f"{cMtr.ltr}8": "Actual Generation ",
        "A9": "Possible Generation ",
        f"{cFlag.ltr}9": "MWh",
        f"{cMtr.ltr}9": "Possible Generation ",
        "A10": "Meter Generation ",
        f"{cMtr.ltr}10": "peak actual kWh ",
        "A11": "Site Capacity [MWdc] ",
        "B11": capacity_DC,
        f"{cMtr.ltr}11": "peak possible kWh ",
        "A12": "Site Capacity [MWac] ",
        "B12": capacity,
        f"{cMtr.ltr}12": "production ratio (raw) ",
        "A13": "Insolation [kWh/m^2] ",
        f"{cMtr.ltr}13": "production ratio ",
        "A14": "DTN Insolation [kWh/m^2] ",
        "B14": dtn_insolation,
        "A15": "Module Tcoeff [%] ",
        "B15": mod_Tcoeff,
        f"{cTime.ltr}{rHeader}": "Timestamp",
        f"{cPOA.ltr}{rHeader}": "POA",
        f"{cPOAb.ltr}{r1b4hdr}": poasource,
        f"{cPOAb.ltr}{rHeader}": "badPOA",
        f"{cMTemp.ltr}{rHeader}": "modTemp",
        f"{cAct.ltr}{rHeader}": "Actual [MWh]",
        f"{cPoss.ltr}{rHeader}": "Possible [MWh]",
        f"{cTcell.ltr}{rHeader}": "Tcell [C]",
        f"{cTcell.ltr}{r2b4hdr}": "Tcell_typ_avg:",
        f"{cENDCi.ltr}{rHeader}": "ENDCi [MWh]",
        f"{cPRatio.ltr}{rHeader}": "Perf. Ratio",
        f"{cCrLoss.ltr}{rHeader}": "Curt. Loss",
        f"{cCurtSP.ltr}{rHeader}": "Curt. SP",
        f"{cFlag.ltr}{rHeader}": "Flag",
        f"{cCurtSP.ltr}{r1b4hdr}": "flag sensitivity:",
        f"{cCurtSP.ltr}{r2b4hdr}": "irradiance threshold:",
        f"{cCurtSP.ltr}12": "avg.",
        f"{cMtr.ltr}{r2b4hdr}": "data % non-null:",
        f"{cMtr.ltr}{r1b4hdr}": f"missing values (of {df4xl.shape[0]}):",
        f"{cMtr.ltr}{rHeader}": "Meter Generation",
        f"{cInv_calc}{rHeader}": "calcCol",
        f"{cPVLib}{r1b4hdr}": "PVLib Inverter Data:",
        f"{cAvail}{r1b4hdr}": "Inverter Hourly Availability:",
        f"{cPVLadj}{r1b4hdr}": "PPC-Adjusted PVLib Data:",
        f"{cMWhLost}{r1b4hdr}": "Inverter Downtime kWh Lost:",
    }

    ## add for Maplewoods
    if IS_MAPLEWOOD:
        dict_cellVals.update({f"{cPossAdj.ltr}{rHeader}": "Possible_Adj [MWh]"})

    # write cell values from dictionary
    print_event("writing table header cell values")
    for cell_ in dict_cellVals:
        ws[cell_] = dict_cellVals[cell_]

    # misc cell address reference variables
    ws[f"{cMtr.ltr}1"] = "# of inverters:"
    numInv_cell = f"${cInv}$1"
    ws[numInv_cell] = numInv  # for curt. total equation (facilitates copying formula b/w reports)

    siteDCcapacity_cell = "$B$11"  # manual update
    sitecapacity_cell = "$B$12"  # manual update
    siteTcoeff_cell = "$B$15"  # manual update
    RAWratio_row = "12"
    adjRatio_row = "13"

    avgRAWratio_cell = f"${cFlag.ltr}${RAWratio_row}"
    ws[avgRAWratio_cell] = f'=AVERAGEIF({cInv}{RAWratio_row}:{cInv_end}{RAWratio_row},">0")'

    sensFactor_cell = f"${cFlag.ltr}${r1b4hdr}"  # sensitivity factor
    ws[sensFactor_cell] = 0.05  # default 5%; deviation limit b/w Actual & Meter power

    irrThreshold_cell = f"${cFlag.ltr}${r2b4hdr}"  # irradiance threshold
    ws[irrThreshold_cell] = 25  # default 25 kW/m^2

    tcellTypeAvg_cell = f"${cENDCi.ltr}${r2b4hdr}"
    ws[tcellTypeAvg_cell] = (
        f"=SUMPRODUCT(${cPOA.ltr}${rStart}:${cPOA.ltr}${rEnd},${cTcell.ltr}${rStart}:${cTcell.ltr}${rEnd})"
        f"/"
        f"SUM(${cPOA.ltr}${rStart}:${cPOA.ltr}${rEnd})"
    )

    """FORMULAS FOR ROWS IN 'INVERTER TOTALS' SECTION"""
    eqn_nrgLoss = f"={cInv}4+{cInv}6"
    eqn_dcHealth = f"={cInv}9-{cInv}8-{cInv}7-{cInv}6"
    eqn_availTotal = f"=IF({cInv}8>0,({cInv}8+{cInv}7)/({cInv}8+{cInv}7+{cInv}6),0)"
    eqn_downtime = f"=SUM({cMWhLost}{rStart}:{cMWhLost}{rEnd})/1000"
    eqn_curtTotal = (
        f'=(SUMIFS({cPVLadj}{rStart}:{cPVLadj}{rEnd},${cPOA.ltr}${rStart}:${cPOA.ltr}${rEnd},">="&{irrThreshold_cell},'
        f'${cCrLoss.ltr}${rStart}:${cCrLoss.ltr}${rEnd},">0")'
        f'-(SUMIFS({cInv}{rStart}:{cInv}{rEnd},${cPOA.ltr}${rStart}:${cPOA.ltr}${rEnd},">="&{irrThreshold_cell},'
        f'${cCrLoss.ltr}${rStart}:${cCrLoss.ltr}${rEnd},">0",${cFlag.ltr}${rStart}:${cFlag.ltr}${rEnd},1)'
        f'+(SUMIFS(${cMtr.ltr}${rStart}:${cMtr.ltr}${rEnd},${cCrLoss.ltr}${rStart}:${cCrLoss.ltr}${rEnd},">0",'
        f'${cFlag.ltr}${rStart}:${cFlag.ltr}${rEnd},"<>1")*1000/{numInv_cell})))/1000'
    )
    eqn_invActual = (
        f"=(SUM({cInv}{rStart}:{cInv}{rEnd})/1000)"
        f"+(((SUMIFS(${cMtr.ltr}${rStart}:${cMtr.ltr}${rEnd},${cFlag.ltr}${rStart}:${cFlag.ltr}${rEnd},0))/0.98)"
        f"*1000/COUNTA(${cInv}${rHeader}:${cInv_end}${rHeader}))/1000"
        f"+(SUMIFS(${cInv_calc}${rStart}:${cInv_calc}${rEnd},${cFlag.ltr}${rStart}:${cFlag.ltr}${rEnd},2,"
        f"{cInv}{rStart}:{cInv}{rEnd},0))/1000"
    )
    eqn_invPoss = f"=SUM({cPVLadj}{rStart}:{cPVLadj}{rEnd})/1000"
    eqn_pctNN = f"=1-({cInv}{r1b4hdr}/{df4xl.shape[0]})"
    eqn_prodRatioRAW = (
        f"=SUMIFS({cInv}{rStart}:{cInv}{rEnd},"  # sum inverter power if:
        f'${cPOA.ltr}${rStart}:${cPOA.ltr}${rEnd},">="&{irrThreshold_cell},'  # condition1, POA>threshold
        f"${cCrLoss.ltr}${rStart}:${cCrLoss.ltr}${rEnd},0,"  # condition2, CurtLoss=0
        f"{cAvail}{rStart}:{cAvail}{rEnd},60)"  # condition3, Availability=60
        f"/"  # div
        f"SUMIFS(${cPOA.ltr}${rStart}:${cPOA.ltr}${rEnd},"  # sum of POA if:
        f'${cPOA.ltr}${rStart}:${cPOA.ltr}${rEnd},">="&{irrThreshold_cell},'  # condition1, POA>threshold
        f"${cCrLoss.ltr}${rStart}:${cCrLoss.ltr}${rEnd},0,"  # condition2, CurtLoss=0
        f"{cAvail}{rStart}:{cAvail}{rEnd},60)"  # condition3, Availability=60
    )
    eqn_prodRatioAdj = f"=IF(ISERROR({cInv}{RAWratio_row}),{avgRAWratio_cell},{cInv}{RAWratio_row})"

    """FORMULAS FOR COLUMNS IN DATA TABLE"""
    eqn_siteActual = f"=SUM({cInv}{rStart}:{cInv_end}{rStart})/1000"
    eqn_sitePoss = (
        f"=IF((SUM({cPVLib}{rStart}:{cPVLib_end}{rStart})/1000)>{sitecapacity_cell},{sitecapacity_cell},"
        f"SUM({cPVLib}{rStart}:{cPVLib_end}{rStart})/1000)"
    )

    eqn_Tcell = f"={cMTemp.ltr}{rStart}+({cPOA.ltr}{rStart}/1000)*3"
    eqn_ENDCi = (
        f"={siteDCcapacity_cell}*({cPOA.ltr}{rStart}/1000)*"
        f"(1-({siteTcoeff_cell}*({tcellTypeAvg_cell}-{cTcell.ltr}{rStart})))"
    )
    eqn_PRatio = f"={cMtr.ltr}{rStart}/{cENDCi.ltr}{rStart}"

    eqn_curtLoss = (
        f"=IF(AND({cCurtSP.ltr}{rStart}<{sitecapacity_cell},{cCurtSP.ltr}{rStart}<{cPoss.ltr}{rStart},"
        f"{cPoss.ltr}{rStart}-{cMtr.ltr}{rStart}>0),{cPoss.ltr}{rStart}-{cMtr.ltr}{rStart},0)"
    )
    eqn_flag = (
        f"=IF(AND({cAct.ltr}{rStart}=0,{cMtr.ltr}{rStart}>0),0,"
        f"IF({cMtr.ltr}{rStart}>{cAct.ltr}{rStart}*(1+{sensFactor_cell}),2,1))"
    )
    eqn_calcCol = (
        f'=IF(COUNTIFS({cInv}{rStart}:{cInv_end}{rStart},"=0")<>0,'
        f'({cMtr.ltr}{rStart}*1.02-{cAct.ltr}{rStart})*1000/(COUNTIFS({cInv}{rStart}:{cInv_end}{rStart},"=0")),0)'
    )
    eqn_adjFactor = (
        f"=IFERROR(${cPoss.ltr}{rStart}*1000/SUM({cPVLib}{rStart}:{cPVLib_end}{rStart}),1)"
    )
    eqn_adjPVLib = f"={cPVLib}{rStart}*${cAdjFactor}{rStart}"
    eqn_mwhLost = (
        f"=IF(AND(${cPOA.ltr}{rStart}>={irrThreshold_cell},${cCrLoss.ltr}{rStart}=0,{cAvail}{rStart}<60,"
        f"${cFlag.ltr}{rStart}=1,({cInv}${adjRatio_row}*${cPOA.ltr}{rStart}-{cInv}{rStart})>0),"
        f"{cInv}${adjRatio_row}*${cPOA.ltr}{rStart}-{cInv}{rStart},0)"
    )

    """FORMULAS FOR SINGLE CELLS - SITE TOTALS & INSOLATION"""
    eqn_totalenergyloss = "=(C8+C7)/(C8+C7+C3)"
    eqn_totaldchealth = "=(C8+C7)/(C8+C7+C4)"
    eqn_invavailability = "=(C8+C7)/(C8+C7+C6)"
    eqn_totalcurtailment = "=1-(C8/(C8+C7))"
    eqn_insolation = f'=SUMIFS({cPOA.ltr}{rStart}:{cPOA.ltr}{rEnd},{cPOA.ltr}{rStart}:{cPOA.ltr}{rEnd},">=0")/1000'
    eqn_totalmeter = f"=SUM({cMtr.ltr}{rStart}:{cMtr.ltr}{rEnd})"

    # create dictionary of cell formulas to be written
    dict_4mulas = {
        "B3": eqn_totalenergyloss,
        "B4": eqn_totaldchealth,
        "B5": eqn_invavailability,
        "B7": eqn_totalcurtailment,
        "B13": eqn_insolation,
        "C10": eqn_totalmeter,
        f"{cInv}3": eqn_nrgLoss,
        f"{cInv}4": eqn_dcHealth,
        f"{cInv}5": eqn_availTotal,
        f"{cInv}6": eqn_downtime,
        f"{cInv}7": eqn_curtTotal if sitename != "Comanche" else f"=$C$7/${cInv}$1",
        f"{cInv}8": eqn_invActual,
        f"{cInv}9": eqn_invPoss,
        f"{cInv}{RAWratio_row}": eqn_prodRatioRAW,
        f"{cInv}{adjRatio_row}": eqn_prodRatioAdj,
        f"{cInv}{r2b4hdr}": eqn_pctNN,
        f"{cAct.ltr}{rStart}": eqn_siteActual,
        f"{cPoss.ltr}{rStart}": eqn_sitePoss,
        f"{cTcell.ltr}{rStart}": eqn_Tcell,  #####NEW
        f"{cENDCi.ltr}{rStart}": eqn_ENDCi,  #####NEW
        f"{cPRatio.ltr}{rStart}": eqn_PRatio,  #####NEW
        f"{cCrLoss.ltr}{rStart}": eqn_curtLoss,
        f"{cCurtSP.ltr}{rStart}": dfc["Curt_SP"][0],  # overwrite for CAISO data (below)
        f"{cFlag.ltr}{rStart}": eqn_flag,
        f"{cInv_calc}{rStart}": eqn_calcCol,
        f"{cAdjFactor}{rStart}": eqn_adjFactor,
        f"{cPVLadj}{rStart}": eqn_adjPVLib,
        f"{cMWhLost}{rStart}": eqn_mwhLost,
        f"{cMtr.ltr}{rStart}": dfm[sitename].iloc[0],
    }

    # write formulas from dictionary
    print_event("writing cell equations/formulas")
    for cell_ in dict_4mulas:
        ws[cell_] = dict_4mulas[cell_]

    """OTHER SITE TOTALS EQNS"""
    # write formulas for Site Totals summary -EnerLoss, SysHealth, Downtime, Curtailment, Actual/PossGen
    for r in range(3, 10):
        if r != 5:
            ws[f"C{r}"] = f"=SUM({cInv}{r}:{cInv_end}{r})"

    """overwrite curtailment in main ws condition 1"""
    # FOR COMANCHE -- overwrite C7 curtailment value with that from separate Comanche report; move original calc to D7
    if sitename == "Comanche":
        ws["C7"] = comanchePPCval if comanchePPCval else f"=SUM({cInv}7:{cInv_end}7)"
        ws["D7"] = (
            f'<< copied from "Curtailment Report" file (..\Commercial\Comanche Invoices\..)'
            if comanchePPCval
            else "note: reported curt. value not found"
        )

    # copy inverter names for summary table & write # of missing inv entries above data table
    for i in range(cInv_num, cInv_end_num + 1):
        col = get_column_letter(i)
        invname_contents = str(ws[f"{col}{rHeader}"].value).split("_")

        ws[f"{col}2"] = (
            invname_contents[-1] if sitename != "Azalea" else "_".join(invname_contents[-2:])
        )
        ws[f"{col}{r1b4hdr}"] = missinginvdata[i - cInv_num]

    # extend formulas in the 'Inverter Totals' section
    print_event("extending formulas for inv totals")
    invtotals_cols = range(cInv_num + 1, cInv_end_num + 1)
    for c in invtotals_cols:
        col = get_column_letter(c)
        ws[f"{col}3"] = Translator(eqn_nrgLoss, origin=f"{cInv}3").translate_formula(f"{col}3")
        ws[f"{col}4"] = Translator(eqn_dcHealth, origin=f"{cInv}4").translate_formula(f"{col}4")
        ws[f"{col}5"] = Translator(eqn_availTotal, origin=f"{cInv}5").translate_formula(f"{col}5")
        ws[f"{col}6"] = Translator(eqn_downtime, origin=f"{cInv}6").translate_formula(f"{col}6")
        if (sitename == "Comanche") and (comanchePPCval is not None):
            ws[f"{col}7"] = f"=$C$7/${cInv}$1"
        else:
            ws[f"{col}7"] = Translator(eqn_curtTotal, origin=f"{cInv}7").translate_formula(
                f"{col}7"
            )
        ws[f"{col}8"] = Translator(eqn_invActual, origin=f"{cInv}8").translate_formula(f"{col}8")
        ws[f"{col}9"] = Translator(eqn_invPoss, origin=f"{cInv}9").translate_formula(f"{col}9")
        ws[f"{col}{RAWratio_row}"] = Translator(
            eqn_prodRatioRAW, origin=f"{cInv}{RAWratio_row}"
        ).translate_formula(f"{col}{RAWratio_row}")
        ws[f"{col}{adjRatio_row}"] = Translator(
            eqn_prodRatioAdj, origin=f"{cInv}{adjRatio_row}"
        ).translate_formula(f"{col}{adjRatio_row}")
        ws[f"{col}{r2b4hdr}"] = Translator(eqn_pctNN, origin=f"{cInv}{r2b4hdr}").translate_formula(
            f"{col}{r2b4hdr}"
        )

    # extend formulas down columns in data table
    row_range = range(int(rStart) + 1, int(rEnd) + 1)
    for r in row_range:
        ws[f"{cAct.ltr}{r}"] = Translator(
            eqn_siteActual, origin=f"{cAct.ltr}{rStart}"
        ).translate_formula(f"{cAct.ltr}{r}")
        ws[f"{cPoss.ltr}{r}"] = Translator(
            eqn_sitePoss, origin=f"{cPoss.ltr}{rStart}"
        ).translate_formula(f"{cPoss.ltr}{r}")

        ws[f"{cTcell.ltr}{r}"] = Translator(
            eqn_Tcell, origin=f"{cTcell.ltr}{rStart}"
        ).translate_formula(f"{cTcell.ltr}{r}")
        ws[f"{cENDCi.ltr}{r}"] = Translator(
            eqn_ENDCi, origin=f"{cENDCi.ltr}{rStart}"
        ).translate_formula(f"{cENDCi.ltr}{r}")
        ws[f"{cPRatio.ltr}{r}"] = Translator(
            eqn_PRatio, origin=f"{cPRatio.ltr}{rStart}"
        ).translate_formula(f"{cPRatio.ltr}{r}")

        ws[f"{cCurtSP.ltr}{r}"] = dfc["Curt_SP"][r - int(rStart)]
        ws[f"{cCrLoss.ltr}{r}"] = Translator(
            eqn_curtLoss, origin=f"{cCrLoss.ltr}{rStart}"
        ).translate_formula(f"{cCrLoss.ltr}{r}")
        ws[f"{cFlag.ltr}{r}"] = Translator(
            eqn_flag, origin=f"{cFlag.ltr}{rStart}"
        ).translate_formula(f"{cFlag.ltr}{r}")
        ws[f"{cInv_calc}{r}"] = Translator(
            eqn_calcCol, origin=f"{cInv_calc}{rStart}"
        ).translate_formula(f"{cInv_calc}{r}")
        ws[f"{cAdjFactor}{r}"] = Translator(
            eqn_adjFactor, origin=f"{cAdjFactor}{rStart}"
        ).translate_formula(f"{cAdjFactor}{r}")
        pbar.update(10)  # total: len(row_range)

    """WRITE PEAK kWh VALUES FROM INPUT"""
    print_event("writing peak kWh data")
    for i, c in enumerate(range(cInv_num, cInv_end_num + 1)):
        col = get_column_letter(c)
        ws[f"{col}10"] = peakkWh_dict["Inverters"][i]
        ws[f"{col}11"] = peakkWh_dict["PVLib"][i]

    """WRITE METER DATA FROM FILE"""
    print_event("writing meter data")
    for r in row_range:
        ws[f"{cMtr.ltr}{r}"] = dfm[sitename].iloc[r - int(rStart)]
        pbar.update(10)  # total: len(row_range)

    """EXTEND/WRITE ADJUSTED PVLIB FORMULA"""
    print_event("extending adjusted pvlib formulas")
    # extend formula across first row of table
    for c in range(cPVLadj_num + 1, cPVLadj_end_num + 1):
        col = get_column_letter(c)
        ws[f"{col}{rStart}"] = Translator(
            eqn_adjPVLib, origin=f"{cPVLadj}{rStart}"
        ).translate_formula(f"{col}{rStart}")

    # copy inv names from summary table & extend formulas down columns of data table
    adjPVLib_range = range(cPVLadj_num, cPVLadj_end_num + 1)
    for i, c in enumerate(adjPVLib_range):
        col = get_column_letter(c)
        invnamecol = get_column_letter(cInv_num + i)
        ws[f"{col}{rHeader}"] = ws[f"{invnamecol}2"].value

        eqn_origin_cell = (
            f"{col}{rStart}"  # get equation in first cell of column & use to extend down rows
        )
        colEqn_adjPVLib = ws[eqn_origin_cell].value
        for r in row_range:
            ws[f"{col}{r}"] = Translator(colEqn_adjPVLib, origin=eqn_origin_cell).translate_formula(
                f"{col}{r}"
            )
            pbar.update(
                1
            )  # total: len(row_range)*numInv     700*(anywhere from 13 to 72)      [[[10k to 50k]]]****** -- scaling factor: 1

    """EXTEND/WRITE MWh LOST FORMULA"""
    print_event("writing formulas for MWh Lost")
    # copy inverter names for MWh Lost table header & extend formula across rows/columns
    # first, extend formula across first row in table
    for c in range(cMWhLost_num + 1, cMWhLost_num + numInv):
        col = get_column_letter(c)
        ws[f"{col}{rStart}"] = Translator(
            eqn_mwhLost, origin=f"{cMWhLost}{rStart}"
        ).translate_formula(f"{col}{rStart}")

    # next, extend formulas in first row down all rows
    print_event("extending MWh Lost formulas")
    for c in range(cMWhLost_num, cMWhLost_end_num + 1):
        col = get_column_letter(c)
        nmCol_num = cInv_num + c - cMWhLost_num
        nmCol = get_column_letter(nmCol_num)
        ws[f"{col}{rHeader}"] = str(ws[f"{nmCol}{rHeader}"].value).split("_")[-1]

        eqn_origin_cell = (
            f"{col}{rStart}"  # get equation in first cell of column & use to extend down rows
        )
        colEqn_mwhLost = ws[eqn_origin_cell].value
        for r in row_range:
            ws[f"{col}{r}"] = Translator(colEqn_mwhLost, origin=eqn_origin_cell).translate_formula(
                f"{col}{r}"
            )
            pbar.update(
                1
            )  # total: len(row_range)*numInv     700*(anywhere from 13 to 72)      [[[10k to 50k]]]****** -- scaling factor: 1

    """ADD CAISO SHEETS IF DATA EXISTS FOR SITE"""
    if isinstance(df_caiso, pd.DataFrame):
        """FIRST CAISO SHEET"""
        wb.create_sheet("CAISO")
        caiso_ws = wb["CAISO"]

        print_event("adding CAISO DOT data to new sheet")
        for r in dataframe_to_rows(df_caiso, index=False, header=True):
            caiso_ws.append(r)
            pbar.update(1)  # total: df_caiso.shape[0] + 1    [[[9k]]] -- scaling factor: 1

        lastCol = caiso_ws.max_column
        lastRow = caiso_ws.max_row

        cminuteTstamps = "AP"
        cPeriodStart = "AQ"
        cPeriodEnd = "AR"
        cDOT = "AS"
        cSUPP = "AT"

        caiso_ws[f"{cPeriodStart}1"] = "Period Start"
        caiso_ws[f"{cPeriodEnd}1"] = "Period End"
        caiso_ws[f"{cDOT}1"] = "DOT"
        caiso_ws[f"{cSUPP}1"] = "SUPP"

        # use hourly timerange (defined at beginning of function)
        for i, t in enumerate(tstamps):
            caiso_ws[f"{cPeriodStart}{i+2}"] = t

        caisoRows = df_caiso.shape[0]
        hrlyRows = len(tstamps) + 1

        eqn_mTs = f"=DATE(YEAR(C2),MONTH(C2),DAY(C2))+TIME(HOUR(C2),0,0)"
        eqn_pEnd = f"={cPeriodStart}2+TIME(1,0,0)"
        eqn_DOT = (
            f"=SUMIFS($H$2:$H${lastRow},"
            f'${cminuteTstamps}$2:${cminuteTstamps}${lastRow},">="&{cPeriodStart}2,'
            f'${cminuteTstamps}$2:${cminuteTstamps}${lastRow},"<"&{cPeriodEnd}2)/12'
        )
        eqn_SUPP = (
            f"=SUMIFS($N$2:$N${lastRow},"
            f'${cminuteTstamps}$2:${cminuteTstamps}${lastRow},">="&{cPeriodStart}2,'
            f'${cminuteTstamps}$2:${cminuteTstamps}${lastRow},"<"&{cPeriodEnd}2)/12'
        )

        caiso_ws[f"{cminuteTstamps}2"] = eqn_mTs
        caiso_ws[f"{cminuteTstamps}2"].number_format = "yyyy-mm-dd h:mm:ss;@"
        caiso_ws[f"{cPeriodEnd}2"] = eqn_pEnd
        caiso_ws[f"{cPeriodEnd}2"].number_format = "yyyy-mm-dd h:mm:ss;@"
        caiso_ws[f"{cDOT}2"] = eqn_DOT
        caiso_ws[f"{cSUPP}2"] = eqn_SUPP

        print_event("writing CAISO sheet formulas")

        for r in range(3, caisoRows + 2):
            caiso_ws[f"{cminuteTstamps}{r}"] = Translator(
                eqn_mTs, origin=f"{cminuteTstamps}2"
            ).translate_formula(f"{cminuteTstamps}{r}")
            caiso_ws[f"{cminuteTstamps}{r}"].number_format = "yyyy-mm-dd h:mm:ss;@"
            pbar.update(1)  # total: df_caiso.shape[0] - 1    [[[9k]]] -- scaling factor: 1

        for r in range(3, hrlyRows + 1):
            caiso_ws[f"{cPeriodEnd}{r}"] = Translator(
                eqn_pEnd, origin=f"{cPeriodEnd}2"
            ).translate_formula(f"{cPeriodEnd}{r}")
            caiso_ws[f"{cPeriodEnd}{r}"].number_format = "yyyy-mm-dd h:mm:ss;@"
            caiso_ws[f"{cDOT}{r}"] = Translator(eqn_DOT, origin=f"{cDOT}2").translate_formula(
                f"{cDOT}{r}"
            )
            caiso_ws[f"{cSUPP}{r}"] = Translator(eqn_SUPP, origin=f"{cSUPP}2").translate_formula(
                f"{cSUPP}{r}"
            )
            pbar.update(10)  # total: len(tstamps) - 1

        # on main sheet, overwrite curtSP column with caiso-related equation
        print_event("writing CAISO-specific curtSP equation")
        eqn_curtSP = (
            f"=IF(INDEX(CAISO!$AT$2:$AT${hrlyRows},MATCH(A{rStart},CAISO!$AQ$2:$AQ${caisoRows},1))>=0,"
            f"{sitecapacity_cell},INDEX(CAISO!$AS$2:$AS${hrlyRows},MATCH(A{rStart},CAISO!$AQ$2:$AQ${caisoRows},1)))"
        )

        ws[f"{cCurtSP.ltr}{rStart}"] = eqn_curtSP

        for r in row_range:
            ws[f"{cCurtSP.ltr}{r}"] = Translator(
                eqn_curtSP, origin=f"{cCurtSP.ltr}{rStart}"
            ).translate_formula(f"{cCurtSP.ltr}{r}")
            pbar.update(10)  # total: len(row_range)

        """add curtailment eq next to summary table on main sheet (before overwrite C7)"""
        # add sum of CrLoss col to D7 & move original eq from C7 to E7
        ws["D7"] = f"=SUM(${cCrLoss.ltr}${rStart}:${cCrLoss.ltr}${rEnd})"
        ws["E7"] = f"=SUM({cInv}7:{cInv_end}7)"

        """next to summary table, add count of instances when SUPP was negative"""
        ws["F6"] = "count (SUPP<0)"
        ws["F7"] = f'=COUNTIF(CAISO!N2:N{lastRow},"<0")'

        """SECOND CAISO SHEET"""  # curtailment report using 5-min data
        wb.create_sheet("Curtailment_ONW")
        ws3 = wb["Curtailment_ONW"]

        print_event("adding second CAISO curtailment sheet")
        for r in dataframe_to_rows(df_caiso2, index=True, header=False):
            ws3.append(r)

        # column reference vars for new sheet -- df has columns through caisoSP; rest are equations
        (
            ccDate,
            ccPOA,
            ccPVL,
            ccMeter,
            ccPPC,
            ccCaisoSP,
            ccSUPP,
            ccLostNRG,
            ccLostRev,
            ccNumCurt,
            ccRef,
        ) = "ABCDEFGHIJK"
        ws3[f"{ccDate}1"] = "Date"
        ws3[f"{ccPOA}1"] = "POA (Wh/m2)"
        ws3[f"{ccPVL}1"] = "PVLib (MW)"
        ws3[f"{ccMeter}1"] = "OE.MeterMW"
        ws3[f"{ccPPC}1"] = "PPC ($/kWh)"
        ws3[f"{ccCaisoSP}1"] = "CAISO_MW_setpoint"
        ws3[f"{ccSUPP}1"] = "SUPP"  # for reference/sanity checks when reviewing

        # columns w/ equations
        ws3[f"{ccLostNRG}1"] = "Curt. Loss (MWh)"
        ws3[f"{ccLostRev}1"] = "Lost Revenue ($)"
        ws3[f"{ccNumCurt}1"] = "# Curtailments"
        ws3[f"{ccRef}1"] = "Ref"

        # site capacity
        ws3["L1"] = "Capacity"
        ws3["L2"] = capacity
        cccCapacity = "$L$2"

        # equations for 5-minute data
        eq_lostNRGcurt = (
            f"=IF(AND({ccCaisoSP}2<{ccPVL}2,{ccCaisoSP}2<{cccCapacity},"
            f"({ccPVL}2-{ccMeter}2)>0),{ccPVL}2-{ccMeter}2,0)/12"
        )
        eq_lostRevenue = f"={ccLostNRG}2*{ccPPC}2*1000"
        eq_numCurt = f"=IF({ccLostNRG}2>0,1,0)"
        eq_refDay = f"=DATE({t_start.year},MONTH({ccDate}2),DAY({ccDate}2))"

        ws3[f"{ccLostNRG}2"] = eq_lostNRGcurt
        ws3[f"{ccLostRev}2"] = eq_lostRevenue
        ws3[f"{ccNumCurt}2"] = eq_numCurt
        ws3[f"{ccRef}2"] = eq_refDay

        for r in range(3, ws3.max_row + 1):
            ws3[f"{ccLostNRG}{r}"] = Translator(
                eq_lostNRGcurt, origin=f"{ccLostNRG}2"
            ).translate_formula(f"{ccLostNRG}{r}")
            ws3[f"{ccLostRev}{r}"] = Translator(
                eq_lostRevenue, origin=f"{ccLostRev}2"
            ).translate_formula(f"{ccLostRev}{r}")
            ws3[f"{ccNumCurt}{r}"] = Translator(
                eq_numCurt, origin=f"{ccNumCurt}2"
            ).translate_formula(f"{ccNumCurt}{r}")
            ws3[f"{ccRef}{r}"] = Translator(eq_refDay, origin=f"{ccRef}2").translate_formula(
                f"{ccRef}{r}"
            )

        # date range for summary table (use for # of rows to extend formulas)
        summary_range = pd.date_range(t_start, t_end, freq="D")[
            :-1
        ]  # exclude first day of next month
        ccSum, ccLostMWh, ccCurtHrs, ccRevLoss = "LMNO"
        ws3[f"{ccSum}1"] = "Sum"
        ws3[f"{ccLostMWh}1"] = "Curt. Loss (MWh)"
        ws3[f"{ccCurtHrs}1"] = "# Curtailment Hours (Events)"
        ws3[f"{ccRevLoss}1"] = "Lost Revenue ($)"

        # equations for daily summary table
        eq_lostMWh = f"=SUMIF(${ccRef}$2:${ccRef}${ws3.max_row},{ccSum}2,${ccLostNRG}$2:${ccLostNRG}${ws3.max_row})"
        eq_curtHrs = f"=SUMIF(${ccRef}$2:${ccRef}${ws3.max_row},{ccSum}2,${ccNumCurt}$2:${ccNumCurt}${ws3.max_row})"
        eq_revLoss = f"=SUMIFS(${ccLostRev}$2:${ccLostRev}${ws3.max_row},${ccRef}$2:${ccRef}${ws3.max_row},{ccSum}2)"

        ws3[f"{ccSum}2"] = f"={ccDate}2"
        ws3[f"{ccLostMWh}2"] = eq_lostMWh
        ws3[f"{ccCurtHrs}2"] = eq_curtHrs
        ws3[f"{ccRevLoss}2"] = eq_revLoss

        dayTable_endRow = len(summary_range) + 1  # plus header row
        for r in range(3, dayTable_endRow + 1):
            ws3[f"{ccSum}{r}"] = f"={ccSum}{r-1}+1"
            ws3[f"{ccLostMWh}{r}"] = Translator(
                eq_lostMWh, origin=f"{ccLostMWh}2"
            ).translate_formula(f"{ccLostMWh}{r}")
            ws3[f"{ccCurtHrs}{r}"] = Translator(
                eq_curtHrs, origin=f"{ccCurtHrs}2"
            ).translate_formula(f"{ccCurtHrs}{r}")
            ws3[f"{ccRevLoss}{r}"] = Translator(
                eq_revLoss, origin=f"{ccRevLoss}2"
            ).translate_formula(f"{ccRevLoss}{r}")

        # totals row
        rTotals = dayTable_endRow + 1
        ws3[f"{ccSum}{rTotals}"] = "Total"
        for col in [ccLostMWh, ccCurtHrs, ccRevLoss]:
            ws3[f"{col}{rTotals}"] = f"=SUM({col}2:{col}{dayTable_endRow})"

        """overwrite curtailment in main ws condition 2"""
        # add sum of CrLoss col to D7 & move original eq from C7 to E7
        ws["C7"] = f"=Curtailment_ONW!${ccLostMWh}${rTotals}"

    pbar.update(150)

    """ADD SUMMARY SHEET - FOR GADS PURPOSES"""
    print_event("creating GADS summary sheet")
    wb.create_sheet("Summary")
    sum_ws = wb["Summary"]

    # similar summary block as on FlashReport sheet
    sRef = "FlashReport!"

    numInvCell = f"{sRef}{cInv}1"

    seqn_SIHD = f'=COUNTIF({sRef}{cAvail}{rStart}:{cAvail_end}{rEnd},">0")'
    seqn_RSIH = f'=COUNTIF({sRef}{cCrLoss.ltr}{rStart}:{cCrLoss.ltr}{rEnd},">0")*{numInvCell}'
    seqn_FOIHD = f"=SUMPRODUCT(({sRef}{cPOA.ltr}{rStart}:{cPOA.ltr}{rEnd}>0)*({sRef}{cInv}{rStart}:{cInv_end}{rEnd}<=0))"
    seqn_PRatio = f"=ROUND(SUM({sRef}{cMtr.ltr}{rStart}:{cMtr.ltr}{rEnd})/SUM({sRef}{cENDCi.ltr}{rStart}:{cENDCi.ltr}{rEnd}), 2)"

    fcell = lambda addr: sRef + addr
    fcells_ = {
        "total_loss": fcell("C3"),
        "dc_health": fcell("C4"),
        "availability": fcell("B5"),
        "downtime": fcell("C6"),
        "curtailment": fcell("C7"),
        "gross": fcell("C8"),
        "pvlib": fcell("C9"),
        "net": fcell("C10"),
    }

    n_month_hours = round(pd.Timedelta(t_end - t_start).total_seconds() / 60)
    n_month_hours = df4xl.shape[0]
    n_calendar_hours = n_month_hours * numInv

    seqn_GAG = (
        f'=IF({fcells_["gross"]}>(1.15*{fcells_["net"]}), 1.15*{fcells_["net"]},'
        f'IF({fcells_["gross"]}<{fcells_["net"]}, {fcells_["net"]},'
        f'{fcells_["gross"]}))'
    )
    seqn_NAG = f'=ROUND({fcells_["net"]}, 2)'
    seqn_NMC = f'=ROUND({fcell("B12")}, 2)'  # site capacity
    seqn_MPOA = f'=ROUND({fcell("B13")}, 2)'  # insolation
    seqn_EG = f'=ROUND({fcells_["pvlib"]}, 2)'  # possible

    seqn_ASIH = "=B18+B19+B20+B21+B22"  # SIHD + RSIH + FOIHD + MIHD + PIHD
    seqn_AIH = "=B35-B15-B16-B17"  #  month_hours - IRIH - MBIH - RIH
    seqn_RUIHN = "=B14-B13"  # AIH - ASIH
    # seqn_RUIHN = '=B35-B15-B16-B17-B13'    # month_hours - IRIH - MBIH - RIH - ASIH

    gads_id_dict = {
        "Comanche": {
            "entity": "NCR11665",
            "plant": "1000281",
            "inv_group": "2000636",
        },
        "GA4": {
            "entity": "NCR12008",
            "plant": "1000283",
            "inv_group": "2000637",
        },
        "Maplewood 1": {
            "entity": "NCR12162",
            "plant": "1000282",
            "inv_group": "2000634",
        },
        "Maplewood 2": {
            "entity": "NCR12162",
            "plant": "1000282",
            "inv_group": "2000634",
        },
    }
    gads_id = lambda key_: (
        "n/a" if (sitename not in gads_id_dict) else gads_id_dict[sitename].get(key_)
    )

    summary_sheet_vals = [
        [f"Metrics - {sitename}", "Value", "GADS Label"],  # 1
        ["Total Energetic Loss [MWh]", f'={fcells_["total_loss"]}', None],  # 2
        ["DC / System Health [MWh]", f'={fcells_["dc_health"]}', None],  # 3
        ["Inverter Availability [%]", f'={fcells_["availability"]}', None],  # 4
        ["Downtime [MWh]", f'={fcells_["downtime"]}', None],  # 5
        ["Curtailment [MWh]", f'={fcells_["curtailment"]}', None],  # 6
        ["Gross Generation [MWh]", seqn_GAG, "Gross Actual Generation (GAG)"],  # 7
        ["Net Generation [MWh]", seqn_NAG, "Net Actual Generation (NAG)"],  # 8
        ["Site Capacity [MW]", seqn_NMC, "Net Maximum Capacity (NMC)"],  # 9
        ["Insolation [kWh/m^2]", seqn_MPOA, "Monthly Plane of Array (MPOA)"],  # 10
        ["Performance Ratio [%]", seqn_PRatio, "Performance Ratio"],  # 11
        ["Possible Generation [MWh]", seqn_EG, "Expected Generation (EG)"],  # 12
        [None, seqn_ASIH, "Active Solar Inverter Hours (ASIH)"],  # 13
        [None, seqn_AIH, "Active Inverter Hours (AIH)"],  # 14
        [None, 0, "Inactive Reserve Inverter Hours (IRIH)"],  # 15
        [None, 0, "Mothballed Inverter Hours (MBIH)"],  # 16
        [None, 0, "Retired Unit Inverter Hours (RIH)"],  # 17
        [None, seqn_SIHD, "Service Inverter Hours Day (SIHD)"],  # 18
        [None, seqn_RSIH, "Reserve Shutdown Inverter Hours (RSIH)"],  # 19
        [None, seqn_FOIHD, "Forced Outage Inverter Hours Day (FOIHD)"],  # 20
        [None, 0, "Maintenance Inverter Hours Day (MIHD)"],  # 21
        [None, 0, "Planned Inverter Hours Day (PIHD)"],  # 22
        [None, 0, "Resource Unavailable Inverter Hours - Day (RUIHD)"],  # 23
        [None, 0, "Service Inverter Hours Night (SIHN)"],  # 24
        [None, 0, "Forced Outage Inverter Hours Night (FOIHN)"],  # 25
        [None, 0, "Maintenance Inverter Hours Night (MiHN)"],  # 26
        [None, 0, "Planned Inverter Hours Night (PIHN)"],  # 27
        [None, seqn_RUIHN, "Resource Unavailable Inverter Hours - Night (RUIHN)"],  # 28
        [None, gads_id("entity"), "Entity ID"],  # 29
        [None, gads_id("plant"), "Plant ID"],  # 30
        [None, gads_id("inv_goup"), "Inverter Group ID"],  # 31
        [None, t_start.month, "Reporting Period (month)"],  # 32
        [None, t_start.year, "Reporting Year"],  # 33
        [None, "AC", "Inverter Group Availability Status"],  # 34
        [None, n_calendar_hours, "Calendar Hours"],  # 35
    ]
    for i, list_ in enumerate(summary_sheet_vals):
        sum_ws[f"A{i+1}"] = list_[0]
        sum_ws[f"B{i+1}"] = list_[1]
        sum_ws[f"C{i+1}"] = list_[2]

    pbar.update(50)

    ## add checks (to the right of table -- see Comanche 202409 FlashReport file)

    """CELL FORMATTING"""
    print_event("defining format variables/functions")

    # define formatting functions
    def fmt_cell(ws, cell, fill=None, align=None, font=None, numFmt=None, border=None):
        """ARGS:
        ws       = openpyxl worksheet object
        cell     = cell address string
        fill     = cell fill format       (optional)
        align    = cell alignment format  (optional)
        font     = cell font format       (optional)
        numFmt   = cell number format     (optional)
        border   = cell border format     (optional)
        """
        c = ws[cell]
        if border:
            c.border = border
        if fill:
            c.fill = fill
        if align:
            c.alignment = align
        if font:
            c.font = font
        if numFmt:
            c.number_format = numFmt

    def fmt_range(
        ws,
        rng,
        fill=None,
        align=None,
        font=None,
        numFmt=None,
        border=None,
        excCols=None,
        dcols=None,
        dFill=None,
    ):
        """ARGS:
        ws       = openpyxl worksheet object
        rng      = cell range address string
        fill     = cell fill format       (optional)
        align    = cell alignment format  (optional)
        font     = cell font format       (optional)
        numFmt   = cell number format     (optional)
        border   = cell border format     (optional)
        excCols  = list of columns to exclude from number formatting   (optional)
        dcols    = list of column numbers with different fill formats  (optional)
        dFill    = dictionary of fill formats corresponding to dcols   (optional)
        """
        for row in ws[rng]:
            for c in row:
                col = c.column
                if border:
                    c.border = border
                if align:
                    c.alignment = align
                if font:
                    c.font = font
                if numFmt:
                    cond = excCols is None
                    if cond or ((not cond) and col not in excCols):
                        c.number_format = numFmt
                if fill:
                    if dcols is not None and col in dcols:
                        c.fill = dFill[col]
                    else:
                        c.fill = fill

    # merge cells in sheet title block (top left)
    ws.merge_cells("B1:C1")
    ws.merge_cells("A1:A2")

    """FORMAT TITLE/SUMMARY BLOCK, TOP-LEFT OF SHEET"""  # pbar +100
    print_event("applying cell formatting")
    fmt_range(ws, "B3:B10", fill=f.grey2, border=b.bd, numFmt="0.0%")
    fmt_range(ws, "C3:C10", fill=f.grey2, border=b.bd, numFmt="#,##0.00")
    fmt_range(ws, "A3:A10", fill=f.grey3, align=a.right, border=b.bd_mRt)
    fmt_range(ws, "B1:C1", fill=f.grey4, align=a.center, font=fnt.bold11, border=b.bd)
    fmt_range(ws, "B2:C2", fill=f.grey3, align=a.right, font=fnt.bold11, border=b.bd_mBtm)
    fmt_range(ws, "A1:A2", fill=f.grey5, align=a.center2, font=fnt.bold14, border=b.bd_mRt_mBtm)
    fmt_range(ws, "A11:A15", fill=f.grey1, align=a.right, border=b.bd)
    fmt_range(ws, "B11:B15", fill=f.grey1, align=a.right, numFmt="0.00", border=b.bd)
    fmt_cell(ws, "B15", numFmt="0.00%")
    fmt_cell(ws, "A10", border=b.bd_mRt_mBtm)
    fmt_cell(ws, "B10", border=b.bd_mBtm)
    fmt_range(ws, "C1:C10", border=b.bd_mRt)
    fmt_cell(ws, "C2", border=b.bd_mRt_mBtm)
    fmt_cell(ws, "C10", border=b.bd_mRt_mBtm)
    pbar.update(100)

    """FORMAT ROWS B/W INVERTERS TOTAL SECTION & INV DATA TABLE"""  # pbar +100
    print_event("formatting inverter summary table")
    fmt_cell(ws, f"{cMtr.ltr}1", align=a.right, font=fnt.ital9)
    fmt_cell(ws, numInv_cell, align=a.center, font=fnt.ital9)
    fmt_range(ws, f"{cMtr.ltr}{r2b4hdr}:{cInv_end}{r1b4hdr}", align=a.right, font=fnt.ital9)
    fmt_range(ws, f"{cInv}{r2b4hdr}:{cInv_end}{r2b4hdr}", numFmt="0.0%")

    fmt_range(ws, f"{cInv}10:{cInv_end}{adjRatio_row}", fill=f.grey1, font=fnt.sz9, border=b.bd1)
    fmt_range(ws, f"{cInv}10:{cInv_end}11", numFmt="0.00")
    fmt_range(ws, f"{cInv}11:{cInv_end}11", border=b.bd1_dbBtm2)
    fmt_range(ws, f"{cInv}{RAWratio_row}:{cInv_end}{adjRatio_row}", numFmt="0.0000")

    fmt_range(ws, f"{cInv}{adjRatio_row}:{cInv_end}{adjRatio_row}", border=b.bd1_m1Btm)
    fmt_range(
        ws,
        f"{cMtr.ltr}10:{cMtr.ltr}{adjRatio_row}",
        fill=f.grey1a,
        align=a.right,
        font=fnt.sz9,
        border=b.bd2_m2Rt,
    )
    fmt_cell(ws, f"{cMtr.ltr}11", border=b.bd2_m2Rt_dbBtm2)
    fmt_cell(ws, f"{cMtr.ltr}{adjRatio_row}", border=b.bd2_m2Rt_m1Btm)

    fmt_range(
        ws, f"{cCurtSP.ltr}{RAWratio_row}:{cCurtSP.ltr}{r1b4hdr}", align=a.right, font=fnt.sz9
    )
    fmt_range(
        ws,
        f"{cFlag.ltr}{RAWratio_row}:{cFlag.ltr}{r1b4hdr}",
        align=a.center,
        font=fnt.sz9,
        numFmt="0.0",
    )
    fmt_cell(ws, f"{cFlag.ltr}{r1b4hdr}", numFmt="0%")
    fmt_cell(ws, f"{cFlag.ltr}{RAWratio_row}", numFmt="0.000")
    fmt_range(ws, f"{cCurtSP.ltr}{RAWratio_row}:{cInv_end}{RAWratio_row}", font=fnt.ital9)
    pbar.update(100)

    fmt_cell(ws, f"{cTcell.ltr}{r2b4hdr}", align=a.right, font=fnt.sz10)
    fmt_cell(ws, f"{cENDCi.ltr}{r2b4hdr}", font=fnt.sz10, numFmt="0.000")

    """FORMAT INVERTER TOTALS SECTION"""  # pbar +100
    fmt_range(ws, f"{cInv}3:{cInv_end}9", fill=f.grey2, numFmt="0.00", border=b.bd)
    fmt_range(ws, f"{cInv}5:{cInv_end}5", numFmt="0.0%")
    fmt_range(ws, f"{cInv}2:{cInv_end}2", fill=f.grey3, align=a.right, border=b.bd_mBtm)
    fmt_range(ws, f"{cFlag.ltr}3:{cFlag.ltr}9", align=a.right, font=fnt.ital9)
    fmt_range(ws, f"{cMtr.ltr}3:{cMtr.ltr}9", fill=f.grey3, align=a.right, border=b.bd_mRt)
    fmt_cell(
        ws, f"{cMtr.ltr}2", fill=f.grey4, align=a.center, font=fnt.bold11, border=b.bd_mRt_mBtm
    )
    fmt_range(ws, f"{cMtr.ltr}9:{cInv_end}9", border=b.bd_dbBtm)
    fmt_cell(ws, f"{cMtr.ltr}9", border=b.bd_mRt_dbBtm)
    pbar.update(100)

    """FORMAT INVERTER DATA TABLE"""  # pbar +300
    print_event("formatting main data table")
    fmt_range(ws, f"{cInv}{rStart}:{cInv_end}{rEnd}", fill=f.grey1, numFmt="0.000", border=b.bd)
    fmt_range(
        ws, f"{cInv}{rHeader}:{cInv_end}{rHeader}", fill=f.grey2, align=a.right, border=b.bd_mBtm
    )
    fmt_range(ws, f"{cInv_calc}{rStart}:{cInv_calc}{rEnd}", align=a.center, numFmt="0.00")
    fmt_cell(ws, f"{cInv_calc}{rHeader}", align=a.center)
    pbar.update(300)

    """FORMAT PVLIB DATA TABLE"""  # pbar +300
    print_event("formatting PVLib data table")
    fmt_range(
        ws,
        f"{cPVLib}{rStart}:{cPVLib_end}{rEnd}",
        fill=f.grey1,
        align=a.right,
        numFmt="0.000",
        border=b.bd,
    )
    fmt_range(
        ws, f"{cPVLib}{rHeader}:{cPVLib_end}{rHeader}", fill=f.grey2, align=a.left, border=b.bd_mBtm
    )
    fmt_cell(ws, f"{cPVLib}{r1b4hdr}", font=fnt.bold11)
    pbar.update(300)

    """FORMAT TIME-BASED AVAILABILITY TABLE"""  # pbar +300
    print_event("formatting availability table")
    fmt_range(
        ws,
        f"{cAvail}{rStart}:{cAvail_end}{rEnd}",
        fill=f.grey1,
        align=a.right,
        numFmt="0",
        border=b.bd,
    )
    fmt_range(
        ws, f"{cAvail}{rHeader}:{cAvail_end}{rHeader}", fill=f.grey2, align=a.left, border=b.bd_mBtm
    )
    fmt_cell(ws, f"{cAvail}{r1b4hdr}", font=fnt.bold11)
    pbar.update(300)

    """FORMAT ADJUSTED PVLIB TABLE"""  # pbar +300
    print_event("formatting adj. PVLib table")
    fmt_range(
        ws,
        f"{cPVLadj}{rStart}:{cPVLadj_end}{rEnd}",
        fill=f.grey1,
        align=a.right,
        numFmt="0.000",
        border=b.bd,
    )
    fmt_range(
        ws,
        f"{cPVLadj}{rHeader}:{cPVLadj_end}{rHeader}",
        fill=f.grey2,
        align=a.left,
        border=b.bd_mBtm,
    )
    fmt_cell(ws, f"{cPVLadj}{r1b4hdr}", font=fnt.bold11)
    pbar.update(300)

    """FORMAT LOST MWh TABLE"""  # pbar +300
    print_event("formatting lost MWh table")
    fmt_range(
        ws,
        f"{cMWhLost}{rStart}:{cMWhLost_end}{rEnd}",
        fill=f.grey1,
        align=a.right,
        numFmt="0.000",
        border=b.bd,
    )
    fmt_range(
        ws,
        f"{cMWhLost}{rHeader}:{cMWhLost_end}{rHeader}",
        fill=f.grey2,
        align=a.left,
        border=b.bd_mBtm,
    )
    fmt_cell(ws, f"{cMWhLost}{r1b4hdr}", font=fnt.bold11)
    pbar.update(300)

    """REFORMAT ACTUAL/POSSIBLE/METER COLUMNS WITH CORRESPONDING COLORS"""  # pbar +400
    print_event("color coding data table columns")
    dcol_list = [cAct.num, cPoss.num, cMtr.num]
    dictFill_lt = {cAct.num: f.blue1, cPoss.num: f.grey3, cMtr.num: f.rust1}
    dictFill_rg = {cAct.num: f.blue2, cPoss.num: f.blk, cMtr.num: f.rust2}
    fmt_range(
        ws,
        f"{cTime.ltr}{rStart}:{cMtr.ltr}{rEnd}",
        fill=f.grey1,
        numFmt="0.000",
        border=b.bd,
        excCols=[cTime.num, cPOAb.num, cFlag.num],
        dcols=dcol_list,
        dFill=dictFill_lt,
    )
    fmt_range(
        ws,
        f"{cTime.ltr}{rHeader}:{cMtr.ltr}{rHeader}",
        fill=f.grey2,
        align=a.center,
        border=b.bd_mBtm,
        dcols=dcol_list,
        dFill=dictFill_rg,
    )
    fmt_cell(ws, f"{cPOAb.ltr}{r1b4hdr}", align=a.left, font=fnt.ital9)
    pbar.update(300)

    """FORMAT MISC"""  # pbar +200
    print_event("fixing table borders and things..")
    fmt_range(
        ws, f"{cMtr.ltr}{rStart}:{cMtr.ltr}{rEnd}", border=b.bd_mRt
    )  # overwrite borders in meter column
    fmt_range(
        ws, f"{cAct.ltr}{rHeader}:{cPoss.ltr}{rHeader}", font=fnt.white
    )  # white font, Actual/Poss header cells
    fmt_cell(
        ws, f"{cMtr.ltr}{rHeader}", border=b.bd_mRt_mBtm, font=fnt.white
    )  # white font, Meter header cell
    pbar.update(100)

    """format 2nd caiso sheet"""
    if isinstance(df_caiso2, pd.DataFrame):
        fmt_range(
            ws3, f"{ccDate}2:{ccDate}{ws3.max_row}", numFmt="yyyy-mm-dd hh:mm:ss", border=b.bd
        )
        fmt_range(ws3, f"{ccPOA}2:{ccMeter}{ws3.max_row}", numFmt="0.000", border=b.bd)
        fmt_range(ws3, f"{ccPPC}2:{ccPPC}{ws3.max_row}", numFmt="0.0000", border=b.bd)
        fmt_range(ws3, f"{ccCaisoSP}2:{ccNumCurt}{ws3.max_row}", border=b.bd)
        fmt_range(ws3, f"{ccRef}2:{ccRef}{ws3.max_row}", numFmt="yyyy-mm-dd")
        fmt_range(ws3, f"{ccRef}1:K1", align=a.right)
        fmt_range(
            ws3,
            f"{ccDate}1:{ccNumCurt}1",
            align=Alignment(wrap_text=True, horizontal="right"),
            fill=f.bluuu,
            font=fnt.white,
            border=b.bd,
        )
        fmt_range(ws3, f"{ccSum}2:{ccSum}{dayTable_endRow}", numFmt="yyyy-mm-dd", border=b.bd_lr)
        fmt_range(
            ws3,
            f"{ccLostMWh}2:{ccCurtHrs}{dayTable_endRow+1}",
            numFmt="0.00",
            align=a.right,
            border=b.bd_lr,
        )
        fmt_range(
            ws3,
            f"{ccRevLoss}2:{ccRevLoss}{dayTable_endRow+1}",
            numFmt="$#,##0.00",
            align=a.right,
            border=b.bd_lr,
        )
        fmt_range(
            ws3,
            f"{ccSum}1:{ccRevLoss}1",
            align=Alignment(wrap_text=True, horizontal="right"),
            fill=f.grey2,
            border=b.bd,
        )
        fmt_range(
            ws3,
            f"{ccSum}{dayTable_endRow+1}:{ccRevLoss}{dayTable_endRow+1}",
            fill=f.grey1,
            border=b.bd,
        )

        # column widths
        ws3.column_dimensions[ccDate].width = 22
        ws3.column_dimensions["F"].width = 12
        for c in "GHI":
            ws3.column_dimensions[c].width = 16
        for c in "BCDEJLMNO":
            ws3.column_dimensions[c].width = 14

    """FORMAT SUMMARY SHEET"""
    fmt_range(sum_ws, f"A1:C{sum_ws.max_row}", fill=f.grey1, border=b.bd)
    fmt_range(sum_ws, "A1:B1", align=a.center, fill=f.grey2, border=b.bd_mBtm)
    fmt_cell(sum_ws, "C1", align=a.indent, fill=f.grey2, border=b.bd_mBtm)
    fmt_range(sum_ws, f"A2:B{sum_ws.max_row}", align=a.right)
    fmt_range(sum_ws, "B2:B12", numFmt="0.00")
    fmt_cell(sum_ws, "B4", numFmt="0.00%")
    fmt_cell(sum_ws, "B11", numFmt="0.00%")
    fmt_range(sum_ws, "B13:B28", numFmt="0")
    fmt_cell(sum_ws, "B35", numFmt="0")
    fmt_range(sum_ws, f"C2:C{sum_ws.max_row}", align=a.indent)

    # column widths
    sum_ws.column_dimensions["A"].width = 26
    sum_ws.column_dimensions["B"].width = 12
    sum_ws.column_dimensions["C"].width = 50

    """CHANGE COLUMN WIDTHS (main sheet)"""
    # change column widths for defined instances of Col class (via 'width' attribute)
    for col in Col_list:
        ws.column_dimensions[col.ltr].width = col.width

    """CREATE CHARTS"""
    print_event("creating charts")

    # create chart series formatting functions
    def fmt_series_markers(series, symbol=None, size=None, color=None, line_color=None):
        if symbol:
            series.marker.symbol = symbol
        if size:
            series.marker.size = size
        if color:
            series.marker.graphicalProperties.solidFill = color
        if line_color:
            series.marker.graphicalProperties.line.solidFill = line_color

    def fmt_series_lines(series, style=None, width=None, color=None, noFill=None):
        if style:
            series.graphicalProperties.line.dashStyle = style
        if width:
            series.graphicalProperties.line.width = width
        if color:
            series.graphicalProperties.line.solidFill = color
        if noFill:
            series.graphicalProperties.line.noFill = noFill

    def fmt_series_bars(series, fill_type, color1, pType=None, color2=None):
        if fill_type == "solid":
            series.graphicalProperties.solidFill = color1
        if fill_type == "pattern":
            fill = PatternFillProperties(prst=pType)
            fill.foreground = ColorChoice(srgbClr=color1)
            fill.background = ColorChoice(srgbClr=color2)
            series.graphicalProperties.pattFill = fill

    def fmt_chart(
        chart,
        width,
        height,
        title=None,
        c_type=None,
        grouping=None,
        overlap=None,
        gap_width=None,
        legend_pos=None,
        yaxis_title=None,
        yaxis_min=None,
        yaxis_max=None,
        yaxis_nFmt=None,
        yaxis_axId=None,
        xaxis_title=None,
        xaxis_min=None,
        xaxis_max=None,
        xaxis_nFmt=None,
    ):
        chart.width = width
        chart.height = height
        if title:
            chart.title = title
        if c_type:
            chart.type = c_type
        if grouping:
            chart.grouping = grouping
        if overlap:
            chart.overlap = overlap
        if gap_width:
            chart.gapWidth = gap_width
        if legend_pos:
            chart.legend.position = legend_pos
        if yaxis_title:
            chart.y_axis.title = yaxis_title
        if yaxis_min:
            chart.y_axis.scaling.min = yaxis_min
        if yaxis_max:
            chart.y_axis.scaling.max = yaxis_max
        if yaxis_nFmt:
            chart.y_axis.number_format = yaxis_nFmt
        if yaxis_axId:
            chart.y_axis.axId = yaxis_axId
        if xaxis_title:
            chart.x_axis.title = xaxis_title
        if xaxis_min:
            chart.x_axis.scaling.min = xaxis_min
        if xaxis_max:
            chart.x_axis.scaling.max = xaxis_max
        if xaxis_nFmt:
            chart.x_axis.number_format = xaxis_nFmt

    """CREATE SCATTER CHART"""
    print_event("adding scatter plot")
    cht = ScatterChart()
    xvals = Reference(ws, min_col=cPOA.num, min_row=rStart, max_row=rEnd)

    # SCATTER, possible power
    yvals_poss = Reference(ws, min_col=cPoss.num, min_row=rHeader, max_row=rEnd)
    s_possible = Series(yvals_poss, xvals, title_from_data=True)
    cht.series.append(s_possible)

    # SCATTER, meter power
    yvals_meter = Reference(ws, min_col=cMtr.num, min_row=rHeader, max_row=rEnd)
    s_meter = Series(yvals_meter, xvals, title_from_data=True)
    cht.series.append(s_meter)

    # SCATTER, actual power
    yvals_actual = Reference(ws, min_col=cAct.num, min_row=rHeader, max_row=rEnd)
    s_actual = Series(yvals_actual, xvals, title_from_data=True)
    cht.series.append(s_actual)

    # SCATTER series formatting
    fmt_series_markers(s_actual, symbol="triangle", size=3, color=cc.blue2, line_color=cc.blue2)
    fmt_series_markers(s_possible, symbol="square", size=3, color=cc.blk, line_color=cc.blk)
    fmt_series_markers(s_meter, symbol="circle", size=3, color=cc.rust2, line_color=cc.rust2)
    fmt_series_lines(s_actual, noFill=True)
    fmt_series_lines(s_possible, noFill=True)
    fmt_series_lines(s_meter, noFill=True)

    # scatter chart formatting & add to sheet
    fmt_chart(
        cht,
        width=20,
        height=12,
        title="Active Power vs. POA",
        legend_pos="t",
        yaxis_title="Active Power [MWh]",
        yaxis_min=0,
        yaxis_nFmt="0",
        xaxis_title="POA [W/m^2]",
        xaxis_min=0,
        xaxis_nFmt="0",
    )
    ws.add_chart(cht, "F22")
    pbar.update(200)

    """CREATE STACKED BAR/LINE COMBO CHART"""
    print_event("adding bar chart")
    cht2 = BarChart()
    cht2_ = LineChart()
    cht3 = LineChart()
    xvals2 = Reference(ws, min_col=cInv_num, max_col=cInv_end_num, min_row=2)

    # BAR, actual generation (row8)
    yvals_actualGen = Reference(ws, min_col=cMtr.num, max_col=cInv_end_num, min_row=8)
    cht2.add_data(yvals_actualGen, titles_from_data=True, from_rows=True)

    # BAR, downtime (row6)
    yvals_downtime = Reference(ws, min_col=cMtr.num, max_col=cInv_end_num, min_row=6)
    cht2.add_data(yvals_downtime, titles_from_data=True, from_rows=True)

    # BAR, dc/system health (row4)
    yvals_dcHealth = Reference(ws, min_col=cMtr.num, max_col=cInv_end_num, min_row=4)
    cht2.add_data(yvals_dcHealth, titles_from_data=True, from_rows=True)
    pbar.update(200)

    # BAR series variables
    s_actGen = cht2.series[0]
    s_downtime = cht2.series[1]
    s_dcHealth = cht2.series[2]

    # BAR series formatting
    fmt_series_bars(s_actGen, "solid", color1=cc.blue2)
    fmt_series_bars(s_downtime, "solid", color1=cc.red)
    fmt_series_bars(s_dcHealth, "pattern", color1=cc.red, pType="ltUpDiag", color2=cc.wht)

    # LINE, curtailment (row7)
    yvals_curtail = Reference(ws, min_col=cMtr.num, max_col=cInv_end_num, min_row=7)
    cht2_.add_data(yvals_curtail, titles_from_data=True, from_rows=True)

    # LINE, inverter availability (row5)
    yvals_invAvail = Reference(ws, min_col=cMtr.num, max_col=cInv_end_num, min_row=5)
    cht3.add_data(yvals_invAvail, titles_from_data=True, from_rows=True)

    # LINE series variables
    s_curt = cht2_.series[0]
    s_avail = cht3.series[0]

    # LINE series formatting
    fmt_series_markers(s_curt, symbol="dot", size=2, color=cc.yllw, line_color=cc.yllw)
    fmt_series_markers(s_avail, symbol="circle", size=4, color=cc.blk, line_color=cc.blk)
    fmt_series_lines(s_curt, style="solid", width=20000, color=cc.yllw)
    fmt_series_lines(s_avail, style="dash", width=22250, color=cc.blk)
    pbar.update(100)

    # add data labels for inverter availability series
    s_avail.dLbls = DataLabelList()
    s_avail.dLbls.showVal = True
    s_avail.dLbls.dLblPos = "t"

    # stacked bar/line chart formatting
    fmt_chart(
        cht2,
        width=16,
        height=10,
        c_type="col",
        grouping="stacked",
        overlap=100,
        gap_width=80,
        legend_pos="t",
        yaxis_title="Power (MW)",
        yaxis_nFmt="0.0",
        xaxis_title="Inverter ID",
    )
    fmt_chart(
        cht2_,
        width=16,
        height=10,
        grouping="stacked",
        overlap=100,
        legend_pos="t",
        yaxis_title="Curtailment (MWh)",
        yaxis_nFmt="0.00",
        yaxis_axId=200,
        yaxis_max=0.5 * inv_maxMWh,
    )
    fmt_chart(
        cht3,
        width=16,
        height=10,
        grouping="stacked",
        overlap=100,
        legend_pos="t",
        yaxis_title="Inverter Availability",
        yaxis_nFmt="0%",
        yaxis_axId=300,
        yaxis_min=0.8,
    )
    pbar.update(200)

    # set categories & combine charts
    cht2.set_categories(xvals2)
    cht2_.set_categories(xvals2)
    cht3.set_categories(xvals2)
    cht2_.y_axis.crosses = "max"
    cht2 += cht2_
    cht3.y_axis.crosses = "max"
    cht2 += cht3

    # add combined chart to sheet & set dimensions based on # of inverters
    ws.add_chart(cht2, "F47")
    wdth = numInv * 1.7
    fmt_chart(cht2, width=wdth, height=12)
    pbar.update(100)

    """ADD IMAGE OF DATA GAPS"""
    print_event("adding data gap image")
    # save msno figure to temporary directory for adding to excel file
    fig_copy = fig.get_figure()
    tmpdir = tempfile.TemporaryDirectory()
    fig_path = f"{tmpdir.name}\\tempfig.png"
    fig_copy.savefig(fig_path, bbox_inches="tight")
    plt.close(fig_copy)
    # close/hide temporary figure

    img = Image(fig_path)
    img.width = img.width / 2.4
    img.height = img.height / 2.4
    img.anchor = "S22"
    ws.add_image(img)
    ws.sheet_view.zoomScale = 90

    ##testing
    print_event("saving initial workbook to tempdir...")
    with tempfile.TemporaryDirectory() as temp_folder:
        temp_fpath = Path(temp_folder, "temp.xlsx")
        wb.save(filename=temp_fpath)
        wb.close()

        print_event("calculating formulas...")
        # pythoncom.CoInitialize()
        xw_app = xw.App(visible=False)
        wb2 = xw_app.books.open(temp_fpath)
        xw_app.calculate()

        print_event("saving workbook...")
        wb2.save()
        wb2.close()
        xw_app.quit()
        # pythoncom.CoUninitialize()

        shutil.copy2(temp_fpath, savepath)

    # print_event('saving workbook...')
    # wb.save(filename=savepath)

    # print_event('calculating formulas...')
    # pythoncom.CoInitialize()
    # xw_app = xw.App(visible=False)
    # wb2 = xw_app.books.open(savepath)
    # xw_app.calculate()

    # print_event('saving workbook...')
    # wb2.save(savepath)
    # wb2.close()
    # xw_app.quit()
    # pythoncom.CoUninitialize()

    """SAVE FILE"""
    # print_event('saving workbook...')
    # wb.save(filename=savepath)
    pbar.update(110)  # update last portion of progress bar
    pbar.close()  # close progress bar
    tmpdir.cleanup()  # remove temporary directory w/ msno figure

    disppath = "\\".join(savepath.split("\\")[-6:])
    if not q:
        print(f'Report successfully generated for {sitename}!\n"..{disppath}"')

    return
