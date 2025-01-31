import openpyxl
import numbers
import sys, clr
import numpy as np
import pandas as pd
import datetime as dt
from pathlib import Path

from oetoolbox.utils import oemeta, oepaths
from oetoolbox.dataquery import pireference as ref

sys.path.append(r"C:\Program Files (x86)\PIPC\AF\PublicAssemblies\4.0")
clr.AddReference("OSIsoft.AFSDK")
import OSIsoft
from OSIsoft.AF import PISystems
from OSIsoft.AF.Search import AFEventFrameSearch
from OSIsoft.AF.Asset import AFAttribute, AFAttributeList, AFSearchMode
from OSIsoft.AF.PI import PIPageType, PIPagingConfiguration, PIPoint, PIPointList, PIServers
from OSIsoft.AF.Time import AFTime, AFTimeRange, AFTimeSpan, AFTimeZone, AFTimeZoneFormatProvider
from OSIsoft.AF.Data import (
    AFBoundaryType,
    AFCalculationBasis,
    AFSummaryTypes,
    AFTimestampCalculation,
)


def segmented_date_ranges(start, end, n_days):
    t0, t1 = start, start + pd.DateOffset(days=n_days)
    if t1 > end:
        return [(start, end)]
    daterange_list = []  # init
    while t1 <= end:
        daterange_list.append((t0, t1))
        if t1 == end:
            break
        t0 = t1
        t1 = t1 + pd.DateOffset(days=n_days)
        if t1 > end:
            t1 = end
    return daterange_list


def attPath_name(site, attPath):
    if f"{site}|" in attPath:
        return attPath.split("|")[-1]
    return "_".join(
        list(reversed(attPath.split(f"{site}\\")[-1].replace("|", "\\").split("\\")[1:]))
    )


## misc query parameters
aftimezone = AFTimeZone().CurrentAFTimeZone
aftz_fmtprovider = AFTimeZoneFormatProvider(aftimezone)
summaryType = AFSummaryTypes.Average
summaryType2 = AFSummaryTypes.Range
calcBasis = AFCalculationBasis.TimeWeighted
calcBasisDiscrete = AFCalculationBasis.TimeWeightedDiscrete
tstamp_calc = AFTimestampCalculation.Auto
config = PIPagingConfiguration(PIPageType.TagCount, 100000)


piserver = PIServers().DefaultPIServer


def pipoint_list(pipoint_names):
    pipointList = PIPointList()
    for pipt in pipoint_names:
        pipointList.Add(PIPoint.FindPIPoint(piserver, pipt))
    return pipointList


def af_attribute_list(attPath_list):
    AFserver = OSIsoft.AF.PISystems().DefaultPISystem
    DB = AFserver.Databases.get_Item("Onward Energy")
    attributeList = AFAttributeList()
    for attPath in attPath_list:
        try:
            attributeList.Add(AFAttribute.FindAttribute(attPath, DB))
        except:
            continue
    return attributeList


def pipoints_from_attPaths(attPath_list):
    AFserver = OSIsoft.AF.PISystems().DefaultPISystem
    DB = AFserver.Databases.get_Item("Onward Energy")
    pipointList = PIPointList()
    for attPath in attPath_list:
        att_ = AFAttribute.FindAttribute(attPath, DB)
        pipointList.Add(att_.PIPoint)
    return pipointList


def pipoints_from_af_atts(attributeList):
    pipointList = PIPointList()
    for att in attributeList:
        pipointList.Add(att.PIPoint)
    return pipointList


# function used to format df from output of summaries method
def format_pi_dataframe(df_, tz, freq):
    df = df_.copy().apply(pd.to_numeric, errors="coerce")
    df.index = pd.to_datetime(df.index.astype(str), utc=True)
    df = df.tz_convert(tz=tz).tz_localize(None)
    start, end = df.index.min(), df.index.max()
    ref_idx = pd.date_range(start, end, freq=freq.replace("m", "min"))

    if df.index.duplicated().any():
        df = df.loc[~df.index.duplicated(keep="first")]

    if df.shape[0] != ref_idx.shape[0]:
        df = df.reindex(ref_idx)

    df = df.rename_axis("Timestamp")
    return df


nowtime = lambda: dt.datetime.now().strftime("%H:%M:%S")

## functions to handle attPaths & get column name(s) when writing query values to dataframe
contents_ = lambda attPath: attPath.split("\\")
get_site = lambda attPath: [x for x in contents_(attPath) if x in ref.all_pi_sites][0]
site_index = lambda attPath: contents_(attPath).index(get_site(attPath))
asset_grp = lambda attPath: contents_(attPath)[site_index(attPath) + 1]
att_level = lambda attPath: len(contents_(attPath)[site_index(attPath) + 1 : -1])
att_struct = lambda attPath: contents_(attPath.replace("|", "\\"))[site_index(attPath) + 1 :]


# function to get att_struct from afvals (for query below)
def get_att_struct(afvals, site):
    struct_ = [afvals.Attribute.Name]
    ele = afvals.Attribute.Element
    while ele.Name != site:
        struct_.append(ele.Name)
        ele = ele.Parent
    return struct_


long_format_cols = ["Group_ID", "Asset_ID", "Subasset_ID", "Subasset2_ID", "Subasset3_ID"]


def query_pi_data(
    site,
    start,
    end,
    freq,
    pipoints=None,
    attPath_list=None,
    calc_basis=None,
    format_=None,
    col_name_source=None,
    subrange_size=None,
    keeptzinfo=False,
    q=True,
):
    qprint = lambda msg, end="\n": None if q else print(msg, end=end)
    if (pipoints is None) and (attPath_list is None):
        print("need either pipoints or attributes")
        return
    querying_pipts = pipoints is not None
    if querying_pipts:
        tag_names = pipoints
        n_items = len(pipoints)
        df_format = "wide"
        qprint(f"{nowtime()}: Generating PIPointList..", end=" ")
        data_object = pipoint_list(pipoints)
        qprint("done!")

    else:
        tag_names = [attPath_name(site, p) for p in attPath_list]
        n_items = len(attPath_list)
        qprint(f"{nowtime()}: Generating AFAttributeList..", end=" ")
        data_object = af_attribute_list(attPath_list).Data
        qprint("done!")

        # determine default format for output data (wide/long)
        if any(f"{site}|" in p for p in attPath_list):  # if any site-level attributes in list
            df_format = "wide"  # use wide format
        else:
            all_att_levels = [att_level(p) for p in attPath_list]
            all_lvls_below_2 = all((lvl < 2) for lvl in all_att_levels)
            df_format = "wide" if (all_lvls_below_2 and n_items < 100) else "long"

        if df_format == "long":
            # check asset groups (if all same, omits group name from columns)
            all_asset_groups = [asset_grp(p) for p in attPath_list]
            all_same_group = len(list(set(all_asset_groups))) == 1

    # override dataframe format (if arg specified)
    if format_ in ["wide", "long"]:
        df_format = format_

    is_long = df_format == "long"  # variable for use in loop below

    ## function to get column name from afvals (depending on format & col_name_source arg)
    if col_name_source in ["PIPoints", "Attributes"]:
        cols_from_pipts = col_name_source == "PIPoints"
    else:
        cols_from_pipts = querying_pipts

    pd_freq = freq.replace("m", "min")  # if min, otherwise h
    if not any(w.isdigit() for w in pd_freq):  # e.g. if ='h'
        pd_freq = f"1{pd_freq}"  #  change to '1h'
    freq_mins = int(pd.Timedelta(pd_freq).seconds / 60)

    ## define query parameters
    tz = oemeta.data["TZ"].get(site)
    timeSpan = AFTimeSpan.Parse(freq)
    summaryType = AFSummaryTypes.Average
    calcBasis = AFCalculationBasis.TimeWeighted  # default
    if calc_basis in ["discrete", "event"]:
        calcBasis = {
            "discrete": AFCalculationBasis.TimeWeightedDiscrete,
            "event": AFCalculationBasis.EventWeighted,
        }.get(calc_basis)
    tstamp_calc = AFTimestampCalculation.Auto
    config = PIPagingConfiguration(PIPageType.TagCount, 1000)
    query_parameters = [timeSpan, summaryType, calcBasis, tstamp_calc, config]

    ## for bulk queries, create grouped sub-dateranges (effective if range longer than 40 days) - reduced to 10 days if len(tags) > 50
    n_days = (pd.Timestamp(end) - pd.Timestamp(start)).days
    if isinstance(subrange_size, int):
        if 0 < subrange_size < n_days:
            n_days = subrange_size

    ## override if conditions met
    if (n_days > 31) and (n_items > 50) and (freq_mins < 15):
        n_days = 11

    grouped_date_ranges = segmented_date_ranges(pd.Timestamp(start), pd.Timestamp(end), n_days)
    n_segments = len(grouped_date_ranges)

    q_startTime = dt.datetime.now()
    main_df_list = []
    for i, sub_range in enumerate(grouped_date_ranges):
        msg_ = (
            f"sub-range {i+1} of {n_segments}"
            if (n_segments > 1)
            else f"full date range ({n_days = })"
        )
        qprint(f"{nowtime()}: Querying {msg_}")
        start_, end_ = sub_range
        t0, t1 = pd.Timestamp(start_, tz=tz), pd.Timestamp(end_, tz=tz)
        timeRange = AFTimeRange(str(t0), str(t1), aftz_fmtprovider)
        summaries = data_object.Summaries(timeRange, *query_parameters)

        df_list = []
        omit_list = []
        starttime = nowtime()
        status_symbols = [chr(n) for n in (92, 124, 47, 45)]
        for i, query_vals in enumerate(summaries):
            msg_ = f"{starttime}: Retrieving values.. att {i+1}/{n_items}"
            symb_ = status_symbols[(i % 4)] if ((i + 1) != n_items) else " "
            qprint(f"{msg_}  {symb_}", end="\r" if ((i + 1) != n_items) else "\n")
            afvals = query_vals[summaryType]
            values, tstamps, flags = afvals.GetValueArrays()

            bad_cond1 = (len(list(values)) == 1) and (list(flags)[0].ToString() == "Bad")
            bad_cond2 = not any(isinstance(x, numbers.Number) for x in list(values))
            if bad_cond1 or bad_cond2:
                bad_data_name = tag_names[i]
                omit_list.append(bad_data_name)
                continue

            ## long format
            if is_long:
                dfa = pd.DataFrame({"Value": list(values)}, index=list(tstamps))
                if cols_from_pipts:
                    dfa["PIPoint"] = afvals.PIPoint.Name
                else:
                    dfa["Attribute"] = afvals.Attribute.Name
                    # print(afvals.Attribute.Name)
                    att_struct = get_att_struct(afvals, site)[::-1]  # reverse, asset group first
                    col_names = long_format_cols[: len(att_struct) - 1]
                    if all_same_group:
                        del att_struct[0]
                        del col_names[0]

                    for col_, val_ in zip(col_names, att_struct):
                        dfa[col_] = val_

            ## wide format
            else:
                col_ = tag_names[i]
                dfa = pd.DataFrame({col_: list(values)}, index=list(tstamps))

            ## add dataframe to list
            df_list.append(dfa)

        if not df_list:
            qprint("!! ERROR !! - no data found for range!\n")
            continue
            # return pd.DataFrame()

        ## in case any pi tags were not found
        if (i + 1) != n_items:
            qprint("")

        if omit_list and (not q):
            print("The following pipoints/attributes were omitted due to bad data:")
            for omit_name in omit_list:
                print(f"    {omit_name}")

        ## format data & localize timezone/reindex for DST if needed
        qprint(f"{nowtime()}: Writing to dataframe..", end=" ")

        join_axis = 0 if is_long else 1
        df_ = pd.concat(df_list, axis=join_axis)
        df_.index = pd.to_datetime(df_.index.astype(str), utc=True)
        df_ = df_.tz_convert(tz=tz)
        if not keeptzinfo:
            df_ = df_.tz_localize(None)
        df_ = df_.rename_axis("Timestamp")

        value_cols = [
            c
            for c in df_.columns
            if (c not in long_format_cols) and (c not in ["Attribute", "PIPoint"])
        ]
        df_[value_cols] = df_[value_cols].apply(pd.to_numeric, errors="coerce")

        if not is_long:
            expected_idx = pd.date_range(
                df_.index.min(), df_.index.max(), freq=freq.replace("m", "min")
            )

            if not keeptzinfo:
                if df_.index.duplicated().any():
                    df_ = df_.loc[~df_.index.duplicated(keep="first")]

            if df_.shape[0] != expected_idx.shape[0]:
                df_ = df_.reindex(expected_idx)
                qprint("(note: reindexed)")

            for col in omit_list:
                df_[col] = np.nan

            df_ = df_[list(sorted(df_.columns))]

        qprint("done!")
        main_df_list.append(df_.copy())

    if n_segments == 1:
        df = main_df_list[0]
    else:
        df = pd.concat(main_df_list, axis=0, ignore_index=False)

    return df


## developed for oakfield wtg icing request ### WIP as of 2025/01/14 -Pete
def query_pi_recordedvalues(
    site, start, end, attPath_list=None, pipoint_list=None, subrange_size=None, q=True
):
    qprint = lambda msg, end="\n": None if q else print(msg, end=end)

    if (attPath_list is not None) and (pipoint_list is None):
        asset_att_list = [p.split("\\")[-1].split("|") for p in attPath_list]
        qprint(f"{nowtime()}: Generating PIPointList..", end=" ")
        pipointList = pipoints_from_attPaths(attPath_list)
        qprint("done!\n")
    elif (pipoint_list is not None) and (attPath_list is None):
        pipointList = pipoint_list
        asset_att_list = None
    elif (pipoint_list is not None) and (attPath_list is not None):
        print(f"Both pipoint_list and attPath_list provided - defaulting to pipoint_list")
        pipointList = pipoint_list
        asset_att_list = None
    elif not pipoint_list and not attPath_list:
        print(f"Error! Need either pipoint_list or attPath_list as input")

    print(f"pipointList: {pipointList}")
    n_items = len(pipointList)
    ## define query parameters
    tz = oemeta.data["TZ"].get(site)
    aftimezone = AFTimeZone().CurrentAFTimeZone
    aftz_fmtprovider = AFTimeZoneFormatProvider(aftimezone)
    boundaryType = AFBoundaryType.Interpolated
    pageType = PIPagingConfiguration(PIPageType.TagCount, 1000)

    n_days = (pd.Timestamp(end) - pd.Timestamp(start)).days
    if isinstance(subrange_size, int):
        if 0 < subrange_size < n_days:
            n_days = subrange_size
    grouped_date_ranges = segmented_date_ranges(pd.Timestamp(start), pd.Timestamp(end), n_days)
    n_segments = len(grouped_date_ranges)

    main_df_list = []
    for i, sub_range in enumerate(grouped_date_ranges):
        msg_ = (
            f"sub-range {i+1} of {n_segments}:"
            if (n_segments > 1)
            else f"full date range ({n_days = })"
        )

        qprint(f"Querying {msg_}")
        start_, end_ = sub_range
        t0, t1 = pd.Timestamp(start_, tz=tz), pd.Timestamp(end_, tz=tz)
        timeRange = AFTimeRange(str(t0), str(t1), aftz_fmtprovider)
        query_output = pipointList.RecordedValues(timeRange, boundaryType, "", False, pageType)

        df_list = []
        starttime = nowtime()
        status_symbols = [chr(n) for n in (92, 124, 47, 45)]
        for i, afvals in enumerate(query_output):
            msg_ = f"{starttime}: Retrieving values.. att {i+1}/{n_items}"
            symb_ = status_symbols[(i % 4)] if ((i + 1) != n_items) else " "
            qprint(f"{msg_}  {symb_}", end="\r" if ((i + 1) != n_items) else "\n")
            if asset_att_list is not None:
                asset_id, att_name = asset_att_list[i]
            else:
                att_name = pipointList[i]
                asset_id = str(i)
            values, tstamps, flags = afvals.GetValueArrays()
            df_list.append(
                pd.DataFrame(
                    {"Value": list(values), "Attribute": att_name, "Asset_ID": asset_id},
                    index=list(tstamps),
                )
            )

        ## format data & localize timezone/reindex for DST if needed
        qprint(f"{nowtime()}: Writing to dataframe..", end=" ")
        df_ = pd.concat(df_list, axis=0)
        df_.index = pd.to_datetime(df_.index.astype(str), utc=True)
        df_ = df_.tz_convert(tz=tz).tz_localize(None)
        df_ = df_.rename_axis("Timestamp")

        qprint("done!\n")
        main_df_list.append(df_.copy())

    df = pd.concat(main_df_list, axis=0, ignore_index=False)
    df["Value"] = pd.to_numeric(df["Value"], errors="coerce")
    return df


# returns wide format dataframe
def query_pipoints(site, start, end, freq, pipoints, dev=False):
    tz = oemeta.data["TZ"].get(site)
    t0, t1 = pd.Timestamp(start, tz=tz), pd.Timestamp(end, tz=tz)

    # define query parameters
    timeRange = AFTimeRange(str(t0), str(t1), aftz_fmtprovider)
    timeSpan = AFTimeSpan.Parse(freq)
    summaryType = AFSummaryTypes.Average
    calcBasis = AFCalculationBasis.TimeWeighted
    if dev:
        calcBasis = AFCalculationBasis.TimeWeightedDiscrete
    tstamp_calc = AFTimestampCalculation.Auto
    config = PIPagingConfiguration(PIPageType.TagCount, 100000)
    query_parameters = [timeRange, timeSpan, summaryType, calcBasis, tstamp_calc, config]

    # get pipoints & query values using "Summaries" method
    pipointList = pipoint_list(pipoints)
    summaries = pipointList.Summaries(*query_parameters)

    df_list = []
    for query_vals in summaries:
        afvals = query_vals[summaryType]
        values, tstamps, flags = afvals.GetValueArrays()
        dfa = pd.DataFrame({afvals.PIPoint.Name: list(values)}, index=list(tstamps))
        df_list.append(dfa)

    df_ = pd.concat(df_list, axis=1)
    df = format_pi_dataframe(df_, tz, freq)
    return df


# basic query, single pipoint
def query_pipoint(site, start, end, freq, pipoint, dev=False):
    tz = oemeta.data["TZ"].get(site)
    t0, t1 = pd.Timestamp(start, tz=tz), pd.Timestamp(end, tz=tz)

    # define query parameters
    timeRange = AFTimeRange(str(t0), str(t1), aftz_fmtprovider)
    timeSpan = AFTimeSpan.Parse(freq)
    summaryType = AFSummaryTypes.Average
    calcBasis = AFCalculationBasis.TimeWeighted
    if dev:
        calcBasis = AFCalculationBasis.TimeWeightedDiscrete
    tstamp_calc = AFTimestampCalculation.Auto
    config = PIPagingConfiguration(PIPageType.TagCount, 100000)
    query_parameters = [timeRange, timeSpan, summaryType, calcBasis, tstamp_calc, config]

    # get pipoint & query values using "Summaries" method
    pipt = PIPoint.FindPIPoint(piserver, pipoint)
    query_output = pipt.Summaries(*query_parameters)
    afvals = query_output[summaryType]
    values, tstamps, flags = afvals.GetValueArrays()
    df_ = pd.DataFrame(index=list(tstamps), data=list(values), columns=[pipoint])
    df = format_pi_dataframe(df_, tz, freq)
    return df


# function to build dictionary of attribute paths from monthly query atts in pireference pqmeta
solarAFpath = "\\\\CORP-PISQLAF\\Onward Energy\\Renewable Fleet\\Solar Assets"


def get_query_attribute_paths(site):
    site_dict = ref.pqmeta.get(site)
    siteAFpath = f"{solarAFpath}\\{site}"

    # inverters
    inv_att_paths = []
    inverters_ = oemeta.data["AF_Solar_V3"][site].get("Inverters")
    if inverters_:
        inv_names = list(inverters_["Inverters_Assets"].keys())
        invPath = f"{siteAFpath}\\Inverters"
        inv_att_paths = [f"{invPath}\\{inv}|OE.ActivePower" for inv in inv_names]

    # combiners
    cmb_att_paths = []
    cmb_atts = site_dict.get("inv atts")
    cmb_sub_atts = site_dict.get("inv sub atts")
    if cmb_atts:
        for inv in inv_names:
            cmb_att_paths.extend([f"{invPath}\\{inv}|{att}" for att in cmb_atts])
    elif cmb_sub_atts:
        inv1 = inv_names[0]  # get cmb names from first inv (same for all)
        inv1_dict = inverters_["Inverters_Assets"].get(inv1)
        cmb_names = list(inv1_dict[f"{inv1}_Subassets"].keys())
        if site == "Imperial Valley":  # temp
            cmb_names = [c for c in cmb_names if c != "Combiner Boxes"]  # temp
        for inv in inv_names:
            for cmb in cmb_names:
                cmb_att_paths.extend([f"{invPath}\\{inv}\\{cmb}|{att}" for att in cmb_sub_atts])

    # met stations
    met_att_paths = []
    met_atts = site_dict.get("metsta atts")
    metPath = f"{siteAFpath}\\Met Stations"
    if met_atts:
        if isinstance(met_atts, dict):
            for met, att_list in met_atts.items():
                met_att_paths.extend([f"{metPath}\\{met}|{att}" for att in att_list])
        else:
            if "met ids" in site_dict:
                met_names = site_dict["met ids"]
            else:
                metstations_ = oemeta.data["AF_Solar_V3"][site].get("Met Stations")
                met_names = list(metstations_["Met Stations_Assets"].keys())

            for met in met_names:
                met_att_paths.extend([f"{metPath}\\{met}|{att}" for att in met_atts])

    # ppc
    ppc_atts = site_dict.get("ppc atts")
    ppc_att_paths = [f"{siteAFpath}\\PPC|{att}" for att in ppc_atts] if ppc_atts else []

    # trackers
    trk_att_paths = []
    trk_atts = site_dict.get("trk atts")
    trk_sub_atts = site_dict.get("trk sub atts")
    trkPath = f"{siteAFpath}\\Trackers"
    trackers_ = oemeta.data["AF_Solar_V3"][site].get("Trackers")
    trk_names = list(trackers_["Trackers_Assets"].keys()) if trackers_ else []
    if trk_atts:
        for trk in trk_names:
            t_atts = trk_atts.get(trk) if isinstance(trk_atts, dict) else trk_atts
            trk_att_paths.extend([f"{trkPath}\\{trk}|{att}" for att in t_atts])
    elif trk_sub_atts:
        trk1 = trk_names[0]
        trk1_dict = trackers_["Trackers_Assets"].get(trk1)
        mtr_names = list(trk1_dict[f"{trk1}_Subassets"].keys())
        for trk in trk_names:
            for mtr in mtr_names:
                trk_att_paths.extend([f"{trkPath}\\{trk}\\{mtr}|{att}" for att in trk_sub_atts])

    # meter
    meter_att_paths = []
    meter_ = oemeta.data["AF_Solar_V3"][site].get("Meter")
    all_meter_atts = meter_["Meter_Attributes"] if meter_ else []
    if "OE_MeterMW" in all_meter_atts:
        meter_att_paths.append(f"{siteAFpath}\\Meter|OE_MeterMW")

    ##temp
    if site == "GA4":
        meter_att_paths = [f"{siteAFpath}\\Meter|SEL3530.MTR_FSREV1.Sts.P_MW"]

    groups_ = ["Inverters", "Combiners", "Met Stations", "PPC", "Trackers", "Meter"]
    attpaths_ = [
        inv_att_paths,
        cmb_att_paths,
        met_att_paths,
        ppc_att_paths,
        trk_att_paths,
        meter_att_paths,
    ]
    output_dict = {grp: atts for grp, atts in zip(groups_, attpaths_) if atts}

    return output_dict


## sites with defined monthly query attributes in file "pireference.py"
pireference_sites = list(ref.pqmeta.keys())


def get_pqmeta_att_paths(sitelist=None):
    if sitelist is None:
        sitelist = pireference_sites
    else:
        sitelist = [s for s in sitelist if s in pireference_sites]

    attPath_dict = {site: get_query_attribute_paths(site) for site in sitelist}
    return attPath_dict


now = lambda: dt.datetime.now().strftime("%H:%M:%S")

## NEW MONTHLY QUERY SCRIPT
default_freq_dict = {
    "Inverters": "1m",
    "Combiners": "1h",
    "Met Stations": "1m",
    "Trackers": "1h",
    "PPC": "1m",
    "Meter": "1m",
}
valid_freq = lambda frq: (frq[:-1].isnumeric()) and (frq[-1] in ["h", "m"])


def monthly_pi_query(
    site,
    year,
    month,
    query_groups=None,
    override_subrange=None,
    save_files=True,
    local=False,
    overwrite=False,
    q=True,
):
    """ """
    qprint = lambda msg, end="\n": None if q else print(msg, end=end)

    ## get query attribute paths (from pireference)
    attPath_dict = get_query_attribute_paths(site)
    available_query_groups = list(attPath_dict.keys())

    asset_groups = available_query_groups.copy()
    invalid_groups = []
    if query_groups is not None:
        asset_groups = []
        for g in query_groups:
            asset_groups.append(g) if g in available_query_groups else invalid_groups.append(g)

    if "Trackers" in asset_groups:
        asset_groups.remove("Trackers")  # temp

    qprint(f"{site.upper()}\n>> asset groups: {asset_groups}")
    if invalid_groups:
        qprint(f"note: no attributes found for the following query groups: {invalid_groups}")

    ## get start/end time for query
    start = pd.Timestamp(year=year, month=month, day=1)
    end = start + pd.DateOffset(months=1)

    ## get savepath info
    def get_savepath(asset_group):
        if not local:
            sfolder_ = Path(oepaths.frpath(year, month, ext="Solar"), site)
            if not sfolder_.exists():
                sfolder_.mkdir(parents=True)
        else:
            sfolder_ = Path(Path.home(), "Downloads")

        stem_ = f'PIQuery_{asset_group.replace(" ", "")}_{site}_{year}-{month:02d}'
        n_, fname_ = 1, f"{stem_}.csv"

        if not Path(sfolder_, fname_).exists():
            savepath = Path(sfolder_, fname_)
        elif not overwrite:
            while Path(sfolder_, fname_).exists():
                fname_ = f"{stem_} ({n_}).csv"
                n_ += 1
            savepath = Path(sfolder_, fname_)
        else:
            related_files = list(sfolder_.glob(f"{stem_}*.csv"))
            for fp in related_files:
                fp.unlink()
            savepath = Path(sfolder_, fname_)

        return savepath

    ## run query script for selected asset groups
    output_data = {}
    for grp in asset_groups:
        qprint(f"\n{grp.upper()}")
        attPaths = attPath_dict[grp]  # get attPath list
        freq = default_freq_dict[grp]  # get freq/interval
        if isinstance(query_groups, dict):
            if valid_freq(query_groups[grp]):  # only overwrite if valid
                freq = query_groups[grp]

        kwargs_ = dict(attPath_list=attPaths, q=q)
        if isinstance(override_subrange, int):
            kwargs_.update(dict(subrange_size=override_subrange))
        elif grp in ["Combiners", "Trackers"]:
            kwargs_.update(dict(subrange_size=4))
        elif grp == "Met Stations":
            kwargs_.update(dict(subrange_size=16))
        elif grp == "Inverters":
            if len(attPaths) > 50:
                size_ = 3
            else:
                size_ = 7
            kwargs_.update(dict(subrange_size=size_))

        ## query data
        try:
            df = query_pi_data(site, start, end, freq, **kwargs_)
        except:
            qprint("!! FAILED !! continuing..")
            continue

        if df.empty:
            qprint("!! ERROR !! no data returned by query; continuing..")
            continue

        ## reformat (if long data)
        id_cols = [c for c in df.columns if c.endswith("_ID")]
        if id_cols:
            df = df.reset_index(drop=False)
            index_cols = ["Timestamp"] + id_cols
            df = df.pivot(index=index_cols, columns="Attribute", values="Value")
            n_lvls = df.index.nlevels
            reset_lvls = list(range(1, n_lvls))
            df = df.reset_index(level=reset_lvls).rename_axis(None, axis=1)
            sort_cols = id_cols[::-1] + ["Timestamp"]
            df = df.sort_values(by=sort_cols)

        ## add dataframe to output dictionary

        ## temp
        if (site == "GA4") and (grp == "Meter"):
            df = df.div(-1e6)
            df.columns = ["OE_MeterMW"]

        output_data[grp] = df.copy()

        ## save file
        if save_files:
            spath = get_savepath(grp)
            df.to_csv(spath)
            qprint(f'>> saved data file: "{spath.name}"')

    return output_data


def run_monthly_pi_query(
    year, month, sitelist=None, querygroups=None, local=False, firstpass=False, q=True
):
    """ """
    qprint = lambda msg, end="\n": None if q else print(msg, end=end)
    query_sites = (
        pireference_sites if (sitelist is None) else [s for s in sitelist if s in pireference_sites]
    )
    if not query_sites:
        print(f"No valid sites specified in sitelist.\nExiting..")
        return

    for i, site in enumerate(query_sites):

        query_groups = None
        existing_pq_files = list(
            Path(oepaths.frpath(year, month, ext="Solar"), site).glob("PIQuery*.csv")
        )
        existing_qgroups = None
        if firstpass and existing_pq_files:
            existing_qgroups = list(set(fp.name.split("_")[1] for fp in existing_pq_files))
            if "MetStations" in existing_qgroups:
                existing_qgroups.remove("MetStations")
                existing_qgroups.append("Met Stations")
            query_groups = [g for g in get_query_attribute_paths(site) if g not in existing_qgroups]

        if existing_qgroups and (not query_groups):
            qprint(f">> skipping {site} (firstpass=True; existing groups already queried).")
            continue

        if isinstance(querygroups, list):
            querygroups = [g for g in querygroups if g in get_query_attribute_paths(site)]
            if querygroups:
                query_groups = querygroups

        qprint(f"\n**[{now()}]** - START QUERY (site {i+1} of {len(query_sites)}) - ", end="")
        output_data = monthly_pi_query(
            site, year, month, query_groups=query_groups, local=local, q=q
        )
        qprint(f"\n**[{now()}]** - END QUERY - {site}", end=" ")

        if existing_qgroups:
            qprint("(firstpass=True; skipped existing groups)")
        qprint("\n")

    qprint("\n!! DONE !!\n")
    return


def load_PI_table(table_name):
    afserver = OSIsoft.AF.PISystems().DefaultPISystem
    DB = afserver.Databases.get_Item("Onward Energy")
    table = DB.Tables[table_name].Table
    cols = [c.Caption for c in table.Columns]
    return pd.DataFrame([[rw[i] for i in range(len(cols))] for rw in table.Rows], columns=cols)


logfilepath_ = Path(oepaths.solar, "Data", "PI AF", "query_optimization", "query_log.xlsx")


def log_query_stats(query_startTime, site, tag_names, t0, t1, freq_mins, n_grp_days, n_na_total):
    ## query execution times
    q_start = pd.Timestamp(query_startTime)
    q_end = pd.Timestamp(dt.datetime.now())
    q_duration = (q_end - q_start).total_seconds()

    ## query params/metrics
    n_tags = len(tag_names)
    total_days = (t1 - t0).days
    n_daily_events = int(24 * (60 / freq_mins))
    n_total_events = n_tags * total_days * n_daily_events

    # segmented_ranges = segmented_date_ranges(t0, t1, n_grp_days)
    # n_segments = len(segmented_ranges)
    n_segments = 1  # temp
    n_events_per_segment = int(n_total_events / n_segments)

    ## add to log file
    q_times = [str(q_start), str(q_end), q_duration]
    q_params = [str(tag_names), n_tags, freq_mins, n_daily_events, str(t0.date()), str(t1.date())]
    q_stats = [total_days, n_total_events, n_segments, n_events_per_segment, n_na_total]
    log_vals = [site] + q_times + q_params + q_stats

    ## update log file
    wb = openpyxl.load_workbook(logfilepath_)
    ws = wb["Sheet1"]
    # col_names = [ws.cell(1, i+1).value for i in range(ws.max_column)]
    next_row = ws.max_row + 1
    for i, val in enumerate(log_vals):
        col_number = i + 1
        cell_ = ws.cell(next_row, col_number)
        cell_.value = val
        if col_number in [2, 3]:
            cell_.number_format = "yyyy-mm-dd hh:mm:ss"
        elif col_number in [9, 10]:
            cell_.number_format = "yyyy-mm-dd"
        elif col_number == 4:
            cell_.number_format = "0.00"
        elif col_number in [6, 7, 8, 11, 12, 13, 14, 15]:
            cell_.number_format = "0"
    wb.save(logfilepath_)
    wb.close()

    return log_vals


"""
PI Event Frame Search
"""


def load_pi_events(start, end, sitelist, q=True):
    """Load all PI event frames for given site that occur within the specified start/end dates

    Parameters
    ----------
    start : str
        Start date for search range*, format: "YYYY-mm-dd" (+optional: "HH:MM:SS")
    end : str
        End date for search range*, format: "YYYY-mm-dd" (+optional: "HH:MM:SS")
    sitelist : list of str
        List of PI site name(s) to include in search.
    q : bool, default True
        A "quiet" parameter to enable/disable status printouts (enabled if q=False)

    *note: using "overlapped" search mode, which includes all objects whose time range
           overlaps with the specified range at any point in time (AFSearchMode.Overlapped)

    Returns
    -------
    pd.DataFrame
        A pandas dataframe object where each row is an individual event. Additional rows are
        provided for effective start/end times, which account for the potion of the event
        that occurred within the specified search range (e.g. for events overlapping start/end)

    Example
    -------
    >>> df_events = load_pi_events(start="2024-11-01", end="2024-11-15", sitelist=["CID"])
    >>> df_events.head()

    <class 'pandas.core.frame.DataFrame'>
    RangeIndex: 16 entries, 0 to 15
    Data columns (total 12 columns):
     #   Column              Non-Null Count  Dtype
    ---  ------              --------------  -----
     0   Site                16 non-null     object
     1   EventStart          16 non-null     datetime64[ns]
     2   EventEnd            16 non-null     datetime64[ns]
     3   EventDuration       16 non-null     timedelta64[ns]
     4   Category            16 non-null     object
     5   Description         16 non-null     object
     6   PrimaryElement      16 non-null     object
     7   existing_event      16 non-null     bool
     8   ongoing_event       16 non-null     bool
     9   effective_start     16 non-null     datetime64[ns]
     10  effective_end       16 non-null     datetime64[ns]
     11  effective_duration  16 non-null     timedelta64[ns]
    dtypes: bool(2), datetime64[ns](4), object(4), timedelta64[ns](2)
    memory usage: 1.4+ KB

    """
    qprint = lambda str_, end="\n": None if q else print(str_, end=end)
    db = PISystems().DefaultPISystem.Databases.get_Item("Onward Energy")
    startTime, endTime = AFTime(start), AFTime(end)
    searchMode = AFSearchMode.Overlapped

    # using AFEventFrameSearch Constructor (AFDatabase, String, AFSearchMode, AFTime, AFTime, String)
    args_ = [db, "", searchMode, startTime, endTime]  # last arg = query string (site-specific)

    data_cols = [
        "Site",
        "EventStart",
        "EventEnd",
        "EventDuration",
        "Category",
        "Description",
        "PrimaryElement",
    ]
    data_list = []  # init
    for site in sitelist:
        site_data = []  # init, only for printout/status purposes
        qprint(site.ljust(22), end="")
        query_string = f"Name:{site}*"
        search_ = AFEventFrameSearch(*args_, query_string)
        for eFrame_ in search_.FindObjects():
            ele_obj = eFrame_.PrimaryReferencedElement
            ref_element = ele_obj.Name if (ele_obj is not None) else ""
            site_data.append(
                [
                    site,
                    str(eFrame_.StartTime),
                    str(eFrame_.EndTime),
                    str(eFrame_.Duration),
                    eFrame_.CategoriesString,
                    eFrame_.Name,
                    ref_element,
                ]
            )
        qprint(f"events: {len(site_data)}")
        data_list.extend(site_data)

    if len(data_list) < 1:
        qprint("!! no events found !!")
        return pd.DataFrame()

    df = pd.DataFrame(data_list, columns=data_cols)
    for col in ["EventStart", "EventEnd"]:
        df[col] = pd.to_datetime(df[col], format="%m/%d/%Y %I:%M:%S %p")

    df["EventDuration"] = df["EventDuration"].apply(pd.Timedelta)
    start_date, end_date = map(pd.Timestamp, [start, end])
    df["existing_event"] = df["EventStart"].lt(start_date)
    df["ongoing_event"] = df["EventEnd"].gt(end_date)
    df["effective_start"] = df["EventStart"].mask(df["existing_event"], start_date)
    df["effective_end"] = df["EventEnd"].mask(df["ongoing_event"], end_date)
    df["effective_duration"] = df["effective_end"] - df["effective_start"]

    qprint("\nEnd of search.")
    return df
