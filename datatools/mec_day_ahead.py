import datetime as dt
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from pathlib import Path

from ..dataquery.external import query_DTN
from ..utils.datetime import remove_tzinfo_and_standardize_index
from ..utils.helpers import quiet_print_function
from ..utils import oepaths

PRECIP_TYPE_MAPPING = {
    "1": "snow",
    "2": "rain",
    "3": "rain and snow",
    "4": "freezing rain",
    "5": "sleet",
    "6": "drizzle",
    "7": "freezing drizzle",
    "8": "thunderstorms",
}
DTN_WEATHER_FIELDS = [
    "airTemp",
    "feelsLikeTemp",
    "dewPoint",
    "relativeHumidity",
    "precipAmount",
    "precipProb",  # note: only for forecasted data
    "precipType",
]


def _get_dtn_query_args(start=None, end=None, n_days=5, q=True):
    if n_days < 1:
        raise ValueError("Value for n_days must be greater than 1.")
    fields = DTN_WEATHER_FIELDS
    today = pd.Timestamp(dt.datetime.now().date())
    if start is None and end is None:
        start = today + pd.Timedelta(days=1)
        end = start + pd.Timedelta(days=n_days)
        return (start, end, fields)
    if start is not None:
        start = pd.Timestamp(start)
        if start < today:
            if not q:
                print(
                    "\nNote: provided start time is earlier than today's date."
                    "\nSome/all data in output will be historical instead of predicted."
                    "\nAdditionally, the precipProb column will be omitted."
                )
            fields.remove("precipProb")
        if end is None:
            end = start + pd.Timedelta(days=n_days)
            return (start, end, fields)
        end = pd.Timestamp(end)
        if end < start:
            start, end = end, start
        return (start, end, fields)
    raise Exception("Invalid arguments.")


# function to get forecasted weather data from DTN
def get_dtn_weather_forecast(lat, lon, tz, start=None, end=None, n_days=5, q=True):
    """Queries data from DTN and formats precipType field using defined mapping."""
    t_start, t_end, fields = _get_dtn_query_args(start, end, n_days, q=q)
    df = query_DTN(lat, lon, t_start, t_end, interval="", fields=fields, tz=tz, q=q)
    df["precipType"] = df["precipType"].astype(int).astype(str).map(PRECIP_TYPE_MAPPING)
    df = df.rename(columns={"airTemp": "temp", "relativeHumidity": "humidity"})
    df = remove_tzinfo_and_standardize_index(df)
    # convert air temp from c to f
    df["temp"] = df["temp"].mul(9 / 5).add(32)
    return df


# supporting functions
def yInterpolate(x, xvals, yvals, n):
    if len(xvals) < n or len(yvals) < n:
        return -999
    for i in range(n - 1):
        if xvals[i] > xvals[i + 1]:
            return -992
    if x <= xvals[0]:
        if (xvals[1] - xvals[0]) != 0:
            k = (yvals[1] - yvals[0]) / (xvals[1] - xvals[0])
            return yvals[0] + (x - xvals[0]) * k
        else:
            return xvals[0]
    if x >= xvals[n - 1]:
        if (xvals[n - 1] - xvals[n - 2]) != 0:
            k = (yvals[n - 1] - yvals[n - 2]) / (xvals[n - 1] - xvals[n - 2])
            return yvals[n - 1] + (x - xvals[n - 1]) * k
        else:
            return xvals[n - 1]
    for i in range(1, n):
        if x <= xvals[i]:
            if (xvals[i] - xvals[i - 1]) != 0:
                k = (yvals[i] - yvals[i - 1]) / (xvals[i] - xvals[i - 1])
                return yvals[i] + (x - xvals[i]) * k
            else:
                return xvals[i]
    return -999


# Unit conversions
def toSIunit_T(F):
    return (5 / 9) * (F - 32) + 273.15


def fromSIunit_T(K):
    return (K - 273.15) * (9 / 5) + 32


def toSIunit_p(psi):
    return psi * 0.00689475729


def fromSIunit_p(MPa):
    return MPa / 0.00689475729


def fromSIunit_h(kJ_per_kg):
    return kJ_per_kg / 2.326


def psat_T(T):
    T_K = toSIunit_T(T)
    if T_K <= 647.096 and T_K > 273.15:
        return fromSIunit_p(p4_T(T_K))
    else:
        return float("nan")


def p4_T(T):
    """
    'Release on the IAPWS Industrial Formulation 1997 for the Thermodynamic Properties of Water and Steam, September 1997
    'Section 8.1 The Saturation-Pressure Equation
    'Eq 30, Page 33
    """
    teta = T - 0.23855557567849 / (T - 650.17534844798)
    a = teta**2 + 1167.0521452767 * teta - 724213.16703206
    B = -17.073846940092 * teta**2 + 12020.82470247 * teta - 3232555.0322333
    c = 14.91510861353 * teta**2 - 4823.2657361591 * teta + 405113.40542057
    p4_T = (2 * c / (-B + (B**2 - 4 * a * c) ** 0.5)) ** 4
    return p4_T


def xAir(Tdb, PT, Rh, ret):
    if Tdb <= 32:
        return -999
    pws = psat_T(Tdb)
    pws = toSIunit_p(pws) * 1000
    pw = Rh * pws / 100
    ptot = toSIunit_p(PT) * 1000
    W = 0.62198 * pw / (ptot - pw)
    T = toSIunit_T(Tdb)
    htot = 1.005 * (T - 273.15) + W * (2500.9 + 1.805 * (T - 273.15))
    htot = fromSIunit_h(htot) + 7.686951
    DBT_C = toSIunit_T(Tdb) - 273.15
    TWB = DBT_C * Rh / 100
    tdelta = 0.2
    iCount = 0
    while True:
        iCount += 1
        T = fromSIunit_T(TWB + 273.15)
        if T > 32:
            pwswb = psat_T(T)
        pwswb = toSIunit_p(pwswb) * 1000
        Wswb = 0.62198 * pwswb / (ptot - pwswb)
        Wc = ((2501 - 2.381 * TWB) * Wswb - (DBT_C - TWB)) / (2501 + 1.805 * DBT_C - 4.186 * TWB)
        tdelta = abs(Wc - W) * 100
        if Wc / W > 1.0025:
            TWB -= tdelta
        elif Wc / W < 0.9975:
            TWB += tdelta
        else:
            break
        if iCount > 2000:
            break
    if iCount > 200:
        TWB = -999
    else:
        TWB = fromSIunit_T(TWB + 273.15)
    if ret.lower() == "h":
        return htot
    elif ret.lower() == "w":
        return W
    else:
        return TWB


def xAir_SH(Tdb, PT, W, ret):
    ptot = toSIunit_p(PT) * 1000
    pw = ptot * W / (0.62198 + W)
    pws = psat_T(Tdb)
    pws = toSIunit_p(pws) * 1000
    Rh = pw / pws * 100
    return min(Rh, 100)


# functions for U1 1x1
def u1_a_13(Tdb, Rh):
    a30 = 1.0004 + 0.000034267 * Tdb + 0.000047374 * Tdb**2 + -0.00000023046 * Tdb**3
    a50 = 1.0002 + 0.000017804 * Tdb + 0.000046412 * Tdb**2 + -0.00000021062 * Tdb**3
    a68 = 1.0005 + 0.000033135 * Tdb + 0.000044529 * Tdb**2 + -0.00000018273 * Tdb**3
    a75 = 1.0003 + 0.000028673 * Tdb + 0.000043899 * Tdb**2 + -0.00000017091 * Tdb**3
    a90 = 1.0005 + 0.000037959 * Tdb + 0.000042494 * Tdb**2 + -0.00000014744 * Tdb**3

    xvals = [30, 50, 68, 75, 90]
    yvals = [a30, a50, a68, a75, a90]
    n = len(xvals)
    a13 = yInterpolate(Rh, xvals, yvals, n)
    return a13


def u1_b_13(Tdb, Rh):
    b30 = 0.99859 + -0.00026825 * Tdb + 0.000057199 * Tdb**2 + -0.00000037426 * Tdb**3
    b50 = 0.99848 + -0.00027405 * Tdb + 0.000057234 * Tdb**2 + -0.0000003783 * Tdb**3
    b68 = 0.99838 + -0.00027835 * Tdb + 0.000057267 * Tdb**2 + -0.00000038196 * Tdb**3
    b75 = 0.99835 + -0.00028 * Tdb + 0.000057279 * Tdb**2 + -0.00000038335 * Tdb**3
    b90 = 0.99827 + -0.00028351 * Tdb + 0.0000573 * Tdb**2 + -0.00000038624 * Tdb**3

    xvals = [30, 50, 68, 75, 90]
    yvals = [b30, b50, b68, b75, b90]
    n = len(xvals)
    b13 = yInterpolate(Rh, xvals, yvals, n)
    return b13


def u1_a_2(barom):
    a2 = 13.929 - 1.7466 * barom + 0.058891 * barom**2 + 0 * barom**3
    return a2


def u1_b_2(barom):
    b2 = 12.193 - 1.5115 * barom + 0.050946 * barom**2 + 0 * barom**3
    return b2


def u1_d7(DBHI):
    d7 = DBHI / 0.0091
    return d7 / 1000


def U1_exp_Perf(Tamb, Rh, baro, chill, fired, mode):
    """
    'tamb in Deg F / rh in % (0-100) / baro in PSIA /
    'chill = effectiveness (0-100) / fired=MMBtu/hr of heat input (0-555)
    'mode = "MW", "minMW" or "HR"; defaults to MW if error/typo
    """
    tref = 6  # Deg F
    rhref = 68  # %
    baroref = 14.225  # psia

    # validate/correct inputs
    if Rh < 1:
        Rh = Rh * 100
    elif Rh > 100:
        Rh = 100

    if fired < 0:
        fired = 0
    elif fired > 555:
        fired = 555

    if baro > 16:
        baro = baro * 0.4912  # assume given in inHgA

    if chill > 1:
        chill = chill / 100  # convert to percentage
    elif chill < 0:
        chill = 0

    WBT = xAir(Tamb, baro, Rh, "TWB")
    ambW = xAir(Tamb, baro, Rh, "W")

    # test basis values: (from 2021 test results)
    MW_Ref = 345  # reference/contract capacity value
    DBMW_Ref = 37.176  # MW; from D7/w7 curve at Fire_Ref DBHI
    Fire_Ref = 338.3  # mmbtu/hr; value calculated to provide 345MW at test conditions
    HR_Ref = 7172  # heat rate at 345 MW
    MW_min = 200  # estimate based on ramp rate min value, 2021, was 190 per R. Kozitza 5/4/25
    HR_min = 7350  # estimate based on ramp rate min value, 2021

    if mode in ["minMW", "minHR"]:
        MW_Ref = MW_min
        HR_Ref = HR_min
        chill = 0  # turn evaps off at min load
        fired = 0  # turn ducts off at min load
        Fire_Ref = 0

    # Determine expected CIT
    if Tamb >= 59 and chill > 0:  # ON
        CIT = Tamb - chill * (Tamb - WBT)
        CIRH = xAir_SH(CIT, (baro - 4 * 0.0361), ambW, "RH")
    else:
        CIT = Tamb
        CIRH = Rh

    # corrections
    a13 = u1_a_13(tref, rhref) / u1_a_13(CIT, CIRH)
    b13 = u1_b_13(tref, rhref) / u1_b_13(CIT, CIRH)
    a2 = u1_a_2(baroref) / u1_a_2(baro)
    b2 = u1_b_2(baroref) / u1_b_2(baro)
    d7 = u1_d7(Fire_Ref) - u1_d7(fired)
    w7 = Fire_Ref - fired

    MW_Exp = MW_Ref * a13 * a2 - d7
    HR_Exp = (HR_Ref * MW_Ref / (b13 * b2) - w7 * 1000) / (MW_Ref / (a13 * a2) - d7)

    if "HR" in mode:
        return HR_Exp
    return MW_Exp


# functions for U2 1x1
def u2_a_13(Tdb, Rh):
    a30 = 0.99175 + 0.0016094 * Tdb + 0.0000041614 * Tdb**2 + 0.00000013347 * Tdb**3
    a50 = 0.99191 + 0.0016377 * Tdb + 0.000001976 * Tdb**2 + 0.00000015975 * Tdb**3
    a60 = 0.99196 + 0.0016297 * Tdb + 0.0000015098 * Tdb**2 + 0.00000017054 * Tdb**3
    a75 = 0.99232 + 0.0016549 * Tdb + -0.00000059383 * Tdb**2 + 0.00000019789 * Tdb**3
    a90 = 0.99265 + 0.0016819 * Tdb + -0.0000030109 * Tdb**2 + 0.00000023057 * Tdb**3

    xvals = [30, 50, 60, 75, 90]
    yvals = [a30, a50, a60, a75, a90]
    n = len(xvals)
    a13 = yInterpolate(Rh, xvals, yvals, n)
    return a13


def u2_b_13(Tdb, Rh):
    b30 = 0.99104 + 0.0016535 * Tdb + 0.0000046987 * Tdb**2 + 0.000000085959 * Tdb**3
    b50 = 0.99089 + 0.001645 * Tdb + 0.0000051103 * Tdb**2 + 0.000000077386 * Tdb**3
    b60 = 0.99089 + 0.0016474 * Tdb + 0.00000484 * Tdb**2 + 0.000000078544 * Tdb**3
    b75 = 0.99081 + 0.0016448 * Tdb + 0.0000049756 * Tdb**2 + 0.000000073298 * Tdb**3
    b90 = 0.99088 + 0.0016575 * Tdb + 0.0000043361 * Tdb**2 + 0.000000075509 * Tdb**3

    xvals = [30, 50, 60, 75, 90]
    yvals = [b30, b50, b60, b75, b90]
    n = len(xvals)
    b13 = yInterpolate(Rh, xvals, yvals, n)
    return b13


def u2_a_2(barom):
    a2 = 2.814 + -0.17939 * barom + 0.0036459 * barom**2 + 0 * barom**3
    return a2


def u2_b_2(barom):
    b2 = 2.8971 + -0.19735 * barom + 0.004498 * barom**2 + 0 * barom**3
    return b2


def u2_d7(DBHI):
    d7 = DBHI / 0.0091
    return d7 / 1000


def U2_exp_Perf(Tamb, Rh, baro, chill, fired, mode):
    """
    'tamb in Deg F / rh in % (0-100) / baro in PSIA /
    'chill = effectiveness (0-100) / fired=MMBtu/hr of heat input (0-742.8)
    'mode = "MW", "minMW" or "HR"; defaults to MW if error/typo
    """
    tref = 6  # Deg F
    rhref = 68  # %
    baroref = 14.225  # psia

    # validate/correct inputs
    if Rh < 1:
        Rh = Rh * 100
    elif Rh > 100:
        Rh = 100

    if fired < 0:
        fired = 0
    elif fired > 742.8:
        fired = 742.8

    if baro > 16:
        baro = baro * 0.4912  # assume given in inHgA

    if chill > 1:
        chill = chill / 100  # convert to percentage
    elif chill < 0:
        chill = 0

    WBT = xAir(Tamb, baro, Rh, "TWB")
    ambW = xAir(Tamb, baro, Rh, "W")

    # test basis values: (from 2021 test results)
    MW_Ref = 375  # reference/contract capacity value adj 370 to 375 6/9/25 per S. Reinhart
    DBMW_Ref = 607.6 / 9.1  # MW; from D7/w7 curves at Fire_Ref DBHI
    Fire_Ref = 607.6  # mmbtu/hr; value calculated to provide 375MW at test conditions
    HR_Ref = 7366  # heat rate at 375 MW
    MW_min = 200  # estimate based on ramp rate min value, 2021 was 190 per R. Kozitza 5/4/25
    HR_min = 7400  # estimate based on ramp rate min value, 2021

    if mode in ["minMW", "minHR"]:
        MW_Ref = MW_min
        HR_Ref = HR_min
        chill = 0  # turn evaps off at min load
        fired = 0  # turn ducts off at min load
        Fire_Ref = 0

    # Determine expected CIT
    if Tamb >= 59 and chill > 0:  # ON
        CIT = Tamb - chill * (Tamb - WBT)
        CIRH = xAir_SH(CIT, (baro - 4 * 0.0361), ambW, "RH")
    else:
        CIT = Tamb
        CIRH = Rh

    # corrections
    a13 = u2_a_13(tref, rhref) / u2_a_13(CIT, CIRH)
    b13 = u2_b_13(tref, rhref) / u2_b_13(CIT, CIRH)
    a2 = u2_a_2(baroref) / u2_a_2(baro)
    b2 = u2_b_2(baroref) / u2_b_2(baro)
    d7 = u2_d7(Fire_Ref) - u2_d7(fired)
    w7 = Fire_Ref - fired

    MW_Exp = MW_Ref * a13 * a2 - d7
    HR_Exp = (HR_Ref * MW_Ref / (b13 * b2) - w7 * 1000) / (MW_Ref / (a13 * a2) - d7)

    if "HR" in mode:
        return HR_Exp
    return MW_Exp


# functions for 2x1
def CF46LTamb(Tamb, Rh, XZ_COOL, XZ_CASE):
    """
    ' PTC 46 CORRECTION CURVES, COPIED IN FROM EXCEL WORKBOOK PROVIDED BY Zach Fairweather 4/5/19
    ' Mankato 2x1 Thermoflow Model
    ' Capacity Correction Factor vs. Ambient Dry Bulb Temp & Relative Humidity
    ' Cf = Capacity Correction Factor at Tamb & Rh
    ' Tamb = Ambient Dry Bulb Temperature (Deg F)
    ' Rh = Ambient Relative Humidity (%)

    ' XZ_CASE = 0 (2x1, Minimum Load) *varies with ambient unlike set value used in SoCo dispatch
    ' XZ_CASE = 1 (2x1, Base Load)
    ' XZ_CASE = 2 (2x1, Supplemental Fired)
    ' XZ_COOL = 0 (Evap Coolers Off)
    ' XZ_COOL = 1 (Evap Coolers On)
    """
    if XZ_CASE == 0:  # 2x1, Minimum Load
        Cf = (
            0.00000000011975 * Tamb**5
            - 0.000000028868 * Tamb**4
            + 0.0000022778 * Tamb**3
            - 0.000075123 * Tamb**2
            - 0.00084693 * Tamb
            + 1.0073
        )

    elif XZ_CASE == 1:  # 2x1, Base Load
        if XZ_COOL == 0:  # Evap Coolers Off
            # Ref. net output at 6 F, 68% RH from Thermoflow = 640,887 KW
            if Tamb < 0:
                Cf = 1  # Gas turbines at or near shaft limits
            else:
                Cf = (
                    0.0000000005197 * Tamb**4
                    - 0.000000184 * Tamb**3
                    + 0.000002693 * Tamb**2
                    - 0.001597 * Tamb
                    + 1.007
                )

        else:  # Evap Coolers On
            # Ref. net output at 89 F, 53% RH from Thermoflow = 532,464 KW
            C30 = 0.000001268 * Tamb**3 - 0.0003028 * Tamb**2 + 0.02103 * Tamb + 0.6623
            C68 = -0.0000003296 * Tamb**3 + 0.0000796 * Tamb**2 - 0.009917 * Tamb + 1.468
            C90 = -0.0000005084 * Tamb**3 + 0.0001124 * Tamb**2 - 0.01201 * Tamb + 1.497

            if 0 <= Rh < 68:
                Cf = InterpF(Rh, 30, 68, C30, C68)
            else:
                Cf = InterpF(Rh, 68, 90, C68, C90)

    elif XZ_CASE == 2:  # 2x1, Supplemental Fired
        if XZ_COOL == 0:  # Evap Coolers Off
            # Ref. net output at 6 F, 68% RH from Thermoflow = 756,770 KW
            C30 = -0.00000001625 * Tamb**3 - 0.000005977 * Tamb**2 - 0.001248 * Tamb + 1.005
            C68 = -0.000000004607 * Tamb**3 - 0.000009583 * Tamb**2 - 0.001114 * Tamb + 1.004
            C90 = 0.0000000003586 * Tamb**3 - 0.00001023 * Tamb**2 - 0.001129 * Tamb + 1.004

            if Tamb < 0:
                Cf = 1  # Gas turbines at or near shaft limits
            else:
                if 0 <= Rh < 68:
                    Cf = InterpF(Rh, 30, 68, C30, C68)
                else:
                    Cf = InterpF(Rh, 68, 90, C68, C90)

        else:  # Evap Coolers On
            # Ref. net output at 89 F, 53% RH from Thermoflow = 647,810 KW
            C30 = 0.0000000138 * Tamb**3 - 0.00000553 * Tamb**2 - 0.001536 * Tamb + 1.194
            C68 = -0.0000001753 * Tamb**3 + 0.00003633 * Tamb**2 - 0.005164 * Tamb + 1.282
            C90 = -0.0000009077 * Tamb**3 + 0.0002051 * Tamb**2 - 0.01828 * Tamb + 1.611
            if 0 <= Rh < 68:
                Cf = InterpF(Rh, 30, 68, C30, C68)
            else:
                Cf = InterpF(Rh, 68, 90, C68, C90)

    return Cf


def CF46LPatm(Patm, XZ_COOL, XZ_CASE):
    """
    ' PTC 46 CORRECTION CURVES, COPIED IN FROM EXCEL WORKBOOK PROVIDED BY Zach Fairweather 4/5/19
    ' Mankato 2x1 Thermoflow Model
    ' Capacity Correction Factor vs. Atmospheric Pressure
    ' Cf = Capacity Correction Factor at Patm
    ' Patm = Atmospheric Pressure (psia)

    ' XZ_CASE = 0 (2x1, Minimum Load) *varies with ambient unlike set value used in SoCo dispatch
    ' XZ_CASE = 1 (2x1, Base Load)
    ' XZ_CASE = 2 (2x1, Supplemental Fired)
    ' XZ_COOL = 0 (Evap Coolers Off)
    ' XZ_COOL = 1 (Evap Coolers On)
    """
    if XZ_CASE == 1:  # 2x1, Base Load
        Cf = 0.04886 * Patm + 0.305

    elif XZ_CASE == 2:  # 2x1, Supp Fired
        Cf = -0.02463 * Patm**2 + 0.7442 * Patm - 4.602

    else:  # Minimum load
        Cf = 1

    return Cf


def CF46HRPatm(Patm, XZ_COOL, XZ_CASE):
    """
    ' PTC 46 CORRECTION CURVES, COPIED IN FROM EXCEL WORKBOOK PROVIDED BY Zach Fairweather 4/5/19
    ' Mankato 2x1 Thermoflow Model
    ' Heat Rate Correction Factor vs. Atmospheric Pressure
    ' Cf = Heat Rate Correction Factor at Patm
    ' Patm = Atmospheric Pressure (psia)

    ' XZ_CASE = 0 (2x1, Minimum Load) *varies with ambient unlike set value used in SoCo dispatch
    ' XZ_CASE = 1 (2x1, Base Load)
    ' XZ_CASE = 2 (2x1, Supplemental Fired)
    ' XZ_COOL = 0 (Evap Coolers Off)
    ' XZ_COOL = 1 (Evap Coolers On)
    """
    if XZ_CASE == 1:  # 2x1, Base Load
        Cf = 0.0004241 * Patm**2 - 0.006318 * Patm + 1.004

    elif XZ_CASE == 2:  # 2x1, Supp Fired
        Cf = -0.008931 * Patm**2 + 0.2628 * Patm - 0.9312

    else:  # Minimum load
        Cf = 1

    return Cf


def InterpF(x, X1, X2, Y1, Y2):
    """
    ' Interpolation Subroutine
    ' Written by MAS 7/00
    """
    Z = (x - X1) / (X2 - X1)
    Y = Y1 + (Y2 - Y1) * Z
    return Y


def AdjHeatRate(ExpMW, HeatIn):
    """
    ' Calculate heat rate at 740,000 kw (MEC site transmission limit) if load predicted is higher than 740,000 kW
    ' ExpMW = predicted MW output in SF EC off mode for the forecast weather condtions (MW)
    ' HeatIn = fuel heat input corresponding to the predicted output and heat rate (MMBtu/hr)
    ' Heat input values calculated below are total duct burner heat input values based on model test results for unit output vs DB heat input
    """
    ExpkW = ExpMW * 1000  # Convert MW to kW

    # DB heat input at predicted kW output (MMBtu/hr)
    HIcalc = -0.00000001512 * ExpkW**2 + 0.029604 * ExpkW - 12792

    # DB heat input at 740,000 kW (MMBtu/hr)
    HI740 = -0.00000001512 * 740000**2 + 0.029604 * 740000 - 12792

    # reduction in DB heat input needed to get down to 740,000 kW output (MMBtu/hr)
    Diff = HIcalc - HI740

    AdjHI = HeatIn - Diff  # Predicted total heat input to achieve 740,000 kW output (MMBtu/hr)
    adj_rate = AdjHI * 1000000 / 740000  # Btu/kWh
    return adj_rate


def CF46HRTamb(Tamb, Rh, XZ_COOL, XZ_CASE):
    """
    ' PTC 46 CORRECTION CURVES, COPIED IN FROM EXCEL WORKBOOK PROVIDED BY Zach Fairweather 4/5/19
    ' Mankato 2x1 Thermoflow Model
    ' Heat Rate Correction Factor vs. Ambient Dry Bulb Temp & Relative Humidity
    ' Cf = Heat Rate Correction Factor at Tamb & Rh
    ' Tamb = Ambient Dry Bulb Temperature (Deg F)
    ' Rh = Ambient Relative Humidity (%)

    ' XZ_CASE = 0 (2x1, Minimum Load) *varies with ambient unlike set value used in SoCo dispatch
    ' XZ_CASE = 1 (2x1, Base Load)
    ' XZ_CASE = 2 (2x1, Supplemental Fired)
    ' XZ_COOL = 0 (Evap Coolers Off)
    ' XZ_COOL = 1 (Evap Coolers On)
    """
    if XZ_CASE == 0:  # 2x1, Minimum Load
        # Ref. net heat rate at 6 F, 68% RH from Thermoflow = 7208 Btu/kWh
        Cf = (
            -0.000000000015773 * Tamb**5
            + 0.0000000011477 * Tamb**4
            + 0.00000024957 * Tamb**3
            - 0.000021463 * Tamb**2
            + 0.00046375 * Tamb
            + 0.99793
        )

    elif XZ_CASE == 1:  # 2x1, Base Load
        if XZ_COOL == 0:  # Evap Coolers Off
            # Ref. net heat rate at 6 F, 68% RH from Thermoflow = 6901  Btu/kWh
            Cf = (
                -0.00000000003618 * Tamb**5
                + 0.000000004272 * Tamb**4
                + 0.00000006618 * Tamb**3
                - 0.000007082 * Tamb**2
                - 0.0004513 * Tamb
                + 1.003
            )
        else:  # Evap Coolers On
            # Ref. net heat rate at 89 F, 53% RH from Thermoflow =  6994 Btu/kWh
            C30 = -0.0000009045 * Tamb**3 + 0.0002143 * Tamb**2 - 0.01605 * Tamb + 1.36
            C68 = 0.0000004716 * Tamb**3 - 0.0001087 * Tamb**2 + 0.009276 * Tamb + 0.7085
            C90 = 0.000000126 * Tamb**3 - 0.00002163 * Tamb**2 + 0.00216 * Tamb + 0.9037
            if 0 <= Rh < 68:
                Cf = InterpF(Rh, 30, 68, C30, C68)
            else:
                Cf = InterpF(Rh, 68, 90, C68, C90)

    elif XZ_CASE == 2:  # 2x1, Supplemental Fired
        if XZ_COOL == 0:  # Evap Coolers Off
            # Ref. net heat rate at 6 F, 68% RH from Thermoflow = 7296 Btu/kWhW
            C30 = (
                -0.0000000001113 * Tamb**5
                + 0.0000000209 * Tamb**4
                - 0.000001289 * Tamb**3
                + 0.00003917 * Tamb**2
                - 0.0008989 * Tamb
                + 1.004
            )
            C68 = (
                -0.0000000001922 * Tamb**5
                + 0.00000003276 * Tamb**4
                - 0.000001729 * Tamb**3
                + 0.0000401 * Tamb**2
                - 0.0008021 * Tamb
                + 1.004
            )
            C90 = (
                -0.000000000164 * Tamb**5
                + 0.00000002789 * Tamb**4
                - 0.000001459 * Tamb**3
                + 0.00003524 * Tamb**2
                - 0.0007581 * Tamb
                + 1.004
            )
            if 0 <= Rh < 68:
                Cf = InterpF(Rh, 30, 68, C30, C68)
            else:
                Cf = InterpF(Rh, 68, 90, C68, C90)

        else:  # Evap Coolers On
            # Ref. net heat rate at 89 F, 53% RH from Thermoflow = 7473 Btu/kWh
            C30 = 0.0000001939 * Tamb**3 - 0.00004516 * Tamb**2 + 0.004217 * Tamb + 0.8367
            C68 = 0.0000002416 * Tamb**3 - 0.00005078 * Tamb**2 + 0.004481 * Tamb + 0.839
            C90 = -0.00000008564 * Tamb**3 + 0.00002603 * Tamb**2 - 0.001327 * Tamb + 0.9863
            if 0 <= Rh < 68:
                Cf = InterpF(Rh, 30, 68, C30, C68)
            else:
                Cf = InterpF(Rh, 68, 90, C68, C90)

    return Cf


def fac_a_13(tatm, Rh, c, x):  # facility a13 factor (2x1)
    result = CF46LTamb(tatm, Rh, c, x)  # evaps OFF, Fired
    return result


def fac_b_13(tatm, Rh, c, x):  # facility a13 factor (2x1)
    result = CF46HRTamb(tatm, Rh, c, x)  # evaps OFF, Fired
    return result


def fac_a_2(barometer, x):  # facility a2 factor (2x1)
    result = CF46LPatm(barometer, 0, x)  # evaps OFF, Fired
    return result


def fac_b_2(barometer, x):  # facility b2 factor (2x1)
    result = CF46HRPatm(barometer, 0, x)  # evaps OFF, Fired
    return result


def Facility_Perf_R2021(Tamb, Rh, baro, chill, fired, mode):
    if mode not in ["MW", "HR", "minMW", "minHR"]:
        raise ValueError("Invalid mode.")

    # Reference values

    # reference/contract capacity value - 5MW of Fired and HP additional aux loads - was +5MW
    MW_Ref_Fired = 345 + 375 - 5
    MW_Ref_Base = 614.1  # from old SS changed from 619.1 to 614.1 per notes by R. Kozitza 5/4/25
    MW_min = 433.8  # from old SS plus 20 MW per R. Kozitza 5/4/25

    Fire_Ref = 338.3 + 607.6  # mmbtu/hr; combined total
    HR_Ref_Fired = (7195 * 345 + 7366 * 375) / (345 + 375)  # combined heat rate removed + 5#
    HR_Ref_Base = 6413  # from old SS
    HR_min = 7270  # from old SS

    if Tamb < 60:
        chill = 0

    if chill > 0:
        tref, rhref, xChill = [87.7, 59.1, 1]
        rhref = 59.1
        MW_Ref_Base = 528.2  # from old SS was 533.2
        HR_Ref_Base = 6979  # from old SS
        MW_Ref_Fired = 644.3  # from old SS was 654.3
        HR_Ref_Fired = 7309  # from old SS
        xChill = 1
    else:
        tref, rhref, xChill = [6, 68, 0]

    baroref = 14.225  # psia

    # validate/correct inputs
    if Rh < 1:
        Rh = Rh * 100
    elif Rh > 100:
        Rh = 100

    if fired < 0:
        fired = 0
    elif fired > 1297.8:
        fired = 1297.8

    if baro > 16:
        baro = baro * 0.4912  # assume given in inHgA

    # TODO: refactor
    if mode in ["minMW", "minHR"]:
        a13 = fac_a_13(tref, rhref, 0, 0) / fac_a_13(Tamb, Rh, 0, 0)
        b13 = fac_b_13(tref, rhref, 0, 0) / fac_b_13(Tamb, Rh, 0, 0)
        a2 = fac_a_2(baroref, 0) / fac_a_2(baro, 0)
        b2 = fac_b_2(baroref, 0) / fac_b_2(baro, 0)
        MW_Ref = MW_min
        HR_Ref = HR_min
    elif fired > 200:  # assume fully fired, only corrects for TT/RH/PT
        a13 = fac_a_13(tref, rhref, xChill, 2) / fac_a_13(Tamb, Rh, xChill, 2)
        b13 = fac_b_13(tref, rhref, xChill, 2) / fac_b_13(Tamb, Rh, xChill, 2)
        a2 = fac_a_2(baroref, 2) / fac_a_2(baro, 2)
        b2 = fac_b_2(baroref, 2) / fac_b_2(baro, 2)
        MW_Ref = MW_Ref_Fired
        HR_Ref = HR_Ref_Fired
    else:  # use baseload curves
        a13 = fac_a_13(tref, rhref, xChill, 1) / fac_a_13(Tamb, Rh, xChill, 1)
        b13 = fac_b_13(tref, rhref, xChill, 1) / fac_b_13(Tamb, Rh, xChill, 1)
        a2 = fac_a_2(baroref, 1) / fac_a_2(baro, 1)
        b2 = fac_b_2(baroref, 1) / fac_b_2(baro, 1)
        MW_Ref = MW_Ref_Base
        HR_Ref = HR_Ref_Base

    MW_Exp = MW_Ref / a13 / a2
    HR_Exp = HR_Ref / b13 / b2

    if "HR" in mode:
        return HR_Exp
    return MW_Exp


# main functions to predict day ahead using forecasted weather data
CALC_COLS_AND_ARGS_U1_1x1 = {
    "MW_Output_SF_EC_Off_U1_1x1": [14.225, 0, 338, "MW"],
    "MW_Output_SF_EC_On_U1_1x1": [14.225, 0.85, 338, "MW"],
    "Heat_Rate_SF_EC_Off_U1_1x1": [14.225, 0, 338, "HR"],
    "MW_Output_BL_EC_Off_U1_1x1": [14.225, 0, 0, "MW"],
    "MW_Output_BL_EC_On_U1_1x1": [14.225, 0.85, 0, "MW"],
    "Heat_Rate_BL_EC_Off_U1_1x1": [14.225, 0, 0, "HR"],
    "MW_Output_Min_Load_U1_1x1": [14.225, 0, 0, "minMW"],
    "Heat_Rate_Min_Load_U1_1x1": [14.225, 0, 0, "minHR"],
}


def calculate_day_ahead_u1_1x1(df_weather):
    df = df_weather[["temp", "humidity"]].copy()
    for col, args in CALC_COLS_AND_ARGS_U1_1x1.items():
        func = lambda tmp, rh: U1_exp_Perf(tmp, rh, *args)
        df[col] = df.apply(lambda row: func(row["temp"], row["humidity"]), axis=1)
    ext = "U1_1x1"
    output_cols = [
        f"MW_Output_SF_EC_Off_{ext}",
        f"MW_Output_SF_EC_On_{ext}",
        f"MW_Output_BL_EC_Off_{ext}",
        f"MW_Output_BL_EC_On_{ext}",
    ]
    df[f"MEC_Max_Expected_Output_{ext}"] = df[output_cols].max(axis=1)
    return df


CALC_COLS_AND_ARGS_U2_1x1 = {
    "MW_Output_SF_EC_Off_U2_1x1": [14.225, 0, 613, "MW"],
    "MW_Output_SF_EC_On_U2_1x1": [14.225, 85, 613, "MW"],
    "Heat_Rate_SF_EC_Off_U2_1x1": [14.225, 0, 613, "HR"],
    "MW_Output_BL_EC_Off_U2_1x1": [14.225, 0, 0, "MW"],
    "MW_Output_BL_EC_On_U2_1x1": [14.225, 0.85, 0, "MW"],
    "Heat_Rate_BL_EC_Off_U2_1x1": [14.225, 0, 0, "HR"],
    "MW_Output_Min_Load_U2_1x1": [14.225, 0, 0, "minMW"],
    "Heat_Rate_Min_Load_U2_1x1": [14.225, 0, 0, "minHR"],
}


def calculate_day_ahead_u2_1x1(df_weather):
    df = df_weather[["temp", "humidity"]].copy()
    for col, args in CALC_COLS_AND_ARGS_U2_1x1.items():
        func = lambda tmp, rh: U2_exp_Perf(tmp, rh, *args)
        df[col] = df.apply(lambda row: func(row["temp"], row["humidity"]), axis=1)
    ext = "U2_1x1"
    output_cols = [
        f"MW_Output_SF_EC_Off_{ext}",
        f"MW_Output_SF_EC_On_{ext}",
        f"MW_Output_BL_EC_Off_{ext}",
        f"MW_Output_BL_EC_On_{ext}",
    ]
    df[f"MEC_Max_Expected_Output_{ext}"] = df[output_cols].max(axis=1)
    return df


CALC_COLS_AND_ARGS_2x1 = {
    "MW_Output_SF_EC_Off_2x1": [14.225, 0, 1297.8, "MW"],
    "MW_Output_SF_EC_On_2x1": [14.225, 85, 1297.8, "MW"],
    "Heat_Rate_SF_EC_Off_2x1": [14.225, 0, 1297.8, "HR"],
    "MW_Output_BL_EC_Off_2x1": [14.225, 0, 0, "MW"],
    "MW_Output_BL_EC_On_2x1": [14.225, 85, 0, "MW"],
    "Heat_Rate_BL_EC_Off_2x1": [14.225, 0, 0, "HR"],
    "MW_Output_Min_Load_2x1": [14.225, 0, 0, "minMW"],
    "Heat_Rate_Min_Load_2x1": [14.225, 0, 0, "minHR"],
}


def calculate_day_ahead_2x1(df_weather):
    df = df_weather[["temp", "humidity"]].copy()
    for col, args in CALC_COLS_AND_ARGS_2x1.items():
        func = lambda tmp, rh: Facility_Perf_R2021(tmp, rh, *args)
        df[col] = df.apply(lambda row: func(row["temp"], row["humidity"]), axis=1)
    output_cols = [
        "MW_Output_SF_EC_Off_2x1",
        "MW_Output_SF_EC_On_2x1",
        "MW_Output_BL_EC_Off_2x1",
        "MW_Output_BL_EC_On_2x1",
    ]
    df["MEC_Max_Expected_Output_2x1"] = df[output_cols].max(axis=1)
    return df


def calculate_day_ahead(df_weather):
    output = {
        "U1_1x1": calculate_day_ahead_u1_1x1(df_weather),
        "U2_1x1": calculate_day_ahead_u2_1x1(df_weather),
        "2x1": calculate_day_ahead_2x1(df_weather),
    }
    return output


def get_max_temp_timestamps(df):
    """Returns list of Timestamps corresponding to the maximum temperature for each day."""
    if not isinstance(df.index, pd.DatetimeIndex):
        raise TypeError("Dataframe must have a datetime index.")
    if "temp" not in df.columns:
        raise ValueError("Dataframe must have column: 'temp'")
    timestamps = []
    for day in df.index.floor("D").unique():
        tstamp = df.loc[df.index.floor("D") == day, "temp"].idxmax()
        timestamps.append(tstamp)
    return timestamps


def _create_max_output_tables(output):
    max_temp_timestamps = get_max_temp_timestamps(output["2x1"])
    df_list = []
    for tstamp in max_temp_timestamps:
        df_list_2 = []
        for key, df_ in output.items():
            df = df_.loc[tstamp].to_frame()
            df = df.T.rename_axis("Central Hour Beginning").reset_index(drop=False)
            df = df.rename(columns={"temp": "Temp", "humidity": "Humidity"})
            df.columns = [c.rstrip(f"_{key}") for c in df.columns]
            df = df.T.copy()
            df.columns = [key]
            df_list_2.append(df)
        df_list.append(pd.concat(df_list_2, axis=1))
    return df_list


def _header_row(df) -> pd.DataFrame:
    row = pd.Series(df.columns, index=df.columns).to_frame().T
    row.index = [""]
    return row


def _blank_row(df) -> pd.DataFrame:
    row = pd.Series([""] * len(df.columns), index=df.columns).to_frame().T
    row.index = [""]
    return row


def combined_tables_for_excel_sheet(output):
    """Goes to the right of the results summary table."""
    max_output_table_list = _create_max_output_tables(output)
    df_list = []
    for df_ in max_output_table_list:
        df_list.append(pd.concat([_header_row(df_), df_], axis=0))
    df = pd.concat(df_list, axis=0)
    df = df.reset_index(drop=False)
    df.columns = [""] * len(df.columns)
    return df


def create_results_summary_table(output):
    """Table for Results sheet."""
    timestamps = get_max_temp_timestamps(output["2x1"])
    max_temps = [output["2x1"].at[tstamp, "temp"] for tstamp in timestamps]
    dfs = pd.DataFrame({"Central Hour Beginning": timestamps, "Temp (F)": max_temps}).T
    df_list = [dfs, _blank_row(dfs)]
    for key, df_ in output.items():
        pfx = key if key == "2x1" else {"U1_1x1": "Unit 1", "U2_1x1": "Unit 2"}.get(key)
        data = {f"{pfx} Status:": ["Available"] * len(max_temps)}
        if key != "2x1":
            data.update(
                {
                    f"{pfx} Evaps Status:": ["Available"] * len(max_temps),
                    f"{pfx} Evaps Gen in Limit?": ["YES"] * len(max_temps),
                    f"{pfx} Ducts Status:": ["Available"] * len(max_temps),
                }
            )
        dfs_ = pd.DataFrame(data)
        getvals = lambda col, x=0: [(df_.at[tstamp, col] + x) for tstamp in timestamps]
        getmaxvals = lambda cols: [df_.loc[tstamp, cols].max() for tstamp in timestamps]

        constant = 10 if key != "2x1" else 20
        dfs_[f"{pfx} Min:"] = getvals(f"MW_Output_Min_Load_{key}", x=constant)
        dfs_[f"{pfx} Max w/o Ducts:"] = getmaxvals(
            [f"MW_Output_BL_EC_On_{key}", f"MW_Output_BL_EC_Off_{key}"]
        )
        dfs_[f"{pfx} Max w/ Ducts:"] = getmaxvals(
            [f"MW_Output_SF_EC_On_{key}", f"MW_Output_SF_EC_Off_{key}"]
        )
        dfs_ = dfs_.T
        df_list.extend([dfs_, _blank_row(dfs_)])

    df_combined = pd.concat(df_list, axis=0)
    return df_combined


def format_results_summary_table(df_results):
    """uses df_combined from above function"""
    df_ = df_results.copy()
    fmt = lambda x: x.lstrip("Unit 1 ").lstrip("Unit 2 ").lstrip("2x1 ")
    df_.index = df_.index.map(fmt)
    match_keys = ["Status:", "Evaps Gen in Limit?", "Ducts Status:"]
    matches = list(sorted(df_.index.get_indexer_non_unique(match_keys)[0]))

    # add blank rows
    split_list = []
    start_ = 0
    for idx in matches:
        end_ = idx + 1
        split_list.extend([df_.iloc[start_:end_], _blank_row(df_)])
        start_ = end_
    if start_ + 1 < len(df_):
        split_list.append(df_.iloc[start_:])
    df = pd.concat(split_list, axis=0, ignore_index=False)

    # add blank columns
    concat_list = []
    for c in range(df.shape[1]):
        df_blank_col = pd.DataFrame(index=df.index, data={f"{c}_blank": np.nan})
        concat_list.extend([df.iloc[:, [c]], df_blank_col])
    df_final = pd.concat(concat_list, axis=1)
    df_final = df_final.reset_index(drop=False)
    df_final.columns = list(range(df_final.shape[1]))
    df_final.at[0, 0] = ""
    return df_final


def generate_results_sheet_dataframe(output):
    """Creates df for main sheet of output file"""
    # create main summary table
    df_results = create_results_summary_table(output)
    df_summary = format_results_summary_table(df_results=df_results)

    # create detailed summary table
    df_tables = combined_tables_for_excel_sheet(output)

    # combine
    df_results = pd.concat([df_summary, df_tables], axis=1, ignore_index=True)
    df_results.columns = list(range(df_results.shape[1]))
    return df_results


# TEMP
SITE_METADATA = {
    "Mankato": {
        "latitude": 44.167,
        "longitude": -94.002,
        "timezone": "US/Central",
    }
}
template_dir = Path(oepaths.python_projects, "Repositories", "excel_templates")
template_fp = Path(template_dir, "MEC_day_ahead_forecast_TEMPLATE.xlsx")
fmt_datetime = "m-d-yy hh:mm;@"
fmt_1z = "0.0"
fmt_int = "0"


def generate_day_ahead_output_file(df_results, df_weather, forecast_data, savepath, q=True):
    qprint = quiet_print_function(q=q)
    output_keys = ["U1_1x1", "U2_1x1", "2x1"]
    output = {key: val for key, val in forecast_data.items() if key in output_keys}
    if len(output) != 3:
        raise ValueError("Invalid forecast_data format.")
    if isinstance(df_weather.index, pd.DatetimeIndex):
        df_weather = df_weather.reset_index(drop=False)

    wb = openpyxl.load_workbook(template_fp)
    ws = wb["Sheet1"]
    for r, row in enumerate(dataframe_to_rows(df_results, index=False, header=False), start=1):
        for c, val in enumerate(row, start=2):
            if pd.isna(val):
                continue
            cell = ws.cell(row=r, column=c)
            cell.value = val
            if type(val) not in (pd.Timestamp, float):
                continue
            fmt = fmt_1z if type(val) is float else fmt_datetime
            cell.number_format = fmt

    weather_sheet = "Forecasted Weather"
    wb.create_sheet(weather_sheet)
    ws_w = wb[weather_sheet]
    for row in dataframe_to_rows(df_weather, index=False, header=True):
        ws_w.append(row)

    for key, df_ in output.items():
        new_sheet = f"Day Ahead ({key})"
        wb.create_sheet(new_sheet)
        ws_ = wb[new_sheet]
        df_ = df_.reset_index(drop=False)
        for row in dataframe_to_rows(df_, index=False, header=True):
            ws_.append(row)

    wb.save(savepath)
    wb.close()
    qprint(f"\nSaved: {str(savepath)}")
    return


def generate_day_ahead_forecasts(site, start=None, n_days=5, savepath=None, q=True):
    """Generates Excel file with day ahead forecast for the next 5 days"""
    if site not in SITE_METADATA:
        raise KeyError(f"{site=} is not currently supported.")
    qprint = quiet_print_function(q=q)
    if savepath is not None:
        savepath = Path(savepath)
        if not savepath.parent.exists():
            raise ValueError(f"Invalid directory for savepath = {str(savepath)}")
        elif savepath.suffix != ".xlsx":
            new_fname = savepath.stem + ".xlsx"
            qprint(
                "Wrong file extension detected in savepath; "
                f"changing from {savepath.name} to {new_fname}."
            )
            savepath = savepath.with_name(new_fname)

    if start is None:
        forecast_start = pd.Timestamp(dt.datetime.now().date()) + pd.Timedelta(days=1)
    else:
        forecast_start = pd.Timestamp(start)
    qprint(f"{site = }\n{forecast_start = :%Y-%m-%d}\n{n_days = }")

    # get forecasted weather data from DTN
    lat, lon, tz = SITE_METADATA[site].values()
    df_dtn = get_dtn_weather_forecast(lat, lon, tz, forecast_start, n_days=n_days, q=q)

    # calculate day ahead
    output = calculate_day_ahead(df_weather=df_dtn)

    # generate results/summary
    df_results = generate_results_sheet_dataframe(output)

    # write to Excel file
    if savepath is not None:
        generate_day_ahead_output_file(
            df_results=df_results, df_weather=df_dtn, forecast_data=output, savepath=savepath, q=q
        )

    return {**output, "results": df_results, "weather": df_dtn}
