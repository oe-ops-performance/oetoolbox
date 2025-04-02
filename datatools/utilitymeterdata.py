import openpyxl
import os, calendar
import pandas as pd
import datetime as dt
from pathlib import Path
import shutil, tempfile, xmltodict

from ..dataquery import pireference as ref, pitools as pt  # type: ignore
from ..datatools import mdatatransforms  # type: ignore
from ..reporting.tools import load_meter_historian, date_range, site_frpath  # type: ignore
from ..utils import oepaths  # type: ignore


now = lambda: dt.datetime.now()
tstamp = lambda: now().strftime("%H%M%S")


def invalid_yearmonth(year, month):
    if (year is None) or (month is None):
        return True
    ref_tstamp = pd.Timestamp(year, month, 1)
    today_tstamp = pd.Timestamp(now().year, now().month, 1)
    return ref_tstamp > today_tstamp


validate_YM = lambda y, m: [now().year, now().month] if invalid_yearmonth(y, m) else [y, m]


def print_sitelist_block(sitelist, n_=4):
    w, s0, s1 = [len(sitelist), 0, n_]
    while w > 0:
        for site in sitelist[s0:s1]:
            print(f"    {site.ljust(18)}", end="")
        w -= n_
        s0 = s1
        s1 += n_
        print("")
    return


latestfile = lambda fplist: (
    None if not fplist else max((fp.stat().st_ctime, fp) for fp in fplist)[1]
)

# SOLAR SITES - ATLAS
atlas_mpath = Path(oepaths.commercial, "Atlas Portfolio")
atlas_meterfolders = {
    "AZ1": Path(atlas_mpath, "1.05 AZ 1", "1.05.3 Meter Data"),
    # 'FL1': Path(atlas_mpath, '1.10 FL Solar 1'),
    # 'FL4': Path(atlas_mpath, '1.11 FL Solar 4'),
    "GA3": Path(atlas_mpath, "1.01 GA Solar 3", "1.01.06 Meter Data"),
    "GA4": Path(atlas_mpath, "1.02 GA Solar 4", "1.02.06 Meter Data"),
    "Grand View East": Path(atlas_mpath, "1.03 Grand View", "1.03.02 Meter Data"),
    "Grand View West": Path(atlas_mpath, "1.03 Grand View", "1.03.02 Meter Data"),
    "Maplewood 1": Path(atlas_mpath, "1.08 Maplewood 1", "1.08.02 Tenaska", "1.08.02.1 Meter Data"),
    "Maplewood 2": Path(atlas_mpath, "1.09 Maplewood 2", "1.09.02 Tenaska", "1.09.02.1 Meter Data"),
    "MS3": Path(atlas_mpath, "1.07 MS 3", "1.07.02 Meter Data"),
    "Sweetwater": Path(
        atlas_mpath, "1.06 Sweetwater", "1.06.02 Invoices"
    ),  # one more folder; draft/invoice depending on avail
    "Three Peaks": Path(
        atlas_mpath, "1.04 Three Peaks", "1.04.01 Energy Invoices"
    ),  # one more folder; draft/invoice depending on avail
}


def atlas_filepatterns(year=None, month=None):
    year, month = validate_YM(year, month)
    lastdayofmonth = (
        pd.Timestamp(year, month, 1) + pd.DateOffset(months=1) - pd.DateOffset(days=1)
    ).day
    return {
        "AZ1": f"*AZ*Solar*{month:02d}-{year}*",
        # 'FL1': f'',
        # 'FL4': f'',
        "GA3": f"*Solar*Generation*{year}*{month:02d}*",
        "GA4": f"*Generation*{month:02d}*{year}*",
        "Grand View East": f"*Solar*{calendar.month_abbr[month].upper()}*{str(year)[-2:]}*",
        "Grand View West": f"*Solar*{calendar.month_abbr[month].upper()}*{str(year)[-2:]}*",
        "Maplewood 1": f"*RealTimeEnergyDetails*{year}-{month:02d}-01*-{lastdayofmonth:02d}*.xlsx",
        "Maplewood 2": f"*RealTimeEnergyDetails*{year}-{month:02d}-01*-{lastdayofmonth:02d}*.xlsx",
        "MS3": f"*MSSL*{str(year)[-2:]}{month:02d}*",
        "Sweetwater": f"*Solar*Invoice*{calendar.month_name[month]}*{year}*",
        "Three Peaks": f"*Power*Invoice*{calendar.month_name[month]}*{year}*",
    }


# SOLAR SITES - VARSITY
varsity_mpath = lambda y, m: Path(oepaths.varsity_metergen, str(y), f"{m}_{str(y)[-2:]}")


def varsity_filepatterns(year=None, month=None):
    year, month = validate_YM(year, month)
    next_ym = pd.Timestamp(year, month, 1) + pd.DateOffset(months=1)
    next_year, next_month = next_ym.year, next_ym.month
    return {
        "Azalea": f"*{month:02d}*{year}*Azalea*Generation*",
        "Imperial Valley": f"*{month:02d}*IVSC*{year}*",
        "Mulberry": f"2839*{calendar.month_abbr[month].upper()}*{str(year)[-2:]}*",
        "Pavant": f"*Invoice*Pavant*Solar*{next_year}{(next_month):02d}*",
        "Richland": f"*Richland*Generation*",
        "Selmer": f"2838*{calendar.month_abbr[month].upper()}*{str(year)[-2:]}*",
        "Somers": f"*C*_hourly.csv",
    }


stlmtui_siteIDs = {
    "Adams East": "ADMEST_6_SOLAR",
    "Alamo": "VICTOR_1_SOLAR2",
    "Camelot": "CAMLOT_2_SOLAR1",
    "Catalina II": "CATLNA_2_SOLAR2",
    "CID": "CORCAN_1_SOLAR1",
    "Columbia II": "CAMLOT_2_SOLAR2",
    "CW-Corcoran": "CORCAN_1_SOLAR2",
    "CW-Goose Lake": "GOOSLK_1_SOLAR1",
    "Kansas": "LEPRFD_1_KANSAS",
    "Kent South": "KNTSTH_6_SOLAR",
    "Maricopa West": "MARCPW_6_SOLAR1",
    "Old River One": "OLDRV1_6_SOLAR",
    "West Antelope": "ACACIA_6_SOLAR",
}


# WIND SITES
isoNE_path = Path(oepaths.commercial, "ISO-NE", "FTP Files", "ISONE")
ercot_path = Path(oepaths.commercial, "ERCOT", "Hedge Settlements", "Novatus Swap Models")


def wind_meterfolders(year=None, month=None):
    year, month = validate_YM(year, month)
    return {
        "Bingham": Path(isoNE_path, "Bingham", "CMP_MeterReads_Monthly"),
        "Hancock": Path(isoNE_path, "Hancock", "Emera Meter Reads"),
        "Oakfield": Path(isoNE_path, "Oakfield", "Emera Meter Reads"),
        "Palouse": Path(oepaths.commercial, "PALOUSE INVOICES", str(year)),
        "Route 66": Path(
            ercot_path, "Rt66", "Tenaska Files", str(year), f"{calendar.month_name[month]} {year}"
        ),
        "South Plains II": Path(
            ercot_path, "SP2", "Tenaska Files", str(year), f"{calendar.month_name[month]} {year}"
        ),
        "Sunflower": Path(
            oepaths.commercial, "Sunflower Invoices", str(year), f"{month:02d}{year}"
        ),
    }


def wind_filepatterns(year=None, month=None):
    year, month = validate_YM(year, month)
    return {
        "Bingham": f"*Bingham*{calendar.month_name[month]}_{year}.xlsx",
        "Hancock": f"*Hancock*{calendar.month_name[month]}_{year}.xlsx",
        "Oakfield": f"*Oakfield*{calendar.month_name[month]}_{year}.xlsx",
        "Palouse": f"*{calendar.month_abbr[month]}*{year}*Gen.xlsx",
        "Route 66": "*Shadow*Real*Time*Energy*Imbalance*Detail*",
        "South Plains II": "*Shadow*Real*Time*Energy*Imbalance*Detail*",
        "Sunflower": "*SUN*.PRN",
    }


atlas_sites = [*atlas_meterfolders]
varsity_sites = [*varsity_filepatterns()] + [*stlmtui_siteIDs]
pi_sites = ["CW-Marin", "Indy I", "Indy II", "Indy III"]

all_solar_sites = atlas_sites + varsity_sites + pi_sites + ["Comanche"]
all_wind_sites = [*wind_meterfolders()]

ALL_METER_SITES = all_solar_sites + all_wind_sites


def get_meter_filepaths(site, year=None, month=None):
    """Get utility meter data filepaths for the given site, year, and month.

    Parameters
    ----------
    site : str
        Name of site.
    year : int, optional*
        Number of year, 4-digit.
    month : int, optional*
        Number of month.

    *note - defaults to current year/month if either value is None,
        or if the given values exceed current date. (e.g. year=2520, month=2)

    Returns
    -------
    list of pathlib.Path objects
        A list of all utility meter data files found for given inputs
        using the paths/file patterns outlined earlier in this file.
        If multiple files are found, returns sorted list (most recent first).
        If no matching files are found, returns empty list.
    """

    year, month = validate_YM(year, month)

    glb_str = None

    # solar, Atlas sites
    if site in atlas_sites:
        glb_str = atlas_filepatterns(year, month).get(site)
        mpath_ = atlas_meterfolders.get(site)
        if site in ["Three Peaks", "Sweetwater"]:
            draft_fp = list(mpath_.glob("*Draft Invoices"))[0]
            final_fp = list(mpath_.glob("*Final Invoices"))[0]
            mfileexists = lambda fp: (len(list(fp.glob(glb_str))) > 0)
            mpath_ = final_fp if mfileexists(final_fp) else draft_fp

    # solar, Varsity sites
    elif site in varsity_sites:
        mpath_ = varsity_mpath(year, month)
        glb_str = "*stlmtui*.xls"
        glb_str = "*stlmtui*"
        if site not in stlmtui_siteIDs:
            glb_str = varsity_filepatterns(year, month).get(site)

    # solar, PI sites (no utility data available)
    elif site in pi_sites:
        mpath_ = site_frpath(site, year, month)
        glb_str = "PIQuery_Meter*.csv"

    # solar, Comanche
    elif site == "Comanche":
        mpath_ = Path(oepaths.commercial, "Comanche Invoices", str(year), f"{month:02d}{year}")
        glb_str = f"{year}-{month:02d}*Comanche*"

    elif site in all_wind_sites:
        mpath_ = wind_meterfolders(year, month).get(site)
        glb_str = wind_filepatterns(year, month).get(site)

    # get matching meter filepath(s)
    meter_fps = list(mpath_.glob(glb_str)) if glb_str else []
    if site == "Comanche":
        meter_fps = [
            fp
            for fp in meter_fps
            if ("Curtailment" not in fp.name) and (fp.suffix in [".xlsx", ".csv"])
        ]

    elif site in stlmtui_siteIDs:
        meter_fps = [fp for fp in meter_fps if (fp.suffix in [".xlsx", ".xls"])]
        unique_stems = list(set(fp.stem.replace(".xls", "") for fp in meter_fps))
        if len(unique_stems) < len(meter_fps):
            xlsx_fp = lambda stem_: Path(mpath_, f"{stem_}.xlsx")
            preferred_fpath = lambda s: (
                xlsx_fp(s) if xlsx_fp(s).exists() else Path(mpath_, f"{s}.xls")
            )
            meter_fps = []  # reset
            for stem_ in unique_stems:
                meter_fps.append(preferred_fpath(stem_))

    if len(meter_fps) > 1:
        sorted_fp_tups = list(
            sorted([(fp.stat().st_ctime, fp) for fp in meter_fps], reverse=True)
        )  # most recently created first
        meter_fps = [tup[1] for tup in sorted_fp_tups]

    return meter_fps


def load_stlmtui_file(fpath):
    """Loads meter data from combined Varsity files. (of type "stlmtui*.xls")

    Parameters
    ----------
    fpath : str or Path object
        Full filepath to varsity meter file. (from "Varsity Generation DataBase" directory)

    Returns
    -------
    pandas DataFrame or None
        If meter file is successfully loaded, returns a dataframe with datetime index
        and hourly generation data (in MWh) for each site found in file. (columns = sitenames)
    """
    ## june 2024 files are suddenly formatted as true .xls docs (not .xml w/ wrong extension)
    try:
        sheets_ = list(pd.read_excel(fpath, sheet_name=None, nrows=0).keys())
        s_names = [s for s in sheets_ if "Meter Data" in s]
        s_name = s_names[0]
        df_ = pd.read_excel(fpath, header=1, sheet_name=s_name)
    except:
        with tempfile.TemporaryDirectory() as temp_dir:
            xml_fpath = Path(temp_dir, f"{fpath.stem}.xml")
            shutil.copy2(src=fpath, dst=xml_fpath)
            with open(xml_fpath) as xml_file:
                data_dict = xmltodict.parse(xml_file.read())
        worksheets = data_dict["Workbook"]["Worksheet"]
        ws_dict = (
            worksheets
            if isinstance(worksheets, dict)
            else [w for w in worksheets if w["@ss:Name"] == "Meter Data"][0]
        )
        data_list = []
        for row_dict in ws_dict["Table"].get("Row")[1:]:  # skipping first row (header w/ no data)
            row_values = [cell_["Data"].get("#text") for cell_ in row_dict["Cell"]]
            data_list.append(row_values)
        uniquecols = []
        for nm in data_list[0]:
            i, col = 1, nm
            while col in uniquecols:
                col = f"{nm}_{i}"
            uniquecols.append(col)
        df_ = pd.DataFrame(data_list[1:], columns=uniquecols)

    df_ = df_[df_["Measurement Type"].eq("GEN")].copy()
    df_ = df_[["Trade Date", "Interval ID", "Value", "Resource ID"]].copy()
    df_.columns = ["Day", "Hour", "MWh", "Site"]
    df_ = df_.replace({"Site": {id_: site for site, id_ in stlmtui_siteIDs.items()}})
    df_["Hour"] = pd.to_numeric(df_["Hour"]).astype(int)
    df_["MWh"] = pd.to_numeric(df_["MWh"])

    ## new - request 11/5/24 - replace NaN values with zeros
    df_["MWh"] = df_["MWh"].fillna(0)

    df_ = df_.loc[df_.Hour.isin(range(1, 25))].reset_index(drop=True)  # for fall DST

    df_["Timestamp"] = pd.to_datetime(
        df_["Day"] + " " + df_["Hour"].sub(1).astype(str).str.zfill(2) + ":00:00",
        format="%m/%d/%Y %H:%M:%S",
    )
    df_ = df_[["Timestamp", "MWh", "Site"]]
    df = pd.pivot_table(df_, index="Timestamp", values="MWh", columns="Site")
    df = df.rename_axis(None, axis=1)

    expected_idx = mdatatransforms.expected_index(df, freq="h")
    # for col in df.columns:
    #     if df[col].last_valid_index() != expected_idx.max():
    #         print(f'!! WARNING !! - data for {col} ends at {df[col].last_valid_index()}')

    # if df.index.max() != expected_idx.max():
    #     print(f'!! WARNING !! - incomplete date range detected in stlmtui file: "{fpath.name}"')
    if df.shape[0] != expected_idx.shape[0]:
        df = df.reindex(expected_idx)

    return df


def load_varsity_stlmtui_sites(year, month, sitelist=None, q=True, return_fpaths=False):
    """Loads meter data from combined Varsity files for specified sitelist. (of type "stlmtui*.xls")

    Parameters
    ----------
    year : int
        Number of year, 4-digit.
    month : int
        Number of month.
    sitelist : list of str, optional
        A list of site names. If None, defaults to all varsity stlmtui sites.
        Function will go through all available stlmtui files until either all specified
        sites are found, or all files have been loaded/checked.
    q : bool, default True
        If False, enables status printouts.

    Returns
    -------
    pandas DataFrame or None
        If meter data is found for specified sites, returns a dataframe with datetime index
        and hourly generation data (in MWh) for each site. (columns = site names)
    """

    # note: function returns dataframe w/ columns names = site names (i.e. not "MWh")
    qprint = lambda msg, end="\n": print(msg, end=end) if (not q) else None
    qprint("\nGROUPED UTILITY METER DATA FILES")

    target_sites = [*stlmtui_siteIDs]
    if sitelist is not None:
        target_sites, invalid_sites = [], []
        for s in sitelist:
            target_sites.append(s) if s in stlmtui_siteIDs else invalid_sites.append(s)
        if invalid_sites:
            qprint(f">> ignoring the following non-stlmtui sites: {invalid_sites}")

    if not target_sites:
        qprint(">> specified stlmtui sites not recognized. Exiting..")
        return None

    meter_fps = get_meter_filepaths(target_sites[0], year, month)
    if not meter_fps:
        qprint(">> no stlmtui files found for the selected year/month. Exiting..")
        return None

    warnings = []
    fpath_dict = {}
    remaining_sites = target_sites.copy()  # init
    found_sites = []  # init
    df_list = []  # init
    for fpath in meter_fps:
        qprint(f'    File: "{fpath.name}"')
        df_ = load_stlmtui_file(fpath)
        sitesofinterest = [s for s in df_.columns if s in remaining_sites]
        if not sitesofinterest:
            qprint("    \u21db no relevant sites found.")
            continue  # go to next fpath
        df = df_[sitesofinterest].copy()
        qprint(f"    \u21db loaded: {sitesofinterest}")

        expected_idx = mdatatransforms.expected_index(df, freq="h")
        for col in df.columns:
            if df[col].last_valid_index() != expected_idx.max():
                alert_msg = f"        !! ALERT !! data for {col} ends {df[col].last_valid_index()}"
                qprint(alert_msg)
                warnings.append(alert_msg)

        df_list.append(df)

        found_sites.extend(sitesofinterest)
        for s in sitesofinterest:
            fpath_dict[s] = str(fpath)
            remaining_sites.remove(s)

        if len(remaining_sites) == 0:
            break

    if not found_sites:
        qprint(f"    >> no specified site(s) found in existing stlmtui files. Exiting..")
        return None

    qprint(
        f"    Complete; found {len(found_sites)} of {len(target_sites)} specified stlmtui sites!"
    )
    missing_sites = [s for s in target_sites if s not in found_sites]
    if missing_sites:
        qprint(f"    >> {missing_sites = }")

    df_stlmtui = pd.concat(df_list, axis=1)
    if return_fpaths:
        if warnings:
            fpath_dict["warnings"] = warnings
        return df_stlmtui, fpath_dict
    return df_stlmtui


SUPPORTED_FALL_DST_SITES = ["Route 66", "South Plains II"]


# function to get utility meter data for sites with individual files (ie not grouped/stlmtui)
def load_site_meter_data(
    site, year=None, month=None, q=True, localized=False, return_df_and_fpath=False
):
    """Load hourly utility meter data for the given site, year, and month.

    Parameters
    ----------
    site : str
        Name of *site.
    year : int, **optional
        Number of year, 4-digit.
    month : int, **optional
        Number of month.
    q : bool, default True
        If False, enables status printouts.

    *note - there are certain varsity sites that provide data in combined files,
        and will not be able to load. (use "load_varsity_stlmtui_sites" function)

    **note - defaults to current year/month if either value is None,
        or if the given values exceed current date. (e.g. year=2420, month=2)

    Returns
    -------
    pandas DataFrame or None
        If meter file is found and successfully loaded, returns a dataframe with
        the following columns/dtypes: ['Day' (datetime), 'Hour' (int), 'MWh' (float)]
        where the "Hour" column represents "Hour Ending" for the interval (range 1 to 24).
        If there is no file or an error loading the file, returns None.
    """

    qprint = lambda msg, end="\n": print(msg, end=end) if not q else None

    # get meter filepath(s)
    mfps_ = get_meter_filepaths(site, year, month)  # returns sorted list (most recent first)
    if not mfps_:
        qprint("no meter file found.")
        return None
    mfp = max((fp.stat().st_ctime, fp) for fp in mfps_)[1]
    func_name = "load_" + site.lower().replace(" ", "_").replace("-", "_")
    load_df_function = getattr(mdatatransforms, func_name)
    qprint(f'Loading file: "{mfp.name}"', end=" ... ")
    args = [mfp]
    if localized and (site in SUPPORTED_FALL_DST_SITES):
        args.append(True)
    try:
        df = load_df_function(*args)
        qprint("success!")
    except:
        df = None
        qprint("ERROR!")

    if return_df_and_fpath:
        return df, mfp

    return df


# functions for printing out PI vs. Utility comparison
pstr = lambda pct: f"  ({pct:.2f}%)"
nstr = lambda n, r=11: f"{n:.2f}".rjust(r)
totalsline = lambda n1, n2, n3, n4: nstr(n1) + nstr(n2) + nstr(n3, 10) + pstr(n4)

pi_meter_attPaths = ref.solar_meter_attPaths | ref.wind_meter_attPaths


# function to validate a year/month tuple & list of y/m tuples
def valid_ym(item):
    if not isinstance(item, tuple):
        return False
    isvalid = (
        all(isinstance(n, int) for n in item)
        and (item[0] in range(2023, 2123))
        and (item[1] in range(1, 13))
    )
    return isvalid


def valid_ym_list(yearmonth_list):
    if not isinstance(yearmonth_list, list):
        return False
    isvalid = all(valid_ym(item) for item in yearmonth_list)
    return isvalid


valid_year_month_args = lambda y, m, ymlist: any([valid_ym_list(ymlist), valid_ym((y, m))])


# break-out function for comparison of utility & pi meter data
## if df_util is provided, skip call to "load_meter_data"
def compare_utility_and_pi_meter(
    year=None, month=None, yearmonth_list=None, sitelist=None, df_util=None, q=True
):
    """Compares utility meter data with internal PI data for given site(s).

    Parameters
    ----------
    year : int, default None
        Number of year, 4-digit.
    month : int, default None
        Number of month.
    yearmonth_list : list of tuples of int, default None
        A list of tuples containing the year and month - e.g. [(2024, 3), (2024, 4)]
        When provided, function will load/compare meter data for multiple year/months.
        Use this parameter instead of calling function externally in a loop; this will
        result in faster execution time by eliminating multiple calls to the function
        that reads from the meter generation historian file "load_meter_historian"
    sitelist : list of str, default None
        A list of site names. When provided, this parameter will load utility meter
        data from the meter generation historian file for the given sites. PI data
        will then either be loaded from the existing "PIQuery_Meter" file in the
        flashreport folder, or queried in real-time from the PI server.
        If sitelist=None, df_util must be provided (see below).
    df_util : pandas DataFrame
        A dataframe containing utility meter data. (columns = site names)
        This parameter is intended to be used in function "load_meter_data" to
        serve as a check/indicator for data quality issues when loading new
        utility meter data from files.
    q : bool, default True
        If False, enables status printouts.

    Returns
    -------
    pandas DataFrame or None
        If utility meter data is found for given sites, returns a dataframe with
        the following columns: ['Site', 'Year', 'Month', 'PI Meter', 'Utility Meter',
        'Delta (MWh)', 'Delta (%)']. If no utility meter data is found or no valid
        sites are provided, function returns None.

    Example
    -------
    >>> df_compare = umd.compare_utility_and_pi_meter(2024, 4, sitelist=['Camelot', 'GA4', 'Kansas'])
    Loading utility meter historian.. done!
    Comparing with PI data:
        SITE                   PI METER    UTILITY     Δ-MWh
        Camelot                11139.80   11814.42    674.62  (6.06%)
        Kansas                  3587.14    3616.34     29.20  (0.81%)
        GA4                    45600.90   45704.95    104.05  (0.23%)
    """
    if not valid_year_month_args(year, month, yearmonth_list):
        print(
            "Invalid year/month args. Please provide valid year/month, or valid yearmonth_list.\nExiting.."
        )
        return

    if not valid_ym_list(yearmonth_list):
        yearmonth_list = [(year, month)]

    if (df_util is None) and (sitelist is None):
        print("No sitelist specified.\nExiting..")
        return

    ## function for status printouts (if q=False)
    qprint = lambda msg, end="\n": None if q else print(msg, end=end)

    ## check if function is being called from "load_meter_data" function
    from_ext_func = isinstance(df_util, pd.DataFrame)
    if from_ext_func:
        sitelist = list(df_util.columns)

    ## check sites in sitelist
    sitelist = [s for s in sitelist if s in ALL_METER_SITES]
    if not sitelist:
        print("No valid sites found in specified sitelist.\nExiting..")
        return

    ## if normal call to function (ie. not from external func), load meter generation historian
    if not from_ext_func:
        qprint("Loading utility meter historian..", end=" ")
        df_hist = load_meter_historian()
        qprint("done!")
        util_source = "historian"

    ## begin collection of meter data
    output_data = []

    qprint("\nComparing with PI data:")

    a1, a2, a3, a4 = ["SITE", "PI METER", "UTILITY", "\u0394-MWh"]  # header
    uline_ = lambda str_: "-" * len(str_)
    for year, month in list(sorted(yearmonth_list, reverse=True)):
        qprint(f"  [({year}, {month:02d})]")
        qprint(f"    {a1: <20}{a2: >11}{a3: >11}{a4: >10}")  # header
        qprint(
            f"    {uline_(a1): <20}{uline_(a2): >11}{uline_(a3): >11}{uline_(a4): >10}"
        )  # header underline
        if not from_ext_func:
            t0 = pd.Timestamp(year=year, month=month, day=1)
            t1 = t0 + pd.DateOffset(months=1)
            df_util = df_hist[(df_hist.index > t0) & (df_hist.index < t1)].copy()
            df_util = df_util.dropna(axis=1, how="all").copy()

        for site in sitelist:
            qprint(f"    {site.ljust(20)}", end="")
            if site not in df_util.columns:
                qprint(f'{"--":>11}{"n/a":>11}{"--":>10}')  # util n/a
                continue  # go to next site

            site_frpath = Path(oepaths.frpath(year, month, ext="Solar"), site)
            pi_meter_fpaths = list(site_frpath.glob("PIQuery*Meter*.csv"))
            ## load PI data from query file if available, otherwise run query
            if pi_meter_fpaths:
                mfpath = max((fp.stat().st_ctime, fp) for fp in pi_meter_fpaths)[1]
                dfpm = pd.read_csv(mfpath, index_col=0)
                pi_source = str(mfpath).split("Operations")[-1]
            else:
                mtr_attPath = pi_meter_attPaths.get(site)
                if not mtr_attPath:
                    qprint(f'{"n/a":>11}{"--":>11}{"--":>10}')  # pi n/a
                    continue  # go to next site
                qprint("   (querying PI meter data...)", end="\r")
                start = pd.Timestamp(year=year, month=month, day=1)
                end = start + pd.DateOffset(months=1)
                dfpm = pt.query_pi_data(
                    site, start, end, freq="1m", attPath_list=[mtr_attPath], q=True
                )
                qprint(f"    {site.ljust(20)}", end="")
                pi_source = "query"

            # calculate totals & delta
            pi_total = dfpm[dfpm.columns[0]].sum() / 60  # minute-level data
            if pi_total <= 0:
                qprint(f'{"n/a":>11}{"--":>11}{"--":>10}')  # pi n/a
                continue  # go to next site

            util_total = df_util[site].sum()  # hourly data
            delta = util_total - pi_total
            pct = (delta / pi_total) * 100

            original_util_fp = "n/a"
            orig_util_fps = get_meter_filepaths(site, year=year, month=month)
            if orig_util_fps:
                original_util_fp = str(orig_util_fps[0]).split("Commercial")[-1]
            u_src = "historian" if not from_ext_func else "file"
            util_source = f'{u_src}; original file: "..{original_util_fp}"'

            t_line = totalsline(pi_total, util_total, delta, pct)
            if pi_source == "query":
                t_line = t_line + "  *note: real-time query data"
            qprint(t_line)

            output_data.append(
                [site, year, month, pi_total, util_total, delta, pct, pi_source, util_source]
            )

        qprint("")

    dfcols = [
        "Site",
        "Year",
        "Month",
        "PI Meter",
        "Utility Meter",
        "Delta (MWh)",
        "Delta (%)",
        "pi_source",
        "util_source",
    ]
    df_compare = pd.DataFrame(output_data, columns=dfcols)
    return df_compare


# function to load utility meter data for any site(s)
def load_meter_data(year, month, sitelist=None, q=True, keep_fall_dst=False, return_fpaths=False):
    """Load hourly utility meter data for the given year, month, and site(s).

    Parameters
    ----------
    year : int
        Number of year. (4-digit)
    month : int
        Number of month.
    sitelist : list of str, default all sites.
        A list of site names.
    q : bool, default True
        If False, enables status printouts. (includes Utility/PI comparison)

    Returns
    -------
    pandas DataFrame
        A time series of hourly meter generation data for the requested site(s) with
        DatetimeIndex and a column (dtype: float, units: MWh) for each site in list.
        If no data is found for given site(s), returns empty DataFrame.

    Example
    --------
    >>> sitelist = ['Adams East', 'AZ1', 'Bingham', 'Camelot', 'CID', 'GA4', 'Kansas', 'Richland']
    >>> df = load_meter_data(year=2024, month=4, sitelist=sitelist, q=False)
    """
    qprint = lambda msg, end="\n": print(msg, end=end) if not q else None

    qprint(f"\nBEGIN METER DATA COLLECTION FOR {calendar.month_name[month].upper()} {year}")
    if sitelist is not None:
        target_sites = [s for s in sitelist if s in ALL_METER_SITES]
        if not target_sites:
            qprint("(note: specified sitelist is invalid; defaulting to all sites)")
            target_sites = ALL_METER_SITES
    else:
        qprint("(note: sitelist not specified; defaulting to all sites)")
        target_sites = ALL_METER_SITES

    qprint("\nSearching for sites:")
    if not q:
        print_sitelist_block(target_sites)

    # only allows to keep extra timestamp if all sites in list are wind sites (functionality not built out for solar)
    if keep_fall_dst:
        if month != 11:
            keep_fall_dst = False
        if not all(s in SUPPORTED_FALL_DST_SITES for s in target_sites):
            keep_fall_dst = False

    # split sites into stlmtui & non-stlmtui
    main_sites, stlmtui_sites = [], []
    for s in target_sites:
        main_sites.append(s) if (s not in stlmtui_siteIDs) else stlmtui_sites.append(s)

    df_list = []
    fpath_dict = {}

    if main_sites:
        qprint("\nINDIVIDUAL UTILITY METER DATA FILES")
        for site in main_sites:
            qprint(f"    {site.ljust(22)}", end="")
            if not keep_fall_dst:
                output = load_site_meter_data(site, year, month, q=q, return_df_and_fpath=True)
                if output is None:
                    continue
                dfm_, mfp = output
                dfm_.index = pd.to_datetime(
                    dfm_.Day.astype(str)
                    + " "
                    + dfm_.Hour.sub(1).astype(str).str.zfill(2)
                    + ":00:00",
                    format="%Y-%m-%d %H:%M:%S",
                )
                dfm = dfm_[["MWh"]].rename(columns={"MWh": site}).copy()
            else:
                output = load_site_meter_data(
                    site, year, month, q=q, localized=True, return_df_and_fpath=True
                )
                if output is None:
                    continue
                dfm, mfp = output
                dfm.columns = [site]  # df has one column & datetimeindex

            df_list.append(dfm)
            fpath_dict[site] = str(mfp)

    if stlmtui_sites:
        stlmtui_output = load_varsity_stlmtui_sites(
            year=year,
            month=month,
            sitelist=stlmtui_sites,
            q=q,
            return_fpaths=return_fpaths,
        )
        if stlmtui_output is not None:
            df_stlmtui = stlmtui_output[0] if return_fpaths else stlmtui_output
            df_list.append(df_stlmtui)
            if return_fpaths:
                fpath_dict.update(stlmtui_output[1])

    qprint("\nFinished loading files", end="; ")

    if len(df_list) > 0:
        df_meter = pd.concat(df_list, axis=1)
        df_meter = df_meter[list(sorted(df_meter.columns))]
        foundsites = list(df_meter.columns)
        qprint(f"found {len(foundsites)} of {len(target_sites)} sites:")
        if not q:
            print_sitelist_block(foundsites)

            # call function to compare df_meter with PI data (from files, or query)
            if not keep_fall_dst:
                df_compare = compare_utility_and_pi_meter(year, month, df_util=df_meter.copy(), q=q)
            else:
                print("Skipping comparison b/c of DST")

    else:
        df_meter = pd.DataFrame()
        qprint("no data found.")

    qprint("\nEND METER DATA COLLECTION\n")
    if return_fpaths:
        return df_meter, fpath_dict
    return df_meter


histpath = Path(oepaths.meter_generation_historian)


# function to collect utility meter data & update historian Excel file
def compile_utility_meter_data(
    year, month, sitelist=None, overwrite=False, local=False, q=True, output=False
):
    """Loads utility meter data and updates historian file for the given year, month, and site(s).

    Parameters
    ----------
    year : int
        Number of year. (4-digit)
    month : int
        Number of month.
    sitelist : list of str, default all sites
        A list of site names.
    overwrite : bool, default False
        When False, historian file will only be updated for sites without existing data
        for the specified month. If True, sites with existing data will be overwritten.
    local : bool, default False
        If True, a copy of the historian file will be saved to the local Downloads folder
        and the original historian file will NOT be updated.
    q : bool, default True
        If False, enables status printouts.

    Returns
    -------
    None
        This function loads the existing meter generation historian file from path:
        "..Commercial\\Standardized Meter Generation\\Meter_Generation_Historian.xlsm"
        and then either updates/re-saves the file, or saves a new copy (if local=True).
        If overwrite=False and meter data is found for sites with existing data in the
        historian, those sites will be excluded from the update.
        If no meter data is found for specified sites, function exits (and returns None).

    Example
    --------
    >>> sitelist = ['Adams East', 'CID', 'GA4', 'Kansas', 'Richland']
    >>> compile_utility_meter_data(year=2024, month=4, sitelist=sitelist, q=False)
    """
    qprint = lambda msg, end="\n": print(msg, end=end) if not q else None

    if sitelist is None:
        qprint("Note: no sitelist specified - defaulting to all solar/wind sites.\n")
        search_sites = ALL_METER_SITES
    else:
        search_sites = [s for s in sitelist if s in ALL_METER_SITES]
        if not search_sites:
            qprint("No valid sites found in specified sitelist.\nExiting..")
            if output:
                return "SCRIPT FAILED - SPECIFIED SITELIST IS INVALID"
            return

    existing_sites = []
    if not overwrite:
        qprint("Checking historian file for existing data.. (overwrite=False)")
        df_hist = load_meter_historian(year=year, month=month)
        df_hist = df_hist.dropna(axis=1, how="all")  # drop cols with no data
        existing_sites = list(df_hist.columns)
        if existing_sites:
            if not search_sites:
                qprint(">> found existing data for all specified sites!\n\nExiting..\n")
                if output:
                    return "ALL SITES FOUND - EXITING"
                return
            qprint(
                f">> found sites with existing meter data.\n(overwrite=False) excluding: {existing_sites}\n"
            )
            search_sites = [s for s in search_sites if s not in existing_sites]
        else:
            qprint(">> no site data found for selected year/month.\n")

    # condition for updating the historian WITH EXTRA FALL DST TIMESTAMP (tstamp must already exist in historian)
    FALL_DST_CONDITION = (month == 11) and all(s in SUPPORTED_FALL_DST_SITES for s in search_sites)

    # load utility meter data
    dfu, fpath_dict = load_meter_data(
        year=year,
        month=month,
        sitelist=search_sites,
        q=q,
        keep_fall_dst=FALL_DST_CONDITION,
        return_fpaths=True,
    )
    if dfu.empty:
        qprint("!! no meter data found for the specified inputs !!\n\nExiting..\n")
        if output:
            return "NO NEW DATA FOUND - EXITING"
        return

    # format dataframe for transfer to meter gen historian file
    sitecols = list(dfu.columns)
    df_ALL = dfu.copy()
    df_ALL["Day"] = pd.to_datetime(df_ALL.index.date)
    df_ALL["Hour Ending"] = df_ALL.index.hour + 1
    df_ALL["Month"] = df_ALL.index.month
    df_ALL["Year"] = df_ALL.index.year
    newcols = ["Day", "Hour Ending", "Month", "Year"] + sitecols
    df_ALL = df_ALL[newcols].reset_index(drop=True)

    # get expected date range for comparison
    expected_date_range = date_range(year, month, freq="h")
    dt_start = dt.datetime(year=year, month=month, day=1)
    dt_end = expected_date_range[-1].to_pydatetime()

    if not FALL_DST_CONDITION:
        # compare with actual min/max from df
        mfile_start = dfu.index.min()
        mfile_end = dfu.index.max()
        if (dt_start != mfile_start) or (dt_end != mfile_end):
            qprint(f"!! ERROR !! -- timestamp mismatch -- check meter data files\nExiting..")
            if output:
                return "SCRIPT FAILED - TIMESTAMP MISMATCH"
            return

    # load existing meter gen historian
    qprint("Loading meter generation historian file for read/write", end="... ")

    wb = openpyxl.load_workbook(histpath, keep_vba=True)
    ws = wb["Hourly Gen by Project"]
    qprint("done!")

    # function to get datetime value from "Day" and "Hour Ending" cells in excel file
    get_datetime = lambda cellA, cellB: (cellA.value + dt.timedelta(hours=cellB.value - 1))

    # get datetime value from last line of excel file
    last_row = ws.max_row
    while not isinstance(ws[f"A{last_row}"].value, dt.datetime):
        last_row -= 1
    lastA, lastB = ws[f"A{last_row}"], ws[f"B{last_row}"]

    last_entry = get_datetime(lastA, lastB)

    # get list of columns excluding headers cols (ie first row)
    ws_colA_cells = [c for c in ws["A"] if isinstance(c.value, dt.datetime)]
    ws_colB_cells = [c for c in ws["B"] if str(c.value).isnumeric()]

    # get datetime range for entire historian file
    historian_datetimes = [
        get_datetime(cellA, cellB) for cellA, cellB in zip(ws_colA_cells, ws_colB_cells)
    ]
    hist_date_range = pd.to_datetime(historian_datetimes)

    # check existing date range in historian
    all_dates_exist = all(d in hist_date_range for d in expected_date_range)
    some_dates_exist = any(d in hist_date_range for d in expected_date_range)
    missing_dates = [d for d in expected_date_range if d not in hist_date_range]
    if all_dates_exist:
        msg1_, msg2_ = "all", "new data will be added to existing rows"
    elif some_dates_exist:
        msg1_, msg2_ = "some", f"{len(missing_dates)} missing rows will be added to file"
    else:
        msg1_, msg2_ = "no", "new data will be appended to file"
    qprint(f">> {msg1_} timestamps found in file for {calendar.month_name[month]}-{year}; {msg2_}.")

    # check for duplicates
    related_hist_dates = [d for d in hist_date_range if d in expected_date_range]
    unique_hist_dates = list(set(related_hist_dates))
    n_duplicates = len(related_hist_dates) - len(unique_hist_dates)
    if (n_duplicates == 1) and (month == 11):
        qprint("\nFall DST duplicate timestamp found in historian file!")
        extra_tstamp = [t for t in related_hist_dates if related_hist_dates.count(t) > 1][0]
        if not FALL_DST_CONDITION:
            qprint(
                "Utility data does not contain the extra DST timestamp; inserting blank row before transferring to historian."
            )
            dup_idx = related_hist_dates.index(extra_tstamp)
            df_ALL = pd.concat(
                [
                    df_ALL.iloc[:dup_idx, :],
                    df_ALL.iloc[dup_idx : dup_idx + 1, :],
                    df_ALL.iloc[dup_idx:, :],
                ],
                axis=0,
                ignore_index=True,
            )
        else:
            qprint("All specified sites have extra timestamp; adding to file.")

    elif n_duplicates > 0:
        qprint(f"\n!! WARNING !! - ({n_duplicates}) duplicate timestamps found in file")
        if not overwrite:
            qprint(
                f'{" "*16}>> set overwrite=True to fix index (note: will re-pull existing data for range)'
            )
            qprint("Exiting..")
            if output:
                return "SCRIPT FAILED - DUPLICATE TIMESTAMP ISSUE"
            return
        qprint(f">> overwrite=True; removing duplicates & overwriting data for selected sites")

    # update historian with new data
    # case 1 --- all new timestamps (i.e. no existing data for selected year/month)
    # if no_dates_exist:
    if last_entry < dt_start:
        rStart = last_row + 1
        rEnd = last_row + expected_date_range.shape[0]
        # qprint(f'\n{rStart = } (new row)\n{rEnd = } (new row)')
    else:
        # find starting row (if exists)
        if expected_date_range[0] in hist_date_range:
            rStart = [
                cA.row
                for cA, cB in zip(ws_colA_cells, ws_colB_cells)
                if (get_datetime(cA, cB) == dt_start)
            ][0]
        else:  # otherwise, find last date before expected start
            rStart = [
                cA.row
                for cA, cB in zip(ws_colA_cells, ws_colB_cells)
                if (get_datetime(cA, cB) < dt_start)
            ][-1]

        # find end row (if exists)
        next_month_exists_in_hist = (
            expected_date_range[-1] + pd.Timedelta(hours=1)
        ) in hist_date_range
        if expected_date_range[-1] in hist_date_range:
            rEnd = [
                cA.row
                for cA, cB in zip(ws_colA_cells, ws_colB_cells)
                if (get_datetime(cA, cB) == dt_end)
            ][0]
        elif next_month_exists_in_hist:
            rEnd = [
                cA.row
                for cA, cB in zip(ws_colA_cells, ws_colB_cells)
                if (get_datetime(cA, cB) > dt_end)
            ][0]
        else:
            rEnd = rStart + expected_date_range.shape[0] - 1

        # determine whether rows need to be inserted
        if some_dates_exist and missing_dates and next_month_exists_in_hist:
            n_newrows = len(missing_dates)
            if not overwrite:
                qprint("\n>> set overwrite=True to add missing timestamps/data.\n\nExiting..\n")
                if output:
                    return "SCRIPT FAILED - TIMESTAMP MISMATCH"
                return
            # simplified approach: add all missing rows to end of range & re-write (overwrite) data
            ws.insert_rows(rEnd, n_newrows)
            qprint(f"added {n_newrows} rows to end of date range in historian!")
        # determine whether rows need to be removed (ie duplicates)
        # elif (n_duplicates > 0) and (not FALL_DST_CONDITION) and overwrite:
        #     ws.delete_rows(rEnd, n_duplicates)
        #     qprint(f'removed {n_duplicates} rows from date range in historian!')
        # qprint(f'\n{rStart = }, {rEnd = }')

    # check length of range from start to end row
    n_newrows = rEnd + 1 - rStart  # in historian
    n_datarows = df_ALL.shape[0]  # from file
    if n_newrows != n_datarows:
        qprint("DST-RELATED ERROR!!")
        if output:
            return "SCRIPT FAILED - DST-RELATED ERROR"
        return

    # n_expectedrows = expected_date_range.shape[0]
    # if n_newrows != n_expectedrows:
    #     qprint(f'WARNING! {n_newrows = } ({n_expectedrows = })')

    # find target columns in historian for selected sites with meter data
    hist_cols = [c.value for c in ws[1]]  # cell values from row 1 in ws
    foundsites = list(dfu.columns)
    targetcols = {site: (hist_cols.index(site) + 1) for site in foundsites}

    # variables for checking vals in date/time cols (in loop)
    dtcols_ = ["Day", "Hour Ending", "Month", "Year"]
    datecols = {c: hist_cols.index(c) + 1 for c in dtcols_}

    checkcols = datecols | targetcols
    rowrange = range(rStart, rEnd + 1)

    qprint("\nWriting meter data to Excel worksheet", end="... ")
    for i, row in enumerate(rowrange):
        for col, idx in checkcols.items():
            # get corresponding cell in historian file
            cell = ws.cell(row, idx)
            # write/update value if cell is empty, or if overwrite=True
            if any([(cell.value is None), overwrite]):
                cell.value = df_ALL.at[i, col]
                if col == "Day":
                    cell.number_format = "m/d/yyyy"
    qprint("done!")

    qprint(f">> added new meter data to historian for {len(foundsites)} sites:")
    for fsite in list(sorted(foundsites)):
        qprint(f"    ++ {fsite}")

    # save & close updated file
    if not local:
        savepath = histpath
        idx_ = Path(savepath).parts.index("Commercial")
        disppath = os.sep.join(Path(savepath).parts[idx_:])
    else:
        sfolder_ = Path(Path.home(), "Downloads")
        stem_ = histpath.stem
        n_, fname_ = 1, histpath.name
        while Path(sfolder_, fname_).exists():
            fname_ = f"{histpath.stem} ({n_}){histpath.suffix}"
            n_ += 1
        savepath = Path(sfolder_, fname_)
        disppath = str(savepath)

    qprint("\nSaving file", end="... ")
    wb.save(savepath)
    wb.close()
    qprint(f'done!\n>> path: "..{disppath}"\n\nEND.')

    if output:
        valid_keys = foundsites + ["warnings"]
        return {k: fp for k, fp in fpath_dict.items() if k in valid_keys}
    return
